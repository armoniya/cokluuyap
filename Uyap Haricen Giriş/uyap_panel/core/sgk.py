"""
uyap_panel.core.sgk — SGK toplu sorgu servisi (arayüzden bağımsız).

Bağımsız 'SGK Sorgu' aracının (sgk_sorgu_gui.py) motoru BURADA tek yerde toplanır;
ttkbootstrap GUI ve Django ön yüzleri yalnızca bu katmanı kullanır.

İki parça:
  • SorguMotoru — yerel ofis proxy'sine (127.0.0.1:8800) .ajx istekleri gönderir;
                  ofis (uyap_app.py) bunları canlı e-imza UYAP oturumuyla iletir.
                  (Takip akışındaki iş-kuyruğundan farklı: bu doğrudan proxy çağrısıdır.)
  • SgkBatch     — Excel tabanlı toplu sorgu döngüsü. Kendi iş parçacığında çalışır,
                  iç durum tutar (loglar, satırlar, ilerleme). Her iki ön yüz de
                  aynı snapshot() ile durumu okuyup arayüzü tazeler (poll deseni).

Veri işleme (özetleme, SSK olumlu/olumsuz değerlendirme, tam döküm) orijinal
araçtan birebir taşınmıştır.
"""

import os
import re
import json
import time
import threading
import urllib.request
import urllib.error

from .config import PORT

BASE_URL = "https://avukat.uyap.gov.tr"

# Birim türleri (logdan): İcra -> 2, İcra Dairesi -> 1101
BIRIM_TURU2 = "1101"
BIRIM_TURU3 = "2"

COMMON_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "content-type": "application/json",
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "expires": "0",
    "referer": f"{BASE_URL}/dosya-sorgulama",
}

# Sorgu tanımları: (kolon_basligi, endpoint, ozetleyici_anahtar)
SORGULAR = [
    ("Kamu Çalışan",   "borclu_bilgileri_goruntule_sgk_kamuCalisaniBilgileri.ajx",   "kamuCalisani"),
    ("Kamu Emekli",    "borclu_bilgileri_goruntule_sgk_kamuEmekliBilgileri.ajx",     "kamuEmekli"),
    ("SSK Çalışan",    "borclu_bilgileri_goruntule_sgk_sskCalisaniBilgileri.ajx",    "sskCalisani"),
    ("SSK Emekli",     "borclu_bilgileri_goruntule_sgk_sskEmekliBilgileri.ajx",      "sskEmekli"),
    ("Bağkur Çalışan", "borclu_bilgileri_goruntule_sgk_bagkurCalisaniBilgileri.ajx", "bagkurCalisani"),
    ("Bağkur Emekli",  "borclu_bilgileri_goruntule_sgk_bagkurEmekliBilgileri.ajx",   "bagkurEmekli"),
    ("SSK İş Yeri",    "borclu_bilgileri_goruntule_sgk_isYeriBilgileri.ajx",         "isYeri"),
]

# Tablo başlıkları: 4 sabit + sorgu sütunları
KOLONLAR = ["No", "Ad Soyad", "Birim", "Dosya No"] + [s[0] for s in SORGULAR]
ANAHTARLAR = [s[2] for s in SORGULAR]

SONUC_BASLANGIC_KOL = 5    # E (1-tabanlı)
HATA_ON = "[HATA]"
SIRKET_NOT = "ŞİRKET - sorgulanmadı"
SIRKET_ANAHTARLARI = {"STI", "SIRKETI", "SIRKET", "LTD"}

SORGU_ARASI_BEKLEME = 1.5
SATIR_ARASI_BEKLEME = 0.2
MOLA_HATA_ESIGI = 3
MOLA_SURESI = 60
OTURUM_KONTROL_ESIGI = 6

YAPILANLAR_EKI = "_yapilanlar"
OZET_EKI = "_yapilanlar_ozet"

NOISE_KEYS = {"isNew", "isMock", "hasError", "hasData", "metadata",
              "yeniBirimEkle", "orgKoduDegisti", "isTumunuKopyala", "testMi"}


class OturumHatasi(Exception):
    """UYAP girişi/oturumu düşmüş (login sayfası ya da yetkisiz yanıt döndü)."""


# --------------------------------------------------------------------------- #
#  Türkçe normalizasyon ve veri özetleme/dökme yardımcıları (orijinalden)
# --------------------------------------------------------------------------- #
def tr_normalize(s):
    if s is None:
        return ""
    s = str(s)
    cevrim = str.maketrans("ıİiİğĞüÜşŞöÖçÇ", "IIIIGGUUSSOOCC")
    s = s.translate(cevrim)
    s = s.upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _temizle(v):
    if v is None:
        return ""
    return str(v).strip()


def _ilk_hata(dvo):
    for k in ("hataMesaji", "anaHataMesaji", "sonucMesaj", "mesaj", "islemSonucAciklama"):
        if isinstance(dvo, dict) and dvo.get(k):
            return _temizle(dvo[k])
    return ""


def _g(d, *yol):
    for k in yol:
        if isinstance(d, dict):
            d = d.get(k)
        elif isinstance(d, list) and isinstance(k, int) and -len(d) <= k < len(d):
            d = d[k]
        else:
            return None
    return d


def _parcala(*ciftler):
    out = []
    for et, dg in ciftler:
        dg = _temizle(dg)
        if dg:
            out.append(f"{et}: {dg}" if et else dg)
    return " | ".join(out)


def sirket_mi(isim):
    """İsim tüzel kişi (ŞTİ / Şirketi / Ltd) mi? -> SGK sorgusu yapılmaz."""
    if not isim:
        return False
    kelimeler = set(tr_normalize(isim).split())
    return bool(kelimeler & SIRKET_ANAHTARLARI)


def degerlendir_ssk_calisani(dvo):
    """Döner: (durum, ozet, veri_var).  durum = 'Olumlu' | 'Olumsuz'."""
    if not isinstance(dvo, dict):
        s = _temizle(dvo)
        return ("Olumlu" if s else "Olumsuz"), s, bool(s)
    try:
        sskBilgi = dvo.get("sskBilgi") or {}
        tescil_list = sskBilgi.get("tescilKaydi4AList") or []
        isyeri_sonuc = _g(dvo, "isYeriBilgileri", "data", "sonuc") or {}
        ozet_array = isyeri_sonuc.get("isyeriTescil4AOzetArray") or []
        sig = _g(dvo, "sskBilgiDetay", "sgkIsyeriDVO", "sigortaliBilgi") or {}

        b = isyeri_sonuc.get("basarili")
        guncel_hizmet = (b is True) or (str(b).strip().lower() == "true")

        oz = ozet_array[0] if ozet_array else {}
        unvan = _temizle(sig.get("isyeriUnvan")) or _temizle(oz.get("unvan"))
        vergi = _temizle(oz.get("teVergiNo"))
        donem = _temizle(sig.get("donemYilAy"))
        if not donem and oz.get("sonDonemYil"):
            donem = f"{_temizle(oz.get('sonDonemYil'))}/{_temizle(oz.get('sonDonemAy'))}"
        gun = oz.get("sonDonemHizmetGun")
        sicil = _temizle(sig.get("sicilNo")) or _temizle(_g(tescil_list, 0, "sicilNo"))
        giris = (_temizle(sig.get("ilkIseGirisTarihi"))
                 or _temizle(_g(tescil_list, 0, "kimlikBilgileri", "iseGirisTarihi")))
        adres = _temizle(sig.get("isyeriAdresi"))
        cikis = _temizle(sig.get("istenCikisTarihi"))

        gun_s = ""
        if gun not in (None, ""):
            gun_s = f"son dönem {gun} gün"
            try:
                if int(gun) < 30:
                    gun_s += " (ay içinde çıkış olabilir)"
            except (TypeError, ValueError):
                pass

        if guncel_hizmet and not cikis:
            ozet = _parcala((None, "Çalışıyor"), ("İşyeri", unvan), ("VKN", vergi),
                            ("Son dönem", donem), (None, gun_s), ("Adres", adres),
                            ("Sicil", sicil), ("İşe giriş", giris))
            return "Olumlu", ozet, True

        mesaj = (_temizle(isyeri_sonuc.get("mesaj"))
                 or _temizle(sskBilgi.get("sonucMesaji"))
                 or _temizle(sskBilgi.get("anaHataMesaji")))
        if cikis:
            ozet = _parcala(("İşten çıkış", cikis), ("Son işyeri", unvan), ("VKN", vergi),
                            ("Son dönem", donem), (None, gun_s), ("Adres", adres),
                            ("Sicil", sicil), ("İşe giriş", giris))
            return "Olumsuz", ozet, True
        if unvan or donem or sicil:
            son = _parcala(("Son işyeri", unvan), ("Son dönem", donem),
                           ("Adres", adres), ("Sicil", sicil), ("İşe giriş", giris))
            ozet = mesaj or "Güncel 4a hizmeti bulunamadı"
            if son:
                ozet = f"{ozet} | {son}"
            return "Olumsuz", ozet, True
        if tescil_list:
            kb = _g(tescil_list, 0, "kimlikBilgileri") or {}
            ek = _parcala(("Eski sicil", _temizle(kb.get("sicilNo"))),
                          ("İlk işe giriş", _temizle(kb.get("iseGirisTarihi"))))
            ozet = mesaj or "Güncel 4a hizmeti bulunamadı"
            if ek:
                ozet = f"{ozet} ({ek})"
            return "Olumsuz", ozet, True
        return "Olumsuz", (mesaj or "Kayıt bulunamadı"), False
    except Exception as e:
        return "Olumsuz", f"[değerlendirme hatası: {e}]", False


def ozetle(anahtar, dvo):
    try:
        if not isinstance(dvo, dict):
            return _temizle(dvo)

        if anahtar == "kamuCalisani":
            if dvo.get("durumAciklamasi4C") or dvo.get("kurum"):
                parcalar = [_temizle(dvo.get("durumAciklamasi4C")),
                            _temizle(dvo.get("kurum")),
                            _temizle(dvo.get("unvani"))]
                yer = "/".join(p for p in (_temizle(dvo.get("kurumIl")),
                                           _temizle(dvo.get("kurumIlce"))) if p)
                if yer:
                    parcalar.append(yer)
                if dvo.get("iseBaslamaTarihi"):
                    parcalar.append("İşe baş: " + _temizle(dvo.get("iseBaslamaTarihi")))
                return " | ".join(p for p in parcalar if p)
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "kamuEmekli":
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "sskCalisani":
            parcalar = []
            ssk = dvo.get("sskBilgi") or {}
            liste = ssk.get("tescilKaydi4AList") or []
            if liste:
                kb = (liste[0] or {}).get("kimlikBilgileri") or {}
                sicil = _temizle(kb.get("sicilNo"))
                giris = _temizle(kb.get("iseGirisTarihi"))
                bilgi = "Tescil var"
                if sicil:
                    bilgi += f" (Sicil {sicil}" + (f", işe giriş {giris}" if giris else "") + ")"
                parcalar.append(bilgi)
            isy = (((dvo.get("isYeriBilgileri") or {}).get("data") or {}).get("sonuc") or {})
            if isy.get("mesaj"):
                parcalar.append(_temizle(isy.get("mesaj")))
            detay = dvo.get("sskBilgiDetay") or {}
            if "aktif" in detay:
                parcalar.append("Aktif: " + ("Evet" if detay.get("aktif") else "Hayır"))
            return " | ".join(parcalar) if parcalar else (_ilk_hata(dvo) or "Kayıt yok")

        if anahtar == "sskEmekli":
            liste = dvo.get("emekliKaydi4AList") or []
            if liste:
                return f"Emekli kaydı var ({len(liste)} kayıt)"
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "bagkurCalisani":
            liste = dvo.get("tescilKaydi4BList") or []
            if liste:
                t = liste[0] or {}
                parcalar = [_temizle(dvo.get("islemSonucAciklama")) or "Kayıt bulundu"]
                if _temizle(t.get("bagNo")):
                    parcalar.append("Bağ-No " + _temizle(t.get("bagNo")))
                if _temizle(t.get("terkAciklama")):
                    sterk = "Terk: " + _temizle(t.get("terkAciklama"))
                    if t.get("terkTarihi"):
                        sterk += f" ({t.get('terkTarihi')})"
                    parcalar.append(sterk)
                return " | ".join(p for p in parcalar if p)
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "bagkurEmekli":
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "isYeri":
            return _ilk_hata(dvo) or "Kayıt yok"

        return _ilk_hata(dvo) or "Veri var"
    except Exception as e:
        return f"[özet hatası: {e}]"


def tam_metin(veri, girinti=0):
    satirlar = []
    ond = "    " * girinti
    if isinstance(veri, dict):
        for k, v in veri.items():
            if k in NOISE_KEYS:
                continue
            if v is None or v == "" or v == [] or v == {}:
                continue
            if isinstance(v, (dict, list)):
                alt = tam_metin(v, girinti + 1)
                if alt.strip():
                    satirlar.append(f"{ond}{k}:")
                    satirlar.append(alt)
            else:
                satirlar.append(f"{ond}{k}: {str(v).strip()}")
    elif isinstance(veri, list):
        for i, e in enumerate(veri, 1):
            if isinstance(e, (dict, list)):
                alt = tam_metin(e, girinti + 1)
                if alt.strip():
                    satirlar.append(f"{ond}[{i}]")
                    satirlar.append(alt)
            elif e not in (None, ""):
                satirlar.append(f"{ond}- {str(e).strip()}")
    else:
        satirlar.append(f"{ond}{str(veri).strip()}")
    return "\n".join(satirlar)


def hucre_metni(anahtar, dvo):
    if anahtar == "sskCalisani":
        durum, ozet, veri_var = degerlendir_ssk_calisani(dvo)
        bas = f"{durum} | {ozet}" if ozet else durum
        if not veri_var:
            return bas
        tam = tam_metin(dvo)
        return f"{bas}\n──────────\n{tam}" if tam.strip() else bas

    ozet = ozetle(anahtar, dvo)
    tam = tam_metin(dvo)
    if tam.strip() and tam.strip() != ozet.strip():
        return f"{ozet}\n──────────\n{tam}"
    return ozet


# --------------------------------------------------------------------------- #
#  Sorgu motoru — yerel ofis proxy'sine konuşur (orijinalden, parametrize port)
# --------------------------------------------------------------------------- #
class SorguMotoru:
    def __init__(self, log_fn=None, port=PORT):
        self.log = log_fn or (lambda m: None)
        self.base = f"http://127.0.0.1:{port}"

    def baslat(self):
        self.log(f"🔌 Ofis bağlantısı kullanılacak: {self.base} üzerinden canlı UYAP "
                 "oturumuna komut gönderilecek.")
        self.log("ℹ️ Bunun için 'UYAP Ağ Geçidi' açık ve Paylaş/Al ile bağlantı "
                 "başlatılmış olmalı.")

    def yeniden_baglan(self):
        self.log("🔄 UYAP oturumu ofiste otomatik yönetiliyor; kaldığı yerden devam ediliyor.")
        return True

    def _post(self, endpoint, payload, timeout=90, denemeler=3):
        url = f"{self.base}/{endpoint}"
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        son_hata = None
        for d in range(denemeler):
            try:
                req = urllib.request.Request(url, data=body, method="POST")
                for k, v in COMMON_HEADERS.items():
                    req.add_header(k, v)
                with urllib.request.urlopen(req, timeout=timeout) as r:
                    status = r.status
                    text = r.read().decode("utf-8", "replace")
                isaret = text.lstrip()[:300].lower()
                if status in (401, 403) or isaret.startswith("<!doctype") \
                        or isaret.startswith("<html") or "<title>uyap" in isaret:
                    raise OturumHatasi(f"oturum/yetki hatası (HTTP {status}; login sayfası döndü)")
                try:
                    return status, json.loads(text)
                except Exception:
                    return status, text
            except OturumHatasi:
                raise
            except urllib.error.HTTPError as e:
                son_hata = e
                detay = ""
                try:
                    detay = e.read().decode("utf-8", "replace")[:200]
                except Exception:
                    pass
                self.log(f"   ⚠️ {endpoint} deneme {d + 1}/{denemeler}: HTTP {e.code} {detay}")
                if e.code in (401, 403):
                    raise OturumHatasi(f"ofis yetki hatası (HTTP {e.code})")
            except urllib.error.URLError as e:
                son_hata = e
                self.log(f"   ⚠️ {endpoint} deneme {d + 1}/{denemeler}: ofise ulaşılamadı "
                         f"({getattr(e, 'reason', e)}). UYAP Ağ Geçidi açık ve Paylaş/Al aktif mi?")
            except Exception as e:
                son_hata = e
                self.log(f"   ⚠️ {endpoint} deneme {d + 1}/{denemeler} başarısız: {e}")
            if d < denemeler - 1:
                time.sleep(3 * (d + 1))
        raise son_hata

    def dosya_id_bul(self, yil, sira):
        payload = {
            "dosyaDurumKod": 0, "pageSize": 500, "pageNumber": 1,
            "dosyaYil": int(yil), "dosyaSira": int(sira),
            "birimId": "", "birimTuru2": BIRIM_TURU2, "birimTuru3": BIRIM_TURU3,
        }
        _, veri = self._post("search_phrase_detayli.ajx", payload)
        try:
            kayitlar = veri[0]
            if kayitlar:
                return kayitlar[0].get("dosyaId"), kayitlar[0].get("dosyaNo")
        except Exception:
            pass
        return None, None

    def borclular(self, dosya_id):
        _, veri = self._post("dosya_borclu_list.ajx", {"dosyaId": dosya_id})
        sonuc = []
        if isinstance(veri, list):
            for b in veri:
                kb = (b or {}).get("kisiTumDVO") or {}
                sonuc.append({
                    "kisiKurumId": b.get("kisiKurumId"),
                    "adi": _temizle(kb.get("adi")),
                    "soyadi": _temizle(kb.get("soyadi")),
                    "tc": _temizle(kb.get("tcKimlikNo")),
                })
        return sonuc

    def borclu_sec(self, borclular, isim):
        if not borclular:
            return None, False
        if len(borclular) == 1:
            return borclular[0], True
        hedef = tr_normalize(isim)
        if hedef:
            for b in borclular:
                tam = tr_normalize(b["adi"] + " " + b["soyadi"])
                if tam and (tam == hedef or tam in hedef or hedef in tam):
                    return b, True
            hedef_kelime = set(hedef.split())
            for b in borclular:
                tam_kelime = set(tr_normalize(b["adi"] + " " + b["soyadi"]).split())
                if tam_kelime and tam_kelime <= hedef_kelime:
                    return b, True
        return borclular[0], False

    def sgk_sorgula(self, dosya_id, kisi_kurum_id, anahtarlar, kontrol):
        sonuc = {}
        payload = {"dosyaId": dosya_id, "kisiKurumId": kisi_kurum_id}
        for _, endpoint, anahtar in SORGULAR:
            if anahtar not in anahtarlar:
                continue
            if not kontrol():
                break
            try:
                durum, veri = self._post(endpoint, payload)
                if isinstance(veri, dict):
                    dvo = veri.get("sorguSonucDVO", veri)
                    sonuc[anahtar] = (True, hucre_metni(anahtar, dvo))
                else:
                    sonuc[anahtar] = (False, f"{HATA_ON} HTTP {durum}")
            except Exception as e:
                sonuc[anahtar] = (False, f"{HATA_ON} {e}")
            time.sleep(SORGU_ARASI_BEKLEME)
        return sonuc


# --------------------------------------------------------------------------- #
#  Toplu sorgu — Excel + iş parçacığı + iç durum (her iki ön yüz snapshot okur)
# --------------------------------------------------------------------------- #
def _kisa(metin):
    if metin is None:
        return ""
    t = str(metin).split("\n", 1)[0]
    return (t[:90] + "…") if len(t) > 90 else t


class SgkBatch:
    """Excel tabanlı toplu SGK sorgusu.

    Kullanım:
        b = SgkBatch(src_path, port=PORT)
        b.yukle()                       # workbook + satırları hazırla
        b.basla(secili=set([...]), mode="normal")   # arka planda çalışır
        b.snapshot(since_seq, since_log)             # arayüz tazeleme (poll)
        b.duraklat() / b.devam() / b.durdur()

    Sonuçlar yüklenen kök dosyaya DEĞİL, yanındaki '*_yapilanlar.xlsx' (tam) ve
    '*_yapilanlar_ozet.xlsx' (özet) dosyalarına yazılır; varsa ondan devam eder.
    """

    def __init__(self, src_path, *, port=PORT):
        self.src_path = src_path
        self.port = port

        self.wb = None
        self.ws = None
        self.excel_yolu = None
        self.ozet_yolu = None
        self.satir_no = []           # veri satırı indeksleri (Excel, 1-tabanlı)

        # İç durum — snapshot ile okunur
        self._lock = threading.RLock()
        self._logs = []              # [str]
        self._rows = {}              # excel_r -> row dict (canlı)
        self._seq = 0
        self._row_seq = {}           # excel_r -> seq
        self.status_text = ""
        self.progress = {"done": 0, "total": 0}
        self.running = False
        self.bitti = False

        self.pause_event = threading.Event()
        self.stop_event = threading.Event()
        self._thread = None

    # ── log / satır güncelleme ──
    def log(self, mesaj):
        with self._lock:
            self._logs.append(str(mesaj))

    def _set_status(self, text):
        with self._lock:
            self.status_text = text

    def _row_payload(self, r):
        return {
            "r": r,
            "no": self._cell(r, 1), "ad": self._cell(r, 2),
            "birim": self._cell(r, 3), "dosya": self._cell(r, 4),
            "sonuc": [_kisa(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                      for i in range(len(SORGULAR))],
            "tam": [self._cell(r, SONUC_BASLANGIC_KOL + i) for i in range(len(SORGULAR))],
            "durum": self._satir_durumu(r),
        }

    def _cell(self, r, c):
        v = self.ws.cell(r, c).value
        return "" if v is None else str(v)

    def _row_guncelle(self, r):
        with self._lock:
            self._seq += 1
            self._row_seq[r] = self._seq
            self._rows[r] = self._row_payload(r)

    # ── yükleme ──
    def yukle(self):
        import openpyxl
        kok, uzanti = os.path.splitext(self.src_path)
        for ek in (OZET_EKI, YAPILANLAR_EKI):
            if kok.endswith(ek):
                kok = kok[: -len(ek)]
                break
        self.excel_yolu = kok + YAPILANLAR_EKI + uzanti
        self.ozet_yolu = kok + OZET_EKI + uzanti

        if os.path.exists(self.excel_yolu):
            self.wb = openpyxl.load_workbook(self.excel_yolu)
            self.log(f"↩️ Yapılanlar dosyasından devam ediliyor: {os.path.basename(self.excel_yolu)}")
        else:
            self.wb = openpyxl.load_workbook(self.src_path)
        self.ws = self.wb.active

        if not self._baslik_satiri_mi():
            self.ws.insert_rows(1)
            for c, ad in enumerate(KOLONLAR, start=1):
                self.ws.cell(1, c, value=ad)

        self.satir_no = []
        with self._lock:
            self._rows = {}
            self._row_seq = {}
        for r in range(2, self.ws.max_row + 1):
            temel = [self.ws.cell(r, c).value for c in range(1, 5)]
            if all(v is None for v in temel):
                continue
            self.satir_no.append(r)
            self._rows[r] = self._row_payload(r)
        self._kaydet()              # başlık/biçimi hemen yaz
        self.log(f"📂 Yüklendi: {os.path.basename(self.src_path)} ({len(self.satir_no)} satır)")
        self.log(f"💾 Tam veri → {os.path.basename(self.excel_yolu)}  ·  "
                 f"Özet → {os.path.basename(self.ozet_yolu)}")
        return self.rows_listesi()

    def _baslik_satiri_mi(self):
        try:
            deg = [str(self.ws.cell(1, c).value or "").strip().lower() for c in range(1, 5)]
            return deg == [b.lower() for b in KOLONLAR[:4]]
        except Exception:
            return False

    # ── hücre durumu yardımcıları ──
    @staticmethod
    def _bos(v):
        return v is None or str(v).strip() == ""

    def _hata_cell(self, v):
        return (not self._bos(v)) and str(v).lstrip().startswith(HATA_ON)

    def gereken_anahtarlar(self, r, mode, secili):
        gerekli = set()
        for i, anahtar in enumerate(ANAHTARLAR):
            if anahtar not in secili:
                continue
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if mode == "retry":
                if self._hata_cell(v):
                    gerekli.add(anahtar)
            else:
                if self._bos(v) or self._hata_cell(v):
                    gerekli.add(anahtar)
        return gerekli

    def _satir_tamamlandi_mi(self, r, secili=None):
        secili_var = False
        for i, anahtar in enumerate(ANAHTARLAR):
            if secili is not None and anahtar not in secili:
                continue
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if secili is None and self._bos(v):
                continue  # genel durum: dolu sütunlara bak
            secili_var = True
            if self._bos(v) or self._hata_cell(v):
                return False
        return secili_var

    def _satir_durumu(self, r):
        """Satır rengi/etiketi: 'sirket' | 'tamam' | 'hata' | ''."""
        degerler = [self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
                    for i in range(len(SORGULAR))]
        dolu = [v for v in degerler if not self._bos(v)]
        if not dolu:
            return ""
        if all(str(v).strip() == SIRKET_NOT for v in dolu):
            return "sirket"
        if any(self._hata_cell(v) for v in dolu):
            return "hata"
        return "tamam"

    # ── sayım ──
    def hata_sayisi(self, secili):
        return sum(len(self.gereken_anahtarlar(r, "retry", secili)) for r in self.satir_no)

    # ── kayıt ──
    def _kaydet(self):
        if not self.wb or not self.excel_yolu:
            return
        try:
            self.wb.save(self.excel_yolu)
        except PermissionError:
            self.log("⚠️ Yapılanlar dosyası Excel'de açık olabilir, kapatın.")
            return
        except Exception as e:
            self.log(f"⚠️ Kaydetme hatası: {e}")
            return
        self._ozet_yaz()

    def _ozet_yaz(self):
        import openpyxl
        if not self.ozet_yolu or not self.ws:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(list(KOLONLAR))
            for r in self.satir_no:
                satir = [self.ws.cell(r, c).value for c in range(1, 5)]
                satir += [(str(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value).split("\n", 1)[0]
                           if self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value is not None else None)
                          for i in range(len(SORGULAR))]
                ws.append(satir)
            wb.save(self.ozet_yolu)
        except PermissionError:
            self.log("⚠️ Özet dosyası Excel'de açık olabilir, kapatın.")
        except Exception as e:
            self.log(f"⚠️ Özet yazma hatası: {e}")

    # ── kontrol ──
    def basla(self, secili, mode="normal"):
        if self.running or not self.ws:
            return False
        secili = set(secili) & set(ANAHTARLAR)
        if not secili:
            self.log("⚠️ En az bir sorgu sütunu seçilmeli.")
            return False
        self.running = True
        self.bitti = False
        self.stop_event.clear()
        self.pause_event.clear()
        self._thread = threading.Thread(target=self._worker, args=(mode, secili), daemon=True)
        self._thread.start()
        return True

    def duraklat(self):
        self.pause_event.set()

    def devam(self):
        self.pause_event.clear()

    def durdur(self):
        self.stop_event.set()
        self.pause_event.clear()
        self._set_status("Durduruluyor…")

    @property
    def paused(self):
        return self.pause_event.is_set()

    def _bekle_devam(self):
        while self.pause_event.is_set() and not self.stop_event.is_set():
            self._set_status("⏸ Duraklatıldı")
            time.sleep(0.3)
        return not self.stop_event.is_set()

    def _yaz_hepsi(self, r, gereken, offset, metin):
        for a in gereken:
            self.ws.cell(r, SONUC_BASLANGIC_KOL + offset[a], value=metin)
        self._row_guncelle(r)

    # ── worker ──
    def _worker(self, mode, secili):
        motor = SorguMotoru(self.log, port=self.port)
        try:
            motor.baslat()
            satirlar = list(self.satir_no)
            toplam = len(satirlar)
            self.progress = {"done": 0, "total": toplam}
            offset = {anahtar: i for i, anahtar in enumerate(ANAHTARLAR)}
            hata_sayaci = 0
            pespese = 0
            oturum_yenilendi = False

            def oturum_yenile():
                nonlocal oturum_yenilendi, pespese
                oturum_yenilendi = True
                pespese = 0
                try:
                    motor.yeniden_baglan()
                    return True
                except Exception as e:
                    self.log(f"❌ Oturum yenilenemedi: {e}")
                    return False

            def mola_kontrol():
                nonlocal hata_sayaci
                if hata_sayaci >= MOLA_HATA_ESIGI:
                    hata_sayaci = 0
                    self.log(f"⏳ {MOLA_HATA_ESIGI} hata oluştu, {MOLA_SURESI} sn mola veriliyor…")
                    bekleyen = MOLA_SURESI
                    while bekleyen > 0 and not self.stop_event.is_set():
                        self._set_status(f"⏳ Otomatik mola: {bekleyen} sn")
                        time.sleep(1)
                        bekleyen -= 1
                    if not self.stop_event.is_set():
                        self.log("▶ Molaya devam ediliyor.")
                return not self.stop_event.is_set()

            for idx, r in enumerate(satirlar, 1):
                if not self._bekle_devam():
                    break
                if not mola_kontrol():
                    break
                self.progress = {"done": idx - 1, "total": toplam}
                self._set_status(f"İşleniyor {idx}/{toplam} (satır {r})")

                gereken = self.gereken_anahtarlar(r, mode, secili)
                if not gereken:
                    continue

                d = self.ws.cell(r, 4).value
                isim = self.ws.cell(r, 2).value
                if sirket_mi(isim):
                    self.log(f"satır {r}: '{isim}' şirket (tüzel kişi) — sorgu yapılmadı.")
                    for i, anahtar in enumerate(ANAHTARLAR):
                        if anahtar in secili:
                            self.ws.cell(r, SONUC_BASLANGIC_KOL + i, value=SIRKET_NOT)
                    self._row_guncelle(r)
                    continue
                if not d or "/" not in str(d):
                    self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} Geçersiz dosya no: {d}")
                    hata_sayaci += 1
                    continue

                yil, _, sira = str(d).partition("/")
                try:
                    dosya_id, _ = motor.dosya_id_bul(yil.strip(), sira.strip())
                except OturumHatasi as e:
                    self.log(f"⛔ Oturum hatası: {e}")
                    if not oturum_yenilendi and oturum_yenile():
                        continue
                    self._oturum_dur()
                    break
                except Exception as e:
                    self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} arama: {e}")
                    hata_sayaci += 1
                    pespese += 1
                    if pespese >= OTURUM_KONTROL_ESIGI and not oturum_yenilendi:
                        self.log(f"⚠️ {OTURUM_KONTROL_ESIGI} ardışık hata — oturum kontrol ediliyor…")
                        oturum_yenile()
                    continue
                if not dosya_id:
                    self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} Dosya bulunamadı")
                    self.log(f"satır {r}: {d} -> dosya bulunamadı")
                    hata_sayaci += 1
                    pespese += 1
                    if pespese >= OTURUM_KONTROL_ESIGI and not oturum_yenilendi:
                        self.log(f"⚠️ {OTURUM_KONTROL_ESIGI} ardışık 'dosya bulunamadı' — yenileniyor…")
                        oturum_yenile()
                    continue

                try:
                    borclular = motor.borclular(dosya_id)
                except OturumHatasi as e:
                    self.log(f"⛔ Oturum hatası: {e}")
                    if not oturum_yenilendi and oturum_yenile():
                        continue
                    self._oturum_dur()
                    break
                except Exception as e:
                    self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} borçlu: {e}")
                    hata_sayaci += 1
                    continue
                borclu, eslesti = motor.borclu_sec(borclular, isim)
                if not borclu:
                    self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} Borçlu bulunamadı")
                    hata_sayaci += 1
                    continue

                pespese = 0
                oturum_yenilendi = False
                self.log(f"satır {r}: {d} -> {borclu['adi']} {borclu['soyadi']}"
                         + ("" if eslesti else "  (⚠ isim eşleşmedi, ilk borçlu)"))

                ozetler = motor.sgk_sorgula(dosya_id, borclu["kisiKurumId"], gereken, self._bekle_devam)
                for anahtar, (ok, metin) in ozetler.items():
                    self.ws.cell(r, SONUC_BASLANGIC_KOL + offset[anahtar], value=metin)
                    if not ok:
                        hata_sayaci += 1
                if ozetler:
                    self._row_guncelle(r)

                self.progress = {"done": idx, "total": toplam}
                if self.stop_event.is_set():
                    break
                if idx % 10 == 0:
                    self._kaydet()
                time.sleep(SATIR_ARASI_BEKLEME)

            self._kaydet()
            if self.stop_event.is_set():
                self._set_status("Durduruldu")
                self.log("⏹ Durduruldu.")
            else:
                self._set_status("Tamamlandı")
                self.log("🎉 İşlem tamamlandı.")
        except Exception as e:
            self.log(f"❌ Beklenmeyen hata: {e}")
            self._set_status("Hata")
        finally:
            self.running = False
            self.bitti = True

    def _oturum_dur(self):
        self.stop_event.set()
        self._set_status("⛔ Ofis bağlantısı yok")
        self.log("⛔ Yerel ofis bağlantısına (127.0.0.1:8800) ulaşılamadı / yetki reddedildi. "
                 "UYAP Ağ Geçidi açık ve Paylaş/Al aktif mi? İşlem durduruldu.")

    # ── snapshot (arayüz tazeleme) ──
    def rows_listesi(self):
        with self._lock:
            return [self._rows[r] for r in self.satir_no if r in self._rows]

    def snapshot(self, since_seq=0, since_log=0):
        with self._lock:
            updates = [self._rows[r] for r, s in self._row_seq.items()
                       if s > since_seq and r in self._rows]
            return {
                "running": self.running,
                "paused": self.paused,
                "bitti": self.bitti,
                "status": self.status_text,
                "progress": dict(self.progress),
                "seq": self._seq,
                "updates": updates,
                "logs": self._logs[since_log:],
                "total_logs": len(self._logs),
            }
