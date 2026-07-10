"""
uyap_core.mts.sgk_excel — SGK toplu sorgu İSTEMCİ tarafı Excel I/O
==================================================================
Kararlı/SGK Sorgu GUI'sinin openpyxl tabanlı dosya mantığının çekirdeği. İstemcide
çalışır: kullanıcı Excel'i seçer, hangi satır/sütunların sorgulanması gerektiği
belirlenir (devam-etme: dolu ve hatasız hücreler atlanır), iş kuyruğundan dönen
sonuçlar sabit (satır, sütun) konumuna yazılır ve "_yapilanlar.xlsx" + "_yapilanlar_ozet.xlsx"
olarak kaydedilir.

Sorgu motoru ofiste (job_handlers.sgk_toplu_sorgu) çalışır; bu modül yalnızca giriş
hazırlama + sonuç yazma yapar. openpyxl gerektirir (yalnızca istemcide).
"""

import os

from .sgk import (
    SORGULAR, ANAHTAR_OFFSET, TUM_ANAHTARLAR, HATA_ON, SIRKET_NOT,
)

# Kolon düzeni: A..D sabit + E.. sonuç sütunları (SORGULAR sırasıyla).
KOLONLAR = ["No", "Ad Soyad", "Birim", "Dosya No"] + [b for (b, _, _) in SORGULAR]
SONUC_BASLANGIC_KOL = 5         # E
YAPILANLAR_EKI = "_yapilanlar"
OZET_EKI = "_yapilanlar_ozet"


def _bos(v):
    return v is None or str(v).strip() == ""


def _hata_cell(v):
    return (not _bos(v)) and str(v).lstrip().startswith(HATA_ON)


def _ilk_satir(deger):
    return None if deger is None else str(deger).split("\n", 1)[0]


class SgkCalisma:
    """Bir SGK sorgu çalışmasının Excel durumu (yükle → hazırla → sonuç yaz → kaydet)."""

    def __init__(self):
        self.kaynak_yolu = None     # kullanıcının seçtiği dosya (DEĞİŞTİRİLMEZ)
        self.excel_yolu = None      # "_yapilanlar" (tüm sonuçlar)
        self.ozet_yolu = None       # "_yapilanlar_ozet" (kısa)
        self.wb = None
        self.ws = None
        self.satirlar = []          # geçerli veri satırı numaraları (excel r)

    # ── Yükleme ──────────────────────────────────────────────────────────────
    def yukle(self, secilen_yol):
        """Seçilen dosyayı yükler. '_yapilanlar' varsa ondan DEVAM eder. Başlık satırını
        garanti eder. (satir_sayisi) döndürür."""
        import openpyxl
        kok, uzanti = os.path.splitext(secilen_yol)
        for ek in (OZET_EKI, YAPILANLAR_EKI):
            if kok.endswith(ek):
                kok = kok[:-len(ek)]
                break
        self.kaynak_yolu = secilen_yol
        self.excel_yolu = kok + YAPILANLAR_EKI + uzanti
        self.ozet_yolu = kok + OZET_EKI + uzanti

        if os.path.exists(self.excel_yolu):
            self.wb = openpyxl.load_workbook(self.excel_yolu)   # önceki sonuçlardan devam
        else:
            self.wb = openpyxl.load_workbook(secilen_yol)
        self.ws = self.wb.active

        if not self._baslik_satiri_mi():
            self.ws.insert_rows(1)
            for c, ad in enumerate(KOLONLAR, start=1):
                self.ws.cell(1, c, value=ad)

        self.satirlar = []
        for r in range(2, self.ws.max_row + 1):
            if all(self.ws.cell(r, c).value is None for c in range(1, 5)):
                continue
            self.satirlar.append(r)
        return len(self.satirlar)

    def _baslik_satiri_mi(self):
        try:
            deg = [str(self.ws.cell(1, c).value or "").strip().lower() for c in range(1, 5)]
            return deg == [b.lower() for b in KOLONLAR[:4]]
        except Exception:
            return False

    # ── Satır/tablo erişimi (GUI tablosu için) ─────────────────────────────────
    def satir_temel(self, r):
        """(No, Ad Soyad, Birim, Dosya No) — tablo gösterimi için."""
        return [self.ws.cell(r, c).value for c in range(1, 5)]

    def satir_sonuc_kisa(self, r):
        """Sonuç sütunlarının kısa (ilk satır) gösterimi — tablo için."""
        return [_ilk_satir(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                for i in range(len(SORGULAR))]

    def hucre_tam(self, r, anahtar):
        v = self.ws.cell(r, SONUC_BASLANGIC_KOL + ANAHTAR_OFFSET[anahtar]).value
        return "" if v is None else str(v)

    # ── Giriş hazırlama (devam-etme: dolu+hatasız hücreleri atla) ──────────────
    def gereken_anahtarlar(self, r, secili_anahtarlar, mode="normal"):
        """Bu satırda hangi sorguların çalışması gerektiğini belirler.
        mode 'normal': boş veya hatalı hücreler; 'retry': yalnız hatalı hücreler."""
        gerekli = []
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if anahtar not in secili_anahtarlar:
                continue
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if mode == "retry":
                if _hata_cell(v):
                    gerekli.append(anahtar)
            else:
                if _bos(v) or _hata_cell(v):
                    gerekli.append(anahtar)
        return gerekli

    def is_satirlari(self, secili_anahtarlar, mode="normal"):
        """İş kuyruğuna gönderilecek satır parametrelerini üretir (gereken sorgusu olanlar)."""
        out = []
        for r in self.satirlar:
            gereken = self.gereken_anahtarlar(r, secili_anahtarlar, mode)
            if not gereken:
                continue
            out.append({
                "id": r,
                "ad_soyad": self.ws.cell(r, 2).value or "",
                "dosya_no": self.ws.cell(r, 4).value or "",
                "gereken": gereken,
            })
        return out

    # ── Sonuç yazma ────────────────────────────────────────────────────────────
    def sonuc_yaz(self, satir_sonuc):
        """İş kuyruğundan dönen tek satır sonucunu uygun hücrelere yazar.
        satir_sonuc: {"id": <excel r>, "sonuclar": {anahtar: metin}, ...}."""
        r = satir_sonuc.get("id")
        if r is None:
            return
        for anahtar, metin in (satir_sonuc.get("sonuclar") or {}).items():
            off = ANAHTAR_OFFSET.get(anahtar)
            if off is None:
                continue
            self.ws.cell(int(r), SONUC_BASLANGIC_KOL + off, value=metin)

    def satir_tamamlandi(self, r, secili_anahtarlar):
        """Seçili tüm sorgu sütunları dolu ve hatasız mı?"""
        secili_var = False
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if anahtar not in secili_anahtarlar:
                continue
            secili_var = True
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if _bos(v) or _hata_cell(v):
                return False
        return secili_var

    # ── Kaydetme ────────────────────────────────────────────────────────────────
    def kaydet(self):
        """'_yapilanlar' (tam) ve '_yapilanlar_ozet' (kısa) dosyalarını yazar.
        Kullanıcının seçtiği kök dosya DEĞİŞTİRİLMEZ. (hata_mesaji | None) döndürür."""
        if not self.wb or not self.excel_yolu:
            return "Çalışma yüklenmedi."
        try:
            self.wb.save(self.excel_yolu)
        except PermissionError:
            return "Yapılanlar dosyası Excel'de açık olabilir, kapatın."
        except Exception as e:
            return f"Kaydetme hatası: {e}"
        return self._ozet_yaz()

    def _ozet_yaz(self):
        import openpyxl
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(list(KOLONLAR))
            for r in self.satirlar:
                satir = [self.ws.cell(r, c).value for c in range(1, 5)]
                satir += [_ilk_satir(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                          for i in range(len(SORGULAR))]
                ws.append(satir)
            wb.save(self.ozet_yolu)
        except PermissionError:
            return "Özet dosyası Excel'de açık olabilir, kapatın."
        except Exception as e:
            return f"Özet yazma hatası: {e}"
        return None
