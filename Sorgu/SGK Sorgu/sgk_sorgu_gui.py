# -*- coding: utf-8 -*-
"""
SGK Toplu Sorgu - GUI
=====================
mts_session_replayer.py mantığının (UYAP oturumu + APIRequestContext ile
backend isteklerini tekrar oynatma) Excel tabanlı, GUI'li bir uyarlamasıdır.

Akış:
  1. Bir Excel dosyası yüklenir ve pencerede gösterilir.
     - A: sıra/no   B: ad-soyad   C: birim   D: "2025/144223" (icra dosya yıl/no)
  2. "Sorgula" denince her satır için:
       - D sütunu  "/"  ile bölünür -> dosyaYil, dosyaSira
       - search_phrase_detayli.ajx  -> dosyaId
       - dosya_borclu_list.ajx      -> borçlular (kisiKurumId)
       - B sütunundaki isimle eşleşen borçlu seçilir (yoksa ilk borçlu)
       - 7 SGK sorgusu çalıştırılır (Kamu Çalışan/Emekli, SSK Çalışan/Emekli,
         Bağkur Çalışan/Emekli, SSK İş Yeri)
       - Gelen verinin TAMAMI ilgili satırın E,F,G,H,I,J,K sütunlarına yazılır.
         Her sorgunun sütunu SABİTTİR; sorgulanmayan sütun boş kalır.
  3. Sonuçlar yüklenen KÖK dosyaya DEĞİL, yanındaki "X_yapilanlar.xlsx" (tam
     veri) ve "X_yapilanlar_ozet.xlsx" (kısa özet) dosyalarına yazılır.
     - 1. satır BAŞLIK satırıdır.
     - KAYDIRMA YOK: bellekteki ana tablo her kayıtta sabit satır/sütun
       sırasıyla diske yazılır; sonradan eklenen bir sorgu (örn. önce SSK
       Çalışan, sonra Kamu Çalışan) ait olduğu sütuna, doğru satıra yazılır.
     - "X_yapilanlar.xlsx" varsa ondan DEVAM edilir (önceki sonuçlar korunur).
       Hücre dolu ve hatasızsa yeniden sorgulanmaz; hatalılar [HATA] kalır.

Özellikler:
  - Başarılı sorgulanan hücre tekrar sorgulanmaz ("sorgulandı" takibi).
  - Hatalı hücreler  [HATA]  ile işaretlenir; "Hatalıları Tekrar Dene" sadece
    bunları yeniden sorgular.
  - Duraklat / Durdur kontrolleri.
  - Sütun bazlı filtreleme.
  - Hücre / sütun verisini sağ tıklayıp kopyalama.
  - Hücreye tek tıklayınca tüm detayını yan panelde gösterme.

Not: UYAP oturumu mts_session_replayer.py ile aynı şekilde
     C:\\Users\\KalkanHukuk\\Desktop\\UYAPDjango\\uyap_session.json
     ve  chrome_profile  klasöründen yüklenir.
"""

import os
import re
import json
import time
import queue
import threading
import traceback

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
import urllib.request
import urllib.error

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SESSION_PATH = r"C:\Users\KalkanHukuk\Desktop\UYAPDjango\uyap_session.json"  # (artık kullanılmıyor)
BASE_URL = "https://avukat.uyap.gov.tr"
# Tüm UYAP istekleri bu YEREL ofis proxy'sine gider; ofis (uyap_app.py) isteği canlı
# e-imza oturumuyla iletir. Playwright/tarayıcı ve uyap_session.json YOK.
OFFICE_BASE = "http://127.0.0.1:8800"

# Birim türleri (logdan): İcra -> 2, İcra Dairesi -> 1101
BIRIM_TURU2 = "1101"
BIRIM_TURU3 = "2"

# Tüm backend istekleri için ortak başlıklar (replayer'daki "safe" başlıklar)
COMMON_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "content-type": "application/json",
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "expires": "0",
    "referer": f"{BASE_URL}/dosya-sorgulama",
}

# Sorgu tanımları: (kolon_basligi, endpoint, ozetleyici_anahtar)
SORGULAR = [
    ("Kamu Çalışan",   "borclu_bilgileri_goruntule_sgk_kamuCalisaniBilgileri.ajx",   "kamuCalisani"),
    ("Kamu Emekli",    "borclu_bilgileri_goruntule_sgk_kamuEmekliBilgileri.ajx",     "kamuEmekli"),
    ("SSK Çalışan",    "borclu_bilgileri_goruntule_sgk_sskCalisaniBilgileri.ajx",    "sskCalisani"),
    ("SSK Emekli",     "borclu_bilgileri_goruntule_sgk_sskEmekliBilgileri.ajx",      "sskEmekli"),
    ("Bağkur Çalışan", "borclu_bilgileri_goruntule_sgk_bagkurCalisaniBilgileri.ajx", "bagkurCalisani"),
    ("Bağkur Emekli",  "borclu_bilgileri_goruntule_sgk_bagkurEmekliBilgileri.ajx",   "bagkurEmekli"),
    ("SSK İş Yeri",    "borclu_bilgileri_goruntule_sgk_isYeriBilgileri.ajx",         "isYeri"),
]

# Sonuçların yazılacağı ilk sütun (E = 5). E,F,G,H,I,J,K
SONUC_BASLANGIC_KOL = 5
HATA_ON = "[HATA]"   # hatalı hücre işareti
SIRKET_NOT = "ŞİRKET - sorgulanmadı"   # tüzel kişi (ŞTİ/Şirketi) -> SGK sorgusu yapılmaz

# İsimde tüzel kişiyi ele veren anahtar kelimeler (tr_normalize sonrası, BÜYÜK/ASCII)
SIRKET_ANAHTARLARI = {"STI", "SIRKETI", "SIRKET", "LTD"}

# Hız / yük ayarları (UYAP throttle'ını azaltmak için)
SORGU_ARASI_BEKLEME = 1.5     # aynı satırdaki SGK sorguları arası bekleme (sn)
SATIR_ARASI_BEKLEME = 0.2     # satırlar arası bekleme (sn)
MOLA_HATA_ESIGI = 3           # bu kadar hatadan sonra otomatik mola
MOLA_SURESI = 60              # otomatik mola süresi (sn)
# Bu kadar ARDIŞIK "dosya bulunamadı/hata" olursa oturum kopmuş kabul edilir
OTURUM_KONTROL_ESIGI = 6


class OturumHatasi(Exception):
    """UYAP girişi/oturumu düşmüş (login sayfası ya da yetkisiz yanıt döndü)."""

# Yapılanlar (çıktı) dosyası eki: yüklenen "X.xlsx" -> "X_yapilanlar.xlsx"
YAPILANLAR_EKI = "_yapilanlar"
# Özet (kısa) çıktı dosyası eki: "X.xlsx" -> "X_yapilanlar_ozet.xlsx"
OZET_EKI = "_yapilanlar_ozet"

# Tam dökümde atlanacak gürültü alanları
NOISE_KEYS = {"isNew", "isMock", "hasError", "hasData", "metadata",
              "yeniBirimEkle", "orgKoduDegisti", "isTumunuKopyala", "testMi"}


# --------------------------------------------------------------------------- #
#  Türkçe normalizasyon ve veri özetleme/dökme yardımcıları
# --------------------------------------------------------------------------- #
def tr_normalize(s):
    """İsim eşleştirme için: büyük harf + Türkçe karakter sadeleştirme."""
    if s is None:
        return ""
    s = str(s)
    cevrim = str.maketrans("ıİiİğĞüÜşŞöÖçÇ", "IIIIGGUUSSOOCC")
    s = s.translate(cevrim)
    s = s.upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _temizle(v):
    if v is None:
        return ""
    return str(v).strip()


def _ilk_hata(dvo):
    for k in ("hataMesaji", "anaHataMesaji", "sonucMesaj", "mesaj", "islemSonucAciklama"):
        if isinstance(dvo, dict) and dvo.get(k):
            return _temizle(dvo[k])
    return ""


def _g(d, *yol):
    """Güvenli iç içe erişim: _g(dvo, 'a', 'b', 0, 'c')."""
    for k in yol:
        if isinstance(d, dict):
            d = d.get(k)
        elif isinstance(d, list) and isinstance(k, int) and -len(d) <= k < len(d):
            d = d[k]
        else:
            return None
    return d


def _parcala(*ciftler):
    """('Etiket', değer) çiftlerinden boş olmayanları birleştirir."""
    out = []
    for et, dg in ciftler:
        dg = _temizle(dg)
        if dg:
            out.append(f"{et}: {dg}" if et else dg)
    return " | ".join(out)


# --- SSK Çalışan: Olumlu/Olumsuz değerlendirmesi (kullanıcı kuralları) ------- #
#  * İşten çıkış tarihi (istenCikisTarihi) VARSA -> maaş haczi yazılamaz -> Olumsuz
#    (yine de son işyeri bilgileri yazılır).
#  * Tescil/işyeri verisi var ve çıkış yoksa -> hâlen çalışıyor -> Olumlu.
#  * Hiç kayıt yoksa -> Olumsuz (yalnızca mesaj yazılır).
def degerlendir_ssk_calisani(dvo):
    """Döner: (durum, ozet, veri_var).  durum = 'Olumlu' | 'Olumsuz'."""
    if not isinstance(dvo, dict):
        s = _temizle(dvo)
        return ("Olumlu" if s else "Olumsuz"), s, bool(s)
    try:
        sskBilgi = dvo.get("sskBilgi") or {}
        tescil_list = sskBilgi.get("tescilKaydi4AList") or []
        isyeri_sonuc = _g(dvo, "isYeriBilgileri", "data", "sonuc") or {}
        ozet_array = isyeri_sonuc.get("isyeriTescil4AOzetArray") or []
        sig = _g(dvo, "sskBilgiDetay", "sgkIsyeriDVO", "sigortaliBilgi") or {}

        # GÜNCEL 4a hizmeti var mı?  Belirleyici alan: isYeriBilgileri.sonuc.basarili
        # (Eski sgkIsyeriDVO/tescil dolu olsa bile basarili != true ise güncel hizmet
        #  yoktur; örn. "Bu yıl ve bir önceki yıl için 4a hizmet verisi bulunamadı.")
        b = isyeri_sonuc.get("basarili")
        guncel_hizmet = (b is True) or (str(b).strip().lower() == "true")

        # Ortak alanlar (son bilinen işyeri — güncel ya da geçmiş)
        oz = ozet_array[0] if ozet_array else {}
        unvan = _temizle(sig.get("isyeriUnvan")) or _temizle(oz.get("unvan"))
        vergi = _temizle(oz.get("teVergiNo"))
        donem = _temizle(sig.get("donemYilAy"))
        if not donem and oz.get("sonDonemYil"):
            donem = f"{_temizle(oz.get('sonDonemYil'))}/{_temizle(oz.get('sonDonemAy'))}"
        gun = oz.get("sonDonemHizmetGun")
        sicil = _temizle(sig.get("sicilNo")) or _temizle(_g(tescil_list, 0, "sicilNo"))
        giris = (_temizle(sig.get("ilkIseGirisTarihi"))
                 or _temizle(_g(tescil_list, 0, "kimlikBilgileri", "iseGirisTarihi")))
        adres = _temizle(sig.get("isyeriAdresi"))
        cikis = _temizle(sig.get("istenCikisTarihi"))

        gun_s = ""
        if gun not in (None, ""):
            gun_s = f"son dönem {gun} gün"
            try:
                if int(gun) < 30:
                    gun_s += " (ay içinde çıkış olabilir)"
            except (TypeError, ValueError):
                pass

        # OLUMLU: güncel 4a hizmeti var VE işten çıkış tarihi yok -> hâlen çalışıyor
        if guncel_hizmet and not cikis:
            ozet = _parcala((None, "Çalışıyor"), ("İşyeri", unvan), ("VKN", vergi),
                            ("Son dönem", donem), (None, gun_s), ("Adres", adres),
                            ("Sicil", sicil), ("İşe giriş", giris))
            return "Olumlu", ozet, True

        # OLUMSUZ — alt durumlar:
        mesaj = (_temizle(isyeri_sonuc.get("mesaj"))
                 or _temizle(sskBilgi.get("sonucMesaji"))
                 or _temizle(sskBilgi.get("anaHataMesaji")))
        # (a) İşten çıkış tarihi var -> son işyeri bilgileriyle
        if cikis:
            ozet = _parcala(("İşten çıkış", cikis), ("Son işyeri", unvan), ("VKN", vergi),
                            ("Son dönem", donem), (None, gun_s), ("Adres", adres),
                            ("Sicil", sicil), ("İşe giriş", giris))
            return "Olumsuz", ozet, True
        # (b) Güncel hizmet yok ama son bilinen işyeri var -> onu göster
        if unvan or donem or sicil:
            son = _parcala(("Son işyeri", unvan), ("Son dönem", donem),
                           ("Adres", adres), ("Sicil", sicil), ("İşe giriş", giris))
            ozet = mesaj or "Güncel 4a hizmeti bulunamadı"
            if son:
                ozet = f"{ozet} | {son}"
            return "Olumsuz", ozet, True
        # (c) Hiç işyeri bilgisi yok; sadece geçmiş tescil veya hiçbir kayıt
        if tescil_list:
            kb = _g(tescil_list, 0, "kimlikBilgileri") or {}
            ek = _parcala(("Eski sicil", _temizle(kb.get("sicilNo"))),
                          ("İlk işe giriş", _temizle(kb.get("iseGirisTarihi"))))
            ozet = mesaj or "Güncel 4a hizmeti bulunamadı"
            if ek:
                ozet = f"{ozet} ({ek})"
            return "Olumsuz", ozet, True
        return "Olumsuz", (mesaj or "Kayıt bulunamadı"), False
    except Exception as e:
        return "Olumsuz", f"[değerlendirme hatası: {e}]", False


def ozetle(anahtar, dvo):
    """Hücrenin ilk satırı: hızlı bakış için kısa özet."""
    try:
        if not isinstance(dvo, dict):
            return _temizle(dvo)

        if anahtar == "kamuCalisani":
            if dvo.get("durumAciklamasi4C") or dvo.get("kurum"):
                parcalar = [_temizle(dvo.get("durumAciklamasi4C")),
                            _temizle(dvo.get("kurum")),
                            _temizle(dvo.get("unvani"))]
                yer = "/".join(p for p in (_temizle(dvo.get("kurumIl")),
                                           _temizle(dvo.get("kurumIlce"))) if p)
                if yer:
                    parcalar.append(yer)
                if dvo.get("iseBaslamaTarihi"):
                    parcalar.append("İşe baş: " + _temizle(dvo.get("iseBaslamaTarihi")))
                return " | ".join(p for p in parcalar if p)
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "kamuEmekli":
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "sskCalisani":
            parcalar = []
            ssk = dvo.get("sskBilgi") or {}
            liste = ssk.get("tescilKaydi4AList") or []
            if liste:
                kb = (liste[0] or {}).get("kimlikBilgileri") or {}
                sicil = _temizle(kb.get("sicilNo"))
                giris = _temizle(kb.get("iseGirisTarihi"))
                bilgi = "Tescil var"
                if sicil:
                    bilgi += f" (Sicil {sicil}" + (f", işe giriş {giris}" if giris else "") + ")"
                parcalar.append(bilgi)
            isy = (((dvo.get("isYeriBilgileri") or {}).get("data") or {}).get("sonuc") or {})
            if isy.get("mesaj"):
                parcalar.append(_temizle(isy.get("mesaj")))
            detay = dvo.get("sskBilgiDetay") or {}
            if "aktif" in detay:
                parcalar.append("Aktif: " + ("Evet" if detay.get("aktif") else "Hayır"))
            return " | ".join(parcalar) if parcalar else (_ilk_hata(dvo) or "Kayıt yok")

        if anahtar == "sskEmekli":
            liste = dvo.get("emekliKaydi4AList") or []
            if liste:
                return f"Emekli kaydı var ({len(liste)} kayıt)"
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "bagkurCalisani":
            liste = dvo.get("tescilKaydi4BList") or []
            if liste:
                t = liste[0] or {}
                parcalar = [_temizle(dvo.get("islemSonucAciklama")) or "Kayıt bulundu"]
                if _temizle(t.get("bagNo")):
                    parcalar.append("Bağ-No " + _temizle(t.get("bagNo")))
                if _temizle(t.get("terkAciklama")):
                    sterk = "Terk: " + _temizle(t.get("terkAciklama"))
                    if t.get("terkTarihi"):
                        sterk += f" ({t.get('terkTarihi')})"
                    parcalar.append(sterk)
                return " | ".join(p for p in parcalar if p)
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "bagkurEmekli":
            return _ilk_hata(dvo) or "Kayıt yok"

        if anahtar == "isYeri":
            return _ilk_hata(dvo) or "Kayıt yok"

        return _ilk_hata(dvo) or "Veri var"
    except Exception as e:
        return f"[özet hatası: {e}]"


def tam_metin(veri, girinti=0):
    """sorguSonucDVO'nun TAMAMINI okunabilir, girintili metne döker."""
    satirlar = []
    ond = "    " * girinti
    if isinstance(veri, dict):
        for k, v in veri.items():
            if k in NOISE_KEYS:
                continue
            if v is None or v == "" or v == [] or v == {}:
                continue
            if isinstance(v, (dict, list)):
                alt = tam_metin(v, girinti + 1)
                if alt.strip():
                    satirlar.append(f"{ond}{k}:")
                    satirlar.append(alt)
            else:
                satirlar.append(f"{ond}{k}: {str(v).strip()}")
    elif isinstance(veri, list):
        for i, e in enumerate(veri, 1):
            if isinstance(e, (dict, list)):
                alt = tam_metin(e, girinti + 1)
                if alt.strip():
                    satirlar.append(f"{ond}[{i}]")
                    satirlar.append(alt)
            elif e not in (None, ""):
                satirlar.append(f"{ond}- {str(e).strip()}")
    else:
        satirlar.append(f"{ond}{str(veri).strip()}")
    return "\n".join(satirlar)


def hucre_metni(anahtar, dvo):
    """Hücreye yazılacak nihai metin: ilk satır kısa özet + ayraç + TAM veri.
    SSK Çalışan için ilk satır 'Olumlu | ...' / 'Olumsuz | ...' ile başlar ve
    kayıt yoksa yalnızca mesaj yazılır (tam döküm eklenmez)."""
    if anahtar == "sskCalisani":
        durum, ozet, veri_var = degerlendir_ssk_calisani(dvo)
        bas = f"{durum} | {ozet}" if ozet else durum
        if not veri_var:
            return bas
        tam = tam_metin(dvo)
        return f"{bas}\n──────────\n{tam}" if tam.strip() else bas

    ozet = ozetle(anahtar, dvo)
    tam = tam_metin(dvo)
    if tam.strip() and tam.strip() != ozet.strip():
        return f"{ozet}\n──────────\n{tam}"
    return ozet


# --------------------------------------------------------------------------- #
#  Sorgu motoru (ayrı thread'de çalışır)
# --------------------------------------------------------------------------- #
class SorguMotoru:
    def __init__(self, log_fn):
        self.log = log_fn
        self.base = OFFICE_BASE

    def baslat(self):
        # Playwright/tarayıcı YOK. İstekler yerel ofis proxy'sine gider; ofis (uyap_app.py)
        # bunları canlı e-imza UYAP oturumuyla (headless) iletir. uyap_session.json kullanılmaz.
        self.log("🔌 Ofis bağlantısı kullanılacak (Playwright/JSON YOK): "
                 f"{self.base} üzerinden canlı UYAP oturumuna komut gönderilecek.")
        self.log("ℹ️ Bunun için 'UYAP Ağ Geçidi' (uyap_app.py) açık ve Paylaş/Al ile "
                 "bağlantı başlatılmış olmalı.")

    def yeniden_baglan(self):
        """Ofis modunda tarayıcı yoktur; UYAP oturumu ofiste tutulur ve gerekirse orada
        otomatik yenilenir. Burada yeniden başlatılacak bir tarayıcı yok."""
        self.log("🔄 UYAP oturumu ofiste otomatik yönetiliyor; kaldığı yerden devam ediliyor.")
        return True

    def _post(self, endpoint, payload, timeout=90, denemeler=3, max_yeniden=3):
        """Tek POST'u yerel ofis proxy'sine (127.0.0.1:8800) gönderir; ofis isteği canlı
        UYAP oturumuyla iletir. Tarayıcı/JSON oturum YOK. timeout SANİYE cinsindendir."""
        url = f"{self.base}/{endpoint}"
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        son_hata = None
        for d in range(denemeler):
            try:
                req = urllib.request.Request(url, data=body, method="POST")
                for k, v in COMMON_HEADERS.items():
                    req.add_header(k, v)
                with urllib.request.urlopen(req, timeout=timeout) as r:
                    status = r.status
                    text = r.read().decode("utf-8", "replace")
                # Ofis canlı oturumu kendi yeniler; yine de login duvarı/HTML gelirse işaretle.
                isaret = text.lstrip()[:300].lower()
                if status in (401, 403) or isaret.startswith("<!doctype") \
                        or isaret.startswith("<html") or "<title>uyap" in isaret:
                    raise OturumHatasi(f"oturum/yetki hatası (HTTP {status}; login sayfası döndü)")
                try:
                    return status, json.loads(text)
                except Exception:
                    return status, text
            except OturumHatasi:
                raise
            except urllib.error.HTTPError as e:
                son_hata = e
                detay = ""
                try:
                    detay = e.read().decode("utf-8", "replace")[:200]
                except Exception:
                    pass
                self.log(f"   ⚠️ {endpoint} deneme {d + 1}/{denemeler}: HTTP {e.code} {detay}")
                if e.code in (401, 403):
                    raise OturumHatasi(f"ofis yetki hatası (HTTP {e.code})")
            except urllib.error.URLError as e:
                son_hata = e
                self.log(f"   ⚠️ {endpoint} deneme {d + 1}/{denemeler}: ofise ulaşılamadı "
                         f"({getattr(e, 'reason', e)}). uyap_app.py açık ve Paylaş/Al aktif mi?")
            except Exception as e:
                son_hata = e
                self.log(f"   ⚠️ {endpoint} deneme {d + 1}/{denemeler} başarısız: {e}")
            if d < denemeler - 1:
                time.sleep(3 * (d + 1))  # 3s, 6s ... artan bekleme
        raise son_hata

    def dosya_id_bul(self, yil, sira):
        payload = {
            "dosyaDurumKod": 0, "pageSize": 500, "pageNumber": 1,
            "dosyaYil": int(yil), "dosyaSira": int(sira),
            "birimId": "", "birimTuru2": BIRIM_TURU2, "birimTuru3": BIRIM_TURU3,
        }
        _, veri = self._post("search_phrase_detayli.ajx", payload)
        try:
            kayitlar = veri[0]
            if kayitlar:
                return kayitlar[0].get("dosyaId"), kayitlar[0].get("dosyaNo")
        except Exception:
            pass
        return None, None

    def borclular(self, dosya_id):
        _, veri = self._post("dosya_borclu_list.ajx", {"dosyaId": dosya_id})
        sonuc = []
        if isinstance(veri, list):
            for b in veri:
                kb = (b or {}).get("kisiTumDVO") or {}
                sonuc.append({
                    "kisiKurumId": b.get("kisiKurumId"),
                    "adi": _temizle(kb.get("adi")),
                    "soyadi": _temizle(kb.get("soyadi")),
                    "tc": _temizle(kb.get("tcKimlikNo")),
                })
        return sonuc

    def borclu_sec(self, borclular, isim):
        if not borclular:
            return None, False
        if len(borclular) == 1:
            return borclular[0], True
        hedef = tr_normalize(isim)
        if hedef:
            for b in borclular:
                tam = tr_normalize(b["adi"] + " " + b["soyadi"])
                if tam and (tam == hedef or tam in hedef or hedef in tam):
                    return b, True
            hedef_kelime = set(hedef.split())
            for b in borclular:
                tam_kelime = set(tr_normalize(b["adi"] + " " + b["soyadi"]).split())
                if tam_kelime and tam_kelime <= hedef_kelime:
                    return b, True
        return borclular[0], False

    def sgk_sorgula(self, dosya_id, kisi_kurum_id, anahtarlar, kontrol):
        """Sadece 'anahtarlar' içindekileri sorgular.
        Döner: {anahtar: (basarili_mi, metin)}.  kontrol() False ise yarıda keser."""
        sonuc = {}
        payload = {"dosyaId": dosya_id, "kisiKurumId": kisi_kurum_id}
        for _, endpoint, anahtar in SORGULAR:
            if anahtar not in anahtarlar:
                continue
            if not kontrol():
                break
            try:
                durum, veri = self._post(endpoint, payload)
                if isinstance(veri, dict):
                    dvo = veri.get("sorguSonucDVO", veri)
                    sonuc[anahtar] = (True, hucre_metni(anahtar, dvo))
                else:
                    sonuc[anahtar] = (False, f"{HATA_ON} HTTP {durum}")
            except Exception as e:
                sonuc[anahtar] = (False, f"{HATA_ON} {e}")
            time.sleep(SORGU_ARASI_BEKLEME)
        return sonuc

    def kapat(self):
        pass  # ofis modunda kapatılacak tarayıcı/oturum yok


# --------------------------------------------------------------------------- #
#  GUI
# --------------------------------------------------------------------------- #
class SgkSorguGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("SGK Toplu Sorgu")
        self.root.geometry("1400x780")

        self.kaynak_yolu = None   # kullanıcının seçtiği dosya (DEĞİŞTİRİLMEZ; yalnızca tohum)
        self.excel_yolu = None    # "yapılanlar" ana dosyası — tüm sonuçlar buraya yazılır
        self.ozet_yolu = None     # "yapılanlar özet" (kısa) dosyası
        self.wb = None
        self.ws = None            # bellekteki ANA tablo (tüm satırlar, sabit sütunlar, başlıklı)
        self.satir_sayisi = 0

        # Filtre/sıralama durumu
        self.filtreler = []           # eklenen (yığılan) filtreler: [{kol, icerir, icermez}]
        self.sadece_olumlu = False    # "Sadece SGK olumluları göster" açık mı
        self._sirali_kol = None
        self._sirali_ters = False

        self.satir_id = {}      # excel_satir_no -> tree iid
        self.iid_satir = {}     # tree iid -> excel_satir_no
        self.tum_satirlar = []  # [(iid, excel_r)] (filtre için orijinal sıra)

        self.calisiyor = False
        self.pause_event = threading.Event()
        self.stop_event = threading.Event()
        self.kuyruk = queue.Queue()

        self._sag_tik_hucre = None  # (iid, kol_index)

        self._arayuz_kur()
        self.root.after(150, self._kuyruk_isle)

    # ---- Arayüz ---------------------------------------------------------- #
    def _arayuz_kur(self):
        # Üst araç çubuğu
        ust = ttk.Frame(self.root, padding=8)
        ust.pack(fill="x")
        ttk.Button(ust, text="📂 Excel Yükle", command=self.excel_yukle).pack(side="left")
        self.sorgula_btn = ttk.Button(ust, text="▶ Sorgula", command=lambda: self.sorgu_baslat("normal"), state="disabled")
        self.sorgula_btn.pack(side="left", padx=4)
        self.retry_btn = ttk.Button(ust, text="↻ Hatalıları Tekrar Dene", command=lambda: self.sorgu_baslat("retry"), state="disabled")
        self.retry_btn.pack(side="left", padx=4)
        self.duraklat_btn = ttk.Button(ust, text="⏸ Duraklat", command=self.duraklat_toggle, state="disabled")
        self.duraklat_btn.pack(side="left", padx=4)
        self.durdur_btn = ttk.Button(ust, text="⏹ Durdur", command=self.durdur, state="disabled")
        self.durdur_btn.pack(side="left", padx=4)
        self.dosya_lbl = ttk.Label(ust, text="Dosya seçilmedi")
        self.dosya_lbl.pack(side="left", padx=10)
        self.durum_lbl = ttk.Label(ust, text="")
        self.durum_lbl.pack(side="right")

        # Sorgulanacak sütunlar (checkbox)
        sb = ttk.LabelFrame(self.root, text="Sorgulanacak sütunlar", padding=(8, 2))
        sb.pack(fill="x", padx=8, pady=(4, 0))
        self.secili_sorgu = {}
        for baslik, _, anahtar in SORGULAR:
            var = tk.BooleanVar(value=True)
            self.secili_sorgu[anahtar] = var
            ttk.Checkbutton(sb, text=baslik, variable=var).pack(side="left", padx=4)
        ttk.Button(sb, text="Tümünü Seç", command=lambda: self._tum_sorgu(True)).pack(side="right", padx=2)
        ttk.Button(sb, text="Tümünü Kaldır", command=lambda: self._tum_sorgu(False)).pack(side="right", padx=2)

        # Filtre çubuğu
        fb = ttk.Frame(self.root, padding=(8, 0))
        fb.pack(fill="x")
        ttk.Label(fb, text="Sütun:").pack(side="left")
        self.filtre_kol = ttk.Combobox(fb, state="readonly", width=16, values=["Tüm Sütunlar"])
        self.filtre_kol.set("Tüm Sütunlar")
        self.filtre_kol.pack(side="left", padx=4)
        ttk.Label(fb, text="İçerir:").pack(side="left")
        self.filtre_txt = ttk.Entry(fb, width=20)
        self.filtre_txt.pack(side="left", padx=4)
        self.filtre_txt.bind("<Return>", lambda e: self.filtrele())
        ttk.Label(fb, text="İçermez:").pack(side="left")
        self.filtre_haric_txt = ttk.Entry(fb, width=20)
        self.filtre_haric_txt.pack(side="left", padx=4)
        self.filtre_haric_txt.bind("<Return>", lambda e: self.filtrele())
        ttk.Button(fb, text="Filtrele", command=self.filtrele).pack(side="left", padx=2)
        ttk.Button(fb, text="➕ Filtre Ekle", command=self.filtre_ekle).pack(side="left", padx=2)
        ttk.Button(fb, text="Temizle", command=self.filtre_temizle).pack(side="left", padx=2)
        self.olumlu_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(fb, text="Sadece SGK olumluları göster",
                        variable=self.olumlu_var, command=self.olumlu_toggle).pack(side="left", padx=10)

        # İkinci satır: eklenen (yığılan) filtreler + sonuç bilgisi
        fb2 = ttk.Frame(self.root, padding=(8, 0))
        fb2.pack(fill="x")
        ttk.Label(fb2, text="Aktif filtreler:").pack(side="left")
        self.filtre_liste_lbl = ttk.Label(fb2, text="(yok)", foreground="#444")
        self.filtre_liste_lbl.pack(side="left", padx=4)
        self.filtre_bilgi = ttk.Label(fb2, text="")
        self.filtre_bilgi.pack(side="right", padx=8)

        # Orta: sol tablo + sağ detay paneli
        pan = ttk.PanedWindow(self.root, orient="horizontal")
        pan.pack(fill="both", expand=True, padx=8, pady=4)

        sol = ttk.Frame(pan)
        self.kolonlar = ["No", "Ad Soyad", "Birim", "Dosya No"] + [s[0] for s in SORGULAR]
        self.tree = ttk.Treeview(sol, columns=self.kolonlar, show="headings")
        for k in self.kolonlar:
            self.tree.heading(k, text=k, command=lambda c=k: self._sirala(c))
            genislik = 200 if k in ("Ad Soyad", "Birim") else 150
            if k == "No":
                genislik = 55
            elif k == "Dosya No":
                genislik = 95
            self.tree.column(k, width=genislik, anchor="w")
        ydk = ttk.Scrollbar(sol, orient="vertical", command=self.tree.yview)
        xdk = ttk.Scrollbar(sol, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ydk.set, xscrollcommand=xdk.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        ydk.grid(row=0, column=1, sticky="ns")
        xdk.grid(row=1, column=0, sticky="ew")
        sol.rowconfigure(0, weight=1)
        sol.columnconfigure(0, weight=1)
        pan.add(sol, weight=4)

        sag = ttk.Frame(pan)
        ttk.Label(sag, text="Hücre Detayı (tıklayın)").pack(anchor="w")
        self.detay_txt = tk.Text(sag, width=46, wrap="word", state="disabled")
        dydk = ttk.Scrollbar(sag, orient="vertical", command=self.detay_txt.yview)
        self.detay_txt.configure(yscrollcommand=dydk.set)
        self.detay_txt.pack(side="left", fill="both", expand=True)
        dydk.pack(side="left", fill="y")
        pan.add(sag, weight=2)

        # Tamamlanıp arşive taşınan satırlar yeşil gösterilir (tablodan silinmez)
        self.tree.tag_configure("tamam", background="#e6ffe6")
        # Şirket (tüzel kişi) satırları gri gösterilir — sorgulanmadan taşınır
        self.tree.tag_configure("sirket", background="#eeeeee")

        # Tree olayları
        self.tree.bind("<ButtonRelease-1>", self.hucre_tikla)   # tek tık -> detay
        self.tree.bind("<Button-3>", self.sag_tik_menu)         # sağ tık -> kopyala

        # Sağ tık menüsü
        self.menu = tk.Menu(self.root, tearoff=0)
        self.menu.add_command(label="Hücre değerini kopyala", command=self.kopyala_hucre)
        self.menu.add_command(label="Sütunu kopyala (görünen satırlar)", command=self.kopyala_sutun)

        # Alt log
        self.log_txt = tk.Text(self.root, height=7, state="disabled", wrap="word")
        self.log_txt.pack(fill="x", padx=8, pady=(0, 8))

    def log(self, mesaj):
        self.kuyruk.put(("log", mesaj))

    def _log_yaz(self, mesaj):
        self.log_txt.configure(state="normal")
        self.log_txt.insert("end", mesaj + "\n")
        self.log_txt.see("end")
        self.log_txt.configure(state="disabled")

    # ---- Yardımcılar: hücre durumu -------------------------------------- #
    @staticmethod
    def _bos(v):
        return v is None or str(v).strip() == ""

    def _hata_cell(self, v):
        return (not self._bos(v)) and str(v).lstrip().startswith(HATA_ON)

    def _tum_sorgu(self, deger):
        for var in self.secili_sorgu.values():
            var.set(deger)

    def gereken_anahtarlar(self, r, mode):
        """Bu satırda hangi sorguların çalışması gerektiğini belirler.
        Yalnızca checkbox ile seçili sütunlar dikkate alınır."""
        gerekli = set()
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if not self.secili_sorgu[anahtar].get():
                continue  # bu sütun seçili değil
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if mode == "retry":
                if self._hata_cell(v):
                    gerekli.add(anahtar)
            else:  # normal: boş veya hatalı olanlar
                if self._bos(v) or self._hata_cell(v):
                    gerekli.add(anahtar)
        return gerekli

    @staticmethod
    def _sirket_mi(isim):
        """İsim tüzel kişi (ŞTİ / Şirketi / Ltd) mi? -> SGK sorgusu yapılmaz."""
        if not isim:
            return False
        kelimeler = set(tr_normalize(isim).split())
        return bool(kelimeler & SIRKET_ANAHTARLARI)

    def _satir_tamamlandi_mi(self, r):
        """Seçili tüm sorgu sütunları başarılı mı (dolu ve [HATA] değil)?"""
        secili_var = False
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if not self.secili_sorgu[anahtar].get():
                continue
            secili_var = True
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if self._bos(v) or self._hata_cell(v):
                return False
        return secili_var

    @staticmethod
    def _kisa(metin):
        """Tabloda göstermek için tek satıra indir."""
        if metin is None:
            return ""
        t = str(metin).split("\n", 1)[0]
        return (t[:90] + "…") if len(t) > 90 else t

    # ---- Excel ----------------------------------------------------------- #
    def excel_yukle(self):
        if self.calisiyor:
            messagebox.showwarning("Çalışıyor", "Önce sorguyu durdurun.")
            return
        yol = filedialog.askopenfilename(
            title="Excel dosyası seç",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Tümü", "*.*")],
        )
        if not yol:
            return
        # Yapılanlar (ana) ve özet dosya yolları: "X.xlsx" -> "X_yapilanlar.xlsx" / "X_yapilanlar_ozet.xlsx"
        kok, uzanti = os.path.splitext(yol)
        # Kullanıcı yanlışlıkla _yapilanlar/_ozet dosyasını seçtiyse kök adını çıkar
        for ek in (OZET_EKI, YAPILANLAR_EKI):
            if kok.endswith(ek):
                kok = kok[: -len(ek)]
                break
        cikti_yolu = kok + YAPILANLAR_EKI + uzanti
        ozet_yolu = kok + OZET_EKI + uzanti
        try:
            # Daha önce çalışılmışsa yapılanlar dosyasından DEVAM et (tüm satır+sonuçlar onda).
            # Yoksa seçilen dosyayı tohum olarak yükle. Seçilen dosya hiç değiştirilmez.
            if os.path.exists(cikti_yolu):
                self.wb = openpyxl.load_workbook(cikti_yolu)
                self._log_yaz(f"↩️ Yapılanlar dosyasından devam ediliyor: {os.path.basename(cikti_yolu)}")
            else:
                self.wb = openpyxl.load_workbook(yol)
            self.ws = self.wb.active
        except Exception as e:
            messagebox.showerror("Hata", f"Excel açılamadı:\n{e}")
            return
        # Başlık satırı yoksa ekle (en üst satır). Böylece veri her zaman 2. satırdan başlar.
        if not self._baslik_satiri_mi(self.ws):
            self.ws.insert_rows(1)
            for c, ad in enumerate(self.kolonlar, start=1):
                self.ws.cell(1, c, value=ad)

        self.kaynak_yolu = yol
        self.excel_yolu = cikti_yolu
        self.ozet_yolu = ozet_yolu
        self.dosya_lbl.config(text=f"{os.path.basename(yol)}  →  {os.path.basename(cikti_yolu)}")
        self._tabloyu_doldur()
        self.filtre_kol.config(values=["Tüm Sütunlar"] + self.kolonlar)
        self.sorgula_btn.config(state="normal")
        self.retry_btn.config(state="normal")
        self._kaydet()   # başlıkları/biçimi hemen yapılanlar dosyasına yaz
        self._log_yaz(f"📂 Yüklendi: {os.path.basename(yol)}  ({self.satir_sayisi} satır)")
        self._log_yaz(f"💾 Tam veri → {os.path.basename(cikti_yolu)}")
        self._log_yaz(f"💾 Özet → {os.path.basename(ozet_yolu)}")

    def _baslik_satiri_mi(self, ws):
        """1. satır bizim başlık satırımız mı? (A=No, B=Ad Soyad, ... D=Dosya No)"""
        try:
            deg = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, 5)]
            return deg == [b.lower() for b in self.kolonlar[:4]]
        except Exception:
            return False

    def _tabloyu_doldur(self):
        self.tree.delete(*self.tree.get_children())
        self.satir_id, self.iid_satir, self.tum_satirlar = {}, {}, []
        self.satir_sayisi = 0
        # 1. satır başlıktır; veri 2. satırdan başlar
        for r in range(2, self.ws.max_row + 1):
            temel = [self.ws.cell(r, c).value for c in range(1, 5)]
            if all(v is None for v in temel):
                continue
            sonuc = [self._kisa(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                     for i in range(len(SORGULAR))]
            iid = self.tree.insert("", "end", values=temel + sonuc)
            self.satir_id[r] = iid
            self.iid_satir[iid] = r
            self.tum_satirlar.append((iid, r))
            self.satir_sayisi += 1

    # ---- Filtreleme ------------------------------------------------------ #
    def _filtre_haystack(self, iid, kol):
        """Filtre için aranacak metin: hücrenin TAM (kısaltılmamış) içeriği."""
        if kol == "Tüm Sütunlar":
            parcalar = [self._hucre_tam_deger(iid, i) for i in range(len(self.kolonlar))]
            return " ".join(parcalar).lower()
        idx = self.kolonlar.index(kol)
        return self._hucre_tam_deger(iid, idx).lower()

    def _mevcut_kriter(self):
        """Filtre kutularındaki anlık kriter (boşsa None)."""
        icerir = self.filtre_txt.get().strip().lower()
        icermez = self.filtre_haric_txt.get().strip().lower()
        if not icerir and not icermez:
            return None
        return {"kol": self.filtre_kol.get(), "icerir": icerir, "icermez": icermez}

    def _kriter_uyuyor(self, iid, kriter):
        hay = self._filtre_haystack(iid, kriter["kol"])
        if kriter["icerir"] and kriter["icerir"] not in hay:
            return False
        if kriter["icermez"] and kriter["icermez"] in hay:
            return False
        return True

    def _aktif_kriterler(self):
        """Eklenen (yığılan) filtreler + kutudaki anlık kriter."""
        kriterler = list(self.filtreler)
        anlik = self._mevcut_kriter()
        if anlik:
            kriterler.append(anlik)
        return kriterler

    def _satir_olumlu_mu(self, iid):
        """Seçili SGK sütunlarından en az biri 'Olumlu' ile başlıyor mu?"""
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if not self.secili_sorgu[anahtar].get():
                continue
            v = self._hucre_tam_deger(iid, 4 + i).strip().lower()
            if v.startswith("olumlu"):
                return True
        return False

    def _satir_gorunur_mu(self, iid):
        """Tüm aktif filtre + olumlu kriterine göre satır görünür olmalı mı?"""
        if self.sadece_olumlu and not self._satir_olumlu_mu(iid):
            return False
        return all(self._kriter_uyuyor(iid, k) for k in self._aktif_kriterler())

    def _satiri_filtrele(self, iid):
        """Tek satırı aktif filtrelere göre göster/gizle (canlı güncelleme için)."""
        uyar = self._satir_gorunur_mu(iid)
        if uyar:
            if iid not in self.tree.get_children(""):
                self.tree.reattach(iid, "", "end")
        else:
            self.tree.detach(iid)
        return uyar

    def _filtre_etiketi(self):
        """İkinci satırdaki 'aktif filtreler' etiketini günceller."""
        parcalar = []
        for k in self.filtreler:
            p = k["kol"]
            if k["icerir"]:
                p += f" ⊃'{k['icerir']}'"
            if k["icermez"]:
                p += f" ⊅'{k['icermez']}'"
            parcalar.append(p)
        if self.sadece_olumlu:
            parcalar.append("SGK Olumlu")
        self.filtre_liste_lbl.config(text="  •  ".join(parcalar) if parcalar else "(yok)")

    def filtre_ekle(self):
        """Kutudaki kriteri yığına ekler (birden fazla filtre birlikte uygulanır)."""
        anlik = self._mevcut_kriter()
        if not anlik:
            messagebox.showinfo("Bilgi", "Önce İçerir/İçermez kutusuna bir şey yazın.")
            return
        self.filtreler.append(anlik)
        self.filtre_txt.delete(0, "end")
        self.filtre_haric_txt.delete(0, "end")
        self._filtre_etiketi()
        self.filtrele()

    def olumlu_toggle(self):
        self.sadece_olumlu = bool(self.olumlu_var.get())
        self._filtre_etiketi()
        self.filtrele()

    def filtrele(self):
        if not self.tum_satirlar:
            return
        self._filtre_etiketi()
        for iid, _ in self.tum_satirlar:
            self.tree.detach(iid)
        gosterilen = 0
        for iid, _ in self.tum_satirlar:
            if self._satir_gorunur_mu(iid):
                self.tree.reattach(iid, "", "end")
                gosterilen += 1
        if self._aktif_kriterler() or self.sadece_olumlu:
            self.filtre_bilgi.config(text=f"{gosterilen} / {len(self.tum_satirlar)} satır")
        else:
            self.filtre_bilgi.config(text="")

    def filtre_temizle(self):
        self.filtre_txt.delete(0, "end")
        self.filtre_haric_txt.delete(0, "end")
        self.filtre_kol.set("Tüm Sütunlar")
        self.filtreler = []
        self.sadece_olumlu = False
        self.olumlu_var.set(False)
        self.filtre_bilgi.config(text="")
        self._filtre_etiketi()
        for iid, _ in self.tum_satirlar:
            self.tree.reattach(iid, "", "end")

    # ---- Sıralama (başlığa tıklayınca) ---------------------------------- #
    def _sirala(self, kol):
        if not self.tum_satirlar:
            return
        ters = (self._sirali_kol == kol) and (not self._sirali_ters)
        self._sirali_kol = kol
        self._sirali_ters = ters
        idx = self.kolonlar.index(kol)

        def anahtar(iid):
            v = self._hucre_tam_deger(iid, idx)
            if kol == "No":
                try:
                    return (0, float(str(v).replace(",", ".")))
                except (TypeError, ValueError):
                    return (1, str(v).lower())
            return str(v).lower()

        cocuklar = list(self.tree.get_children(""))   # yalnızca görünenler
        try:
            cocuklar.sort(key=anahtar, reverse=ters)
        except TypeError:
            cocuklar.sort(key=lambda i: str(anahtar(i)), reverse=ters)
        for i, iid in enumerate(cocuklar):
            self.tree.move(iid, "", i)
        for k in self.kolonlar:
            ok = (" ▼" if ters else " ▲") if k == kol else ""
            self.tree.heading(k, text=k + ok)

    # ---- Tıklama / kopyalama -------------------------------------------- #
    def _hedef_hucre(self, event):
        iid = self.tree.identify_row(event.y)
        kol = self.tree.identify_column(event.x)  # '#1'..
        if not iid or not kol:
            return None, None
        try:
            kol_index = int(kol.replace("#", "")) - 1
        except ValueError:
            return None, None
        return iid, kol_index

    def _hucre_tam_deger(self, iid, kol_index):
        """Detay/kopyalama için tam (kısaltılmamış) değer."""
        r = self.iid_satir.get(iid)
        if r is None or kol_index < 0:
            return ""
        if kol_index >= 4:  # sonuç sütunu -> Excel'deki tam metin
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + (kol_index - 4)).value
            return "" if v is None else str(v)
        v = self.ws.cell(r, kol_index + 1).value
        return "" if v is None else str(v)

    def hucre_tikla(self, event):
        iid, kol_index = self._hedef_hucre(event)
        if iid is None:
            return
        baslik = self.kolonlar[kol_index] if kol_index < len(self.kolonlar) else ""
        deger = self._hucre_tam_deger(iid, kol_index)
        self.detay_txt.configure(state="normal")
        self.detay_txt.delete("1.0", "end")
        self.detay_txt.insert("1.0", f"【 {baslik} 】\n\n{deger}")
        self.detay_txt.configure(state="disabled")

    def sag_tik_menu(self, event):
        iid, kol_index = self._hedef_hucre(event)
        if iid is None:
            return
        self.tree.selection_set(iid)
        self._sag_tik_hucre = (iid, kol_index)
        self.menu.tk_popup(event.x_root, event.y_root)

    def _panoya(self, metin):
        self.root.clipboard_clear()
        self.root.clipboard_append(metin)

    def kopyala_hucre(self):
        if not self._sag_tik_hucre:
            return
        iid, kol_index = self._sag_tik_hucre
        self._panoya(self._hucre_tam_deger(iid, kol_index))
        self.durum_lbl.config(text="Hücre kopyalandı")

    def kopyala_sutun(self):
        if not self._sag_tik_hucre:
            return
        _, kol_index = self._sag_tik_hucre
        satirlar = []
        for iid in self.tree.get_children(""):   # yalnızca görünen (filtreli) satırlar
            satirlar.append(self._hucre_tam_deger(iid, kol_index))
        self._panoya("\n".join(satirlar))
        self.durum_lbl.config(text=f"Sütun kopyalandı ({len(satirlar)} satır)")

    # ---- Sorgu kontrolleri ---------------------------------------------- #
    def sorgu_baslat(self, mode):
        if self.calisiyor or not self.ws:
            return
        # retry modunda hata var mı?
        if mode == "retry":
            toplam_hata = sum(len(self.gereken_anahtarlar(r, "retry")) for r in self.satir_id)
            if toplam_hata == 0:
                messagebox.showinfo("Bilgi", "Tekrar denenecek hatalı hücre yok.")
                return
        self.calisiyor = True
        self.stop_event.clear()
        self.pause_event.clear()
        self._butonlar(calisiyor=True)
        threading.Thread(target=self._sorgu_worker, args=(mode,), daemon=True).start()

    def duraklat_toggle(self):
        if not self.calisiyor:
            return
        if self.pause_event.is_set():
            self.pause_event.clear()
            self.duraklat_btn.config(text="⏸ Duraklat")
        else:
            self.pause_event.set()
            self.duraklat_btn.config(text="▶ Devam")

    def durdur(self):
        if not self.calisiyor:
            return
        self.stop_event.set()
        self.pause_event.clear()  # duraklatılmışsa çıkabilsin
        self.durum_lbl.config(text="Durduruluyor...")

    def _butonlar(self, calisiyor):
        if calisiyor:
            self.sorgula_btn.config(state="disabled")
            self.retry_btn.config(state="disabled")
            self.duraklat_btn.config(state="normal", text="⏸ Duraklat")
            self.durdur_btn.config(state="normal")
        else:
            self.sorgula_btn.config(state="normal")
            self.retry_btn.config(state="normal")
            self.duraklat_btn.config(state="disabled", text="⏸ Duraklat")
            self.durdur_btn.config(state="disabled")

    def _bekle_devam(self):
        """Duraklatma boyunca bekler; durdurulduysa False döner."""
        while self.pause_event.is_set() and not self.stop_event.is_set():
            self.kuyruk.put(("durum", "⏸ Duraklatıldı"))
            time.sleep(0.3)
        return not self.stop_event.is_set()

    # ---- Worker ---------------------------------------------------------- #
    def _sorgu_worker(self, mode):
        motor = SorguMotoru(self.log)
        try:
            motor.baslat()
        except Exception as e:
            self.log(f"❌ Başlatma hatası: {e}")
            motor.kapat()
            self.kuyruk.put(("bitti", None))
            return

        satirlar = list(self.satir_id.keys())
        toplam = len(satirlar)
        offset = {anahtar: i for i, (_, _, anahtar) in enumerate(SORGULAR)}
        hata_sayaci = 0       # son moladan bu yana biriken hata sayısı
        pespese = 0           # ardışık başarısızlık (dosya bulunamadı / hata)
        oturum_yenilendi = False  # bu başarısızlık serisinde bir kez yenilendi mi

        def oturum_yenile():
            """Oturumu bir kez yeniden başlatır. Dönüş: başarılı mı."""
            nonlocal oturum_yenilendi, pespese
            oturum_yenilendi = True
            pespese = 0
            try:
                motor.yeniden_baglan()
                return True
            except Exception as e:
                self.log(f"❌ Oturum yenilenemedi: {e}")
                return False

        def mola_kontrol():
            """Hata eşiği aşıldıysa kısa mola ver. Stop ise False döner."""
            nonlocal hata_sayaci
            if hata_sayaci >= MOLA_HATA_ESIGI:
                hata_sayaci = 0
                self.log(f"⏳ {MOLA_HATA_ESIGI} hata oluştu, sunucu yükünü azaltmak için "
                         f"{MOLA_SURESI} sn mola veriliyor...")
                bekleyen = MOLA_SURESI
                while bekleyen > 0 and not self.stop_event.is_set():
                    self.kuyruk.put(("durum", f"⏳ Otomatik mola: {bekleyen} sn"))
                    time.sleep(1)
                    bekleyen -= 1
                if not self.stop_event.is_set():
                    self.log("▶ Molaya devam ediliyor.")
            return not self.stop_event.is_set()

        for idx, r in enumerate(satirlar, 1):
            if not self._bekle_devam():
                break
            if not mola_kontrol():
                break
            self.kuyruk.put(("durum", f"İşleniyor {idx}/{toplam}  (satır {r})"))

            gereken = self.gereken_anahtarlar(r, mode)
            if not gereken:
                continue  # bu satır zaten tamamlanmış

            d = self.ws.cell(r, 4).value
            isim = self.ws.cell(r, 2).value
            if self._sirket_mi(isim):
                self.log(f"satır {r}: '{isim}' şirket (tüzel kişi) — sorgu yapılmadı, taşınıyor.")
                self.kuyruk.put(("sirket", r))
                continue
            if not d or "/" not in str(d):
                self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} Geçersiz dosya no: {d}")
                hata_sayaci += 1
                continue

            yil, _, sira = str(d).partition("/")
            try:
                dosya_id, _ = motor.dosya_id_bul(yil.strip(), sira.strip())
            except OturumHatasi as e:
                self.log(f"⛔ Oturum hatası: {e}")
                if not oturum_yenilendi and oturum_yenile():
                    continue  # bu satır boş kalsın, sonra tekrar denenir
                self.kuyruk.put(("oturum_dur", None))
                self.stop_event.set()
                break
            except Exception as e:
                self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} arama: {e}")
                hata_sayaci += 1
                pespese += 1
                if pespese >= OTURUM_KONTROL_ESIGI and not oturum_yenilendi:
                    self.log(f"⚠️ {OTURUM_KONTROL_ESIGI} ardışık hata — oturum kontrol ediliyor...")
                    oturum_yenile()
                continue
            if not dosya_id:
                self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} Dosya bulunamadı")
                self.log(f"satır {r}: {d} -> dosya bulunamadı")
                hata_sayaci += 1
                pespese += 1
                if pespese >= OTURUM_KONTROL_ESIGI and not oturum_yenilendi:
                    self.log(f"⚠️ {OTURUM_KONTROL_ESIGI} ardışık 'dosya bulunamadı' — "
                             "oturum kopmuş olabilir, yenileniyor...")
                    oturum_yenile()
                continue

            try:
                borclular = motor.borclular(dosya_id)
            except OturumHatasi as e:
                self.log(f"⛔ Oturum hatası: {e}")
                if not oturum_yenilendi and oturum_yenile():
                    continue
                self.kuyruk.put(("oturum_dur", None))
                self.stop_event.set()
                break
            except Exception as e:
                self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} borçlu: {e}")
                hata_sayaci += 1
                continue
            borclu, eslesti = motor.borclu_sec(borclular, isim)
            if not borclu:
                self._yaz_hepsi(r, gereken, offset, f"{HATA_ON} Borçlu bulunamadı")
                hata_sayaci += 1
                continue

            # Buraya geldiyse oturum çalışıyor; ardışık başarısızlık sayacını sıfırla
            pespese = 0
            oturum_yenilendi = False

            self.log(f"satır {r}: {d} -> {borclu['adi']} {borclu['soyadi']}"
                     + ("" if eslesti else "  (⚠ isim eşleşmedi, ilk borçlu)"))

            ozetler = motor.sgk_sorgula(dosya_id, borclu["kisiKurumId"], gereken, self._bekle_devam)

            sparse = {}
            for anahtar, (ok, metin) in ozetler.items():
                sparse[offset[anahtar]] = metin
                if not ok:
                    hata_sayaci += 1
            if sparse:
                self.kuyruk.put(("sonuc", (r, sparse)))

            if self.stop_event.is_set():
                break
            if idx % 10 == 0:
                self.kuyruk.put(("kaydet", None))
            time.sleep(SATIR_ARASI_BEKLEME)

        self.kuyruk.put(("kaydet", None))
        if self.stop_event.is_set():
            self.kuyruk.put(("durum", "Durduruldu"))
            self.log("⏹ Durduruldu.")
        else:
            self.kuyruk.put(("durum", "Tamamlandı"))
            self.log("🎉 İşlem tamamlandı.")
        motor.kapat()
        self.kuyruk.put(("bitti", None))

    def _yaz_hepsi(self, r, gereken, offset, metin):
        """Satır düzeyi hata: gereken tüm hücrelere aynı hatayı yaz."""
        sparse = {offset[a]: metin for a in gereken}
        self.kuyruk.put(("sonuc", (r, sparse)))

    # ---- Kuyruk işleme (ana thread) ------------------------------------- #
    def _kuyruk_isle(self):
        try:
            while True:
                tur, veri = self.kuyruk.get_nowait()
                if tur == "log":
                    self._log_yaz(veri)
                elif tur == "durum":
                    self.durum_lbl.config(text=veri)
                elif tur == "sonuc":
                    excel_satir, sparse = veri
                    iid = self.satir_id.get(excel_satir)
                    mevcut = list(self.tree.item(iid, "values")) if iid else None
                    for off, metin in sparse.items():
                        self.ws.cell(excel_satir, SONUC_BASLANGIC_KOL + off, value=metin)
                        if mevcut is not None:
                            mevcut[4 + off] = self._kisa(metin)
                    if iid and mevcut is not None:
                        # Seçili tüm sorgular hatasız bittiyse yeşil işaretle (satır yerinde kalır)
                        etiket = ("tamam",) if self._satir_tamamlandi_mi(excel_satir) else ()
                        self.tree.item(iid, values=mevcut, tags=etiket)
                        # Aktif filtre varsa satırı yeni içeriğine göre tekrar değerlendir
                        if self._satiri_filtrele(iid):
                            self.tree.see(iid)
                elif tur == "sirket":
                    self._sirketi_isaretle(veri)
                elif tur == "oturum_dur":
                    self.durum_lbl.config(text="⛔ Ofis bağlantısı yok")
                    messagebox.showwarning(
                        "Ofis bağlantısı yok",
                        "Yerel ofis bağlantısına (127.0.0.1:8800) ulaşılamadı ya da yetki reddedildi.\n\n"
                        "'UYAP Ağ Geçidi' (uyap_app.py) açık ve 'Paylaş' (ya da 'Al') ile bağlantı "
                        "başlatılmış olmalı. Bağlantıyı başlatıp tekrar ▶ Sorgula'ya basın.\n\n"
                        "İşlem durduruldu; tamamlanan satırlar kaydedildi.")
                elif tur == "kaydet":
                    self._kaydet()
                elif tur == "bitti":
                    self.calisiyor = False
                    self._butonlar(calisiyor=False)
        except queue.Empty:
            pass
        self.root.after(150, self._kuyruk_isle)

    def _sirketi_isaretle(self, r):
        """Tüzel kişi satırı: seçili sütunlara 'ŞİRKET' notu yazar, gri işaretler.
        Satır yerinde kalır; kaydetmede ana dosyaya yazılır (kaydırma yok)."""
        iid = self.satir_id.get(r)
        mevcut = list(self.tree.item(iid, "values")) if iid else None
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if not self.secili_sorgu[anahtar].get():
                continue
            self.ws.cell(r, SONUC_BASLANGIC_KOL + i, value=SIRKET_NOT)
            if mevcut is not None:
                mevcut[4 + i] = self._kisa(SIRKET_NOT)
        if iid and mevcut is not None:
            self.tree.item(iid, values=mevcut, tags=("sirket",))
            self._satiri_filtrele(iid)

    @staticmethod
    def _ilk_satir(deger):
        """Sonuç hücresinin yalnızca ilk satırını (özetini) döndürür."""
        if deger is None:
            return None
        return str(deger).split("\n", 1)[0]

    def _kaydet(self):
        """Sonuçları diske yazar. KAYDIRMA YOK: her sonuç bellekteki ana tablonun
        sabit (satır, sütun) konumunda durur; dosyalar bu tablodan birebir yazılır.
        - Yapılanlar (tam) dosyası: ana tablonun tamamı (başlık + tüm satırlar).
        - Özet dosyası: aynı satırlar, sonuç sütunlarında yalnızca ilk satır.
        Kullanıcının seçtiği kök dosya hiç değiştirilmez."""
        if not self.wb or not self.excel_yolu:
            return
        # 1) Tam veri = ana çalışma kitabını olduğu gibi kaydet
        try:
            self.wb.save(self.excel_yolu)
        except PermissionError:
            self._log_yaz("⚠️ Yapılanlar dosyası Excel'de açık olabilir, kapatın.")
            return
        except Exception as e:
            self._log_yaz(f"⚠️ Kaydetme hatası: {e}")
            return
        # 2) Özet dosyasını ana tablodan yeniden üret
        self._ozet_yaz()

    def _ozet_yaz(self):
        if not self.ozet_yolu or not self.ws:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(list(self.kolonlar))  # başlık satırı
            for r in sorted(self.satir_id.keys()):
                satir = [self.ws.cell(r, c).value for c in range(1, 5)]
                satir += [self._ilk_satir(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                          for i in range(len(SORGULAR))]
                ws.append(satir)
            wb.save(self.ozet_yolu)
        except PermissionError:
            self._log_yaz("⚠️ Özet dosyası Excel'de açık olabilir, kapatın.")
        except Exception as e:
            self._log_yaz(f"⚠️ Özet yazma hatası: {e}")


def main():
    root = tk.Tk()
    SgkSorguGUI(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        input("Hata oluştu, kapatmak için Enter...")
