# -*- coding: utf-8 -*-
"""
MTS ÇEVİRME — Excel toplu işleyici
==================================
Bir Excel listesindeki MTS dosya numaralarını (B sütunu, "2025/1041498" biçimi)
sırayla İCRA takibine çevirir; oluşan İcra Müdürlüğü'nü C, YENİ dosya numarasını
D sütununa yazar.

Her istek 127.0.0.1:8800 ofis proxy'sine POST edilir; ofis (uyap_app.py) onu canlı
e-imza UYAP oturumuyla iletir — tarayıcı/Playwright GEREKMEZ ("UYAP Bağlantısı" /
Paylaş-Al açık olmalı).

NEDEN ELDE YAZILDI: Logger'ın otomatik ürettiği sürüm, kaydedilen oturumu körlemesine
tekrar ediyordu; girdi olarak TCKN'yi (ilk adımdaki işe yaramaz kişi-fotoğrafı
sorgusu) alıyor, dosyayı ise SABİT gömülü 2025/1041498 ile arıyordu. Burada akış
dosya-no girdili ve döngüye uygun hale getirildi.

Tek dosya akışı (her satır için):
    1) search_phrase_detayli      → dosyaYil/dosyaSira ile ara, dosyaId al
    2) (hazırlık, best-effort)    → islem_turleri / ayrinti / acilis_dosya_bilgileri
    3) acilis_alacakli_taraflar   → alacaklı kişi/kurum id
    4) icra_takip_baslat          → takibi başlat; YENİ dosya no yanıt mesajından çekilir
       Örn. yanıt: {"type":"success","message":"... 2026/13219 dosya no ile icra
       takip dosyanız oluşturulmuştur! ..."}

Kullanım:
    python Mts_evirme.py <liste.xlsx>     → B sütununu işler, C sütununa yazar
    python Mts_evirme.py                  → tek dosya örneği çalıştırır (ORNEK_GIRDI)
"""

import re
import sys
import json
import urllib.request
import urllib.error

OFFICE_BASE = "http://127.0.0.1:8800"

COMMON_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "content-type": "application/json",
    "cache-control": "no-cache",
    "referer": "https://avukat.uyap.gov.tr/",
}

# Arayüz/örnek girdi: tek dosya çalıştırmak için. (Excel modunda kullanılmaz.)
PARAMETRELER = [
    ('dosya_no', 'Dosya No (örn. 2025/1041498)', '2025/1041498'),
]
ORNEK_GIRDI = {
    'dosya_no': '2025/1041498',
}

# Panel runner'ı bunu görünce bir "Dosya Seç…" düğmesi gösterir. Excel seçilip
# Çalıştır'a basılırsa, tek-girdi yerine excel_isle(seçilen_yol) çağrılır:
# B sütunundaki dosya no'ları sırayla çevirir, yeni no'yu C sütununa yazar.
EXCEL_GIRDI = {
    'etiket': 'Excel listesi (B = dosya no → C = müdürlük, D = yeni no)',
    'fonksiyon': 'excel_isle',
    'uzanti': [('Excel dosyası', '*.xlsx'), ('Tüm dosyalar', '*.*')],
}

# Aramada kullanılan birim filtreleri (İcra Dairesi). Kaydedilen oturumdan birebir.
ARAMA_SABIT = {
    'dosyaDurumKod': 0,
    'pageSize': 500,
    'pageNumber': 1,
    'birimId': '',
    'birimTuru2': '1101',
    'birimTuru3': '2',
}


def _gonder(endpoint, payload, log=None):
    """Tek bir isteği 8800 ofis proxy'sine POST eder; yanıtı (dict/list/metin) döner."""
    _govde = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    _istek = urllib.request.Request(OFFICE_BASE + "/" + endpoint, data=_govde, method="POST")
    for _ad, _deg in COMMON_HEADERS.items():
        _istek.add_header(_ad, _deg)
    try:
        with urllib.request.urlopen(_istek, timeout=90) as _y:
            _metin = _y.read().decode("utf-8", "replace")
            if log:
                log(f"   {endpoint} -> HTTP {_y.status}")
    except urllib.error.HTTPError as _e:
        if log:
            log(f"   {endpoint} -> HTTP {_e.code}")
        return {"_hata": f"HTTP {_e.code}"}
    except urllib.error.URLError as _e:
        return {"_hata": f"Ofise ulaşılamadı ({getattr(_e, 'reason', _e)}). "
                         "uyap_app.py açık ve Paylaş/Al aktif mi?"}
    try:
        return json.loads(_metin)
    except Exception:
        return _metin


def _dosya_no_coz(ham):
    """'2025/1041498' (ya da '2025 / 1041498') → (2025, 1041498). Çözülemezse None."""
    m = re.search(r'(\d{4})\s*/\s*(\d+)', str(ham or ""))
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _kazi(obj, anahtar):
    """İç içe dict/list yanıtta `anahtar`'ı arar; ilk anlamlı (boş olmayan) değeri döner.
    Eski sürümdeki kırılgan [0][0]['dosyaId'] varsayımının yerine geçer."""
    if isinstance(obj, dict):
        if anahtar in obj and obj[anahtar] not in (None, "", []):
            return obj[anahtar]
        for v in obj.values():
            r = _kazi(v, anahtar)
            if r is not None:
                return r
    elif isinstance(obj, list):
        for v in obj:
            r = _kazi(v, anahtar)
            if r is not None:
                return r
    return None


def _yeni_dosya_no(yanit):
    """Takip başlatma yanıtından (yeni_dosya_no, icra_mudurlugu, mesaj) döner.
    Örnek mesaj:
        'İzmir Abonelik Sözleşmeleri İcra Dairesi 2026/13219 dosya no ile
         icra takip dosyanız oluşturulmuştur! ...'
    Müdürlük = dosya no'dan ÖNCEKİ, 'İcra Dairesi'/'İcra Müdürlüğü' ile biten ad.
    Bulunamazsa müdürlük '' döner; dosya no için yedek olarak sade 'YYYY/NN' aranır."""
    mesaj = ""
    if isinstance(yanit, dict):
        mesaj = yanit.get("message") or yanit.get("mesaj") or ""
        if not mesaj:
            mesaj = json.dumps(yanit, ensure_ascii=False)
    elif isinstance(yanit, str):
        mesaj = yanit

    mudurluk, yeni = "", ""
    m = re.search(r'([^\n]*?İcra\s+(?:Dairesi|Müdürlüğü))\s+(\d{4}\s*/\s*\d+)',
                  mesaj, re.IGNORECASE)
    if m:
        mudurluk = re.sub(r'\s+', ' ', m.group(1)).strip()
        yeni = re.sub(r'\s+', '', m.group(2))
    else:
        m2 = re.search(r'(\d{4}\s*/\s*\d+)', mesaj)
        yeni = re.sub(r'\s+', '', m2.group(1)) if m2 else ""
    return yeni, mudurluk, mesaj


def calistir(girdi=None, log_fn=None):
    """Tek bir dosyayı çevirir. girdi={'dosya_no': '2025/1041498'}.
    Döner: {'dosya_no', 'yeni_dosya_no', 'mesaj'} veya hata için {'_hata': ...}."""
    girdi = girdi or {}
    log = log_fn or (lambda *a, **k: None)

    ham = str(girdi.get("dosya_no") or girdi.get("dosyaNo") or "").strip()
    coz = _dosya_no_coz(ham)
    if not coz:
        return {"dosya_no": ham, "yeni_dosya_no": "",
                "_hata": f"Geçersiz dosya no (beklenen 2025/1041498): {ham!r}"}
    yil, sira = coz
    log(f"▶ {ham}: dosya aranıyor (yıl {yil}, sıra {sira})")

    # 1) Dosyayı bul → dosyaId
    arama = _gonder("search_phrase_detayli.ajx",
                    {**ARAMA_SABIT, "dosyaYil": yil, "dosyaSira": sira}, log)
    if isinstance(arama, dict) and arama.get("_hata"):
        return {"dosya_no": ham, "yeni_dosya_no": "", "_hata": arama["_hata"]}
    dosya_id = _kazi(arama, "dosyaId")
    if not dosya_id:
        return {"dosya_no": ham, "yeni_dosya_no": "",
                "_hata": "Dosya bulunamadı veya dosyaId alınamadı."}

    # 2) Hazırlık adımları (UYAP'ın oturum durumunu kurması için; sonucu zincirlenmez)
    for ep in ("dosya_islem_turleri_sorgula_brd.ajx",
               "dosyaAyrintiBilgileri_brd.ajx",
               "mts_icra_acilis_dosya_bilgileri_brd.ajx"):
        try:
            _gonder(ep, {"dosyaId": dosya_id}, log)
        except Exception:
            pass

    # 3) Alacaklı taraf → alacakliKisiKurumId
    alacakli = _gonder("mts_icra_acilis_alacakli_taraflar_brd.ajx",
                       {"dosyaId": dosya_id}, log)
    alacakli_id = _kazi(alacakli, "ilgiliKisiIlkKisiKurumId")
    if not alacakli_id:
        return {"dosya_no": ham, "yeni_dosya_no": "",
                "_hata": "Alacaklı kişi/kurum id alınamadı."}

    # 4) Takibi başlat → yeni dosya no
    takip = _gonder("mts_icra_takip_baslat_brd.ajx", {
        "dosyaId": dosya_id,
        "alacakliKisiKurumId": alacakli_id,
        "kotaTur": 0,
        "icraTakipSebepId": "1",
        "odemeTipi": "7",
    }, log)
    if isinstance(takip, dict) and takip.get("_hata"):
        return {"dosya_no": ham, "yeni_dosya_no": "", "_hata": takip["_hata"]}

    yeni, mudurluk, mesaj = _yeni_dosya_no(takip)
    if not yeni:
        return {"dosya_no": ham, "icra_mudurlugu": mudurluk, "yeni_dosya_no": "",
                "mesaj": mesaj,
                "_hata": "Takip başlatıldı ama yanıttan yeni dosya no okunamadı. "
                         f"Yanıt: {mesaj[:200]}"}
    log(f"✔ {ham} → {mudurluk} {yeni}")
    return {"dosya_no": ham, "icra_mudurlugu": mudurluk,
            "yeni_dosya_no": yeni, "mesaj": mesaj}


# ─────────────────────────── EXCEL TOPLU İŞLEME ───────────────────────────

def excel_isle(excel_yolu, girdi_sutun="B", mudurluk_sutun="C", dosyano_sutun="D",
               baslik_satiri=1, log_fn=print):
    """Excel'i açar, `girdi_sutun`'daki dosya no'ları sırayla çevirir; sonucu
    İKİ sütuna yazar: `mudurluk_sutun` (C) → İcra Müdürlüğü, `dosyano_sutun` (D) →
    yeni dosya no. Her satırdan sonra kaydeder (uzun listede ilerleme kaybolmaz).
    baslik_satiri: başlık satırının numarası (veri bir altından başlar)."""
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    log = log_fn or (lambda *a, **k: None)
    gi = column_index_from_string(girdi_sutun)
    mu = column_index_from_string(mudurluk_sutun)
    do = column_index_from_string(dosyano_sutun)

    wb = load_workbook(excel_yolu)
    ws = wb.active

    veri_baslangic = baslik_satiri + 1
    # Sonuç sütunlarına başlık ata (boşsa)
    if baslik_satiri >= 1:
        if not ws.cell(baslik_satiri, mu).value:
            ws.cell(baslik_satiri, mu).value = "İcra Müdürlüğü"
        if not ws.cell(baslik_satiri, do).value:
            ws.cell(baslik_satiri, do).value = "Yeni Dosya No"

    def _kaydet():
        try:
            wb.save(excel_yolu)
        except PermissionError:
            log("⚠️ Excel kaydedilemedi — dosya başka bir programda (Excel'de) açık olabilir.")

    toplam = basari = 0
    for r in range(veri_baslangic, ws.max_row + 1):
        ham = ws.cell(r, gi).value
        if ham is None or str(ham).strip() == "":
            continue
        toplam += 1
        log(f"[{r}. satır] {ham}")

        # Zaten çevrilmişse atla (tekrar çalıştırmada ilerlemeyi koru):
        # dosya no (D) dolu ve HATA değilse bu satır tamamdır.
        mevcut = ws.cell(r, do).value
        if mevcut and not str(mevcut).startswith("HATA"):
            log(f"   atlandı (zaten dolu: {mevcut})")
            basari += 1
            continue

        try:
            sonuc = calistir({"dosya_no": str(ham)}, log)
        except Exception as e:
            sonuc = {"_hata": f"beklenmeyen hata: {e}"}

        if sonuc.get("yeni_dosya_no"):
            ws.cell(r, mu).value = sonuc.get("icra_mudurlugu", "")
            ws.cell(r, do).value = sonuc["yeni_dosya_no"]
            basari += 1
        else:
            ws.cell(r, do).value = "HATA: " + (sonuc.get("_hata") or "bilinmiyor")
        _kaydet()

    _kaydet()
    log(f"\nBitti: {basari}/{toplam} dosya çevrildi. Kaydedildi → {excel_yolu}")
    return {"toplam": toplam, "basari": basari, "dosya": excel_yolu}


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if len(sys.argv) >= 2:
        excel_isle(sys.argv[1], log_fn=print)
    else:
        print("Kullanım: python Mts_evirme.py <liste.xlsx>")
        print("  B sütunundaki dosya no'larını (2025/1041498) okur, yeni no'yu C'ye yazar.")
        print("\nTek dosya örneği çalıştırılıyor (ORNEK_GIRDI):\n")
        _sonuc = calistir(ORNEK_GIRDI, log_fn=print)
        print(json.dumps(_sonuc, ensure_ascii=False, indent=2, default=str))
