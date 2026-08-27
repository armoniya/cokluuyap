# -*- coding: utf-8 -*-
"""
Barkod Sorgu — Kapalı Tebligat barkodunu PTT'de sorgula
=========================================================
Excel'deki (Birim, Dosya No) her satır için:
  1. birim adı        -> birimId          (icra_core.birim_id_bul, önbellekli)
  2. birimId+yıl+sıra  -> dosyaId          (search_phrase_detayli.ajx)
  3. dosyaId           -> evrak listesi    (dosya_core.evrak_listesi_getir, sayfalı)
  4. "tur" alanı == "Kapalı Tebligat" olan evrak(lar) filtrelenir (bu değer
     docs/BELGE_ONBELLEK_PLANI.md §3'te CANLI DOĞRULANMIŞ örnek olarak geçiyor)
     — bir dosyada birden fazla olabilir, HEPSİ işlenir.
  5. her eşleşen evrak PDF olarak indirilir (dosya_core.evrak_html_indir HTML
     BEKLEYİP utf-8 ile decode ettiğinden bu içerik için UYGUN DEĞİL — bkz.
     `_evrak_pdf_indir` — bu modül kendi ham-bayt indiricisini kullanır) ve
     içindeki PTT barkodu çıkarılır (bkz. `barkod_bul`).
  6. barkod PTT'nin gönderi takip servisine sorgulanır (`ptt_sorgula`),
     sonuç ekrana (Panel tablosu) yazdırılır.

CANLI DOĞRULAMA (2026-07-28, gerçek ofis oturumu + Chrome ağ yakalama):

  * `view_document_brd.uyap`'ın "Kapalı Tebligat" evrakı için döndürdüğü içerik
    HTML DEĞİL, `Content-Type: application/pdf` ham bir PDF'tir (hem
    "Elektronik Tebligat" hem "Normal Tebligat" alt türlerinde aynı — 3 gerçek
    dosya üzerinde doğrulandı: 2026/94270 elektronik, 2026/94269 ve
    2026/93940 normal/fiziki). `dosya_core.evrak_html_indir` içeriği
    `.decode("utf-8","replace")` ile okuduğundan bu ikili PDF'i GERİ
    DÖNÜLEMEZ biçimde bozar — bu yüzden burada AYRI bir ham-bayt indirici
    (`_evrak_pdf_indir`) kullanılır (dosya_core.py'ye DOKUNULMADI, olası
    başka çağıranları etkilememek için).
  * PDF'in içindeki tebliğ mazbatasında PTT barkodu, Code-39 barkod fontuyla
    (`IDAutomationHC39M`, PDF kaynağında `/F4`) çizilen düz metin olarak
    gömülü — Code-39 kuralı gereği değer `*...*` arasında geçiyor, örn.
    PDF content stream'inde `/F4 10 Tf … (*4007663523293*)Tj`. Bu yüzden
    barkod, PDF'i görüntülemeye/OCR'a gerek KALMADAN metin regex'iyle
    doğrudan ve KESİN biçimde çıkarılabiliyor (bkz. `barkod_bul`).
  * PTT servisi `gonderitakip.ptt.gov.tr` DEĞİL (o alt alan adı artık
    `ptt.gov.tr` anasayfasına yönleniyor) — anasayfadaki "Gönderi Takip"
    kutusunun gittiği gerçek uç nokta `POST https://api.ptt.gov.tr/api/
    ShipmentTracking`, gövde bir NESNE değil HAM JSON STRING (ör. sadece
    `"4007663523293"` — `{"barkod":...}` göndermek 400 döner). Kimlik/çerez
    GEREKMEZ (düz curl ile de çalışıyor).
  * Yanıt biçimi (gerçek barkodla doğrulandı, ör. 2026/94269 dosyasının
    "Normal Tebligat" evrakı — barkod 4007663523293, PTT durumu "TESLİM
    EDİLDİ"): tek elemanlı liste, örn.
    `[{"errorMessage":"BAŞARILI","errorState":true,
       "hareketDongu":[{"aciklama":"KABUL EDİLDİ","il":...,"tarih":...,"saat":...}, …],
       "kabul":{"barkod_no":...,"gonderici":...,"kabul_tarihi":20260723,...},
       "sondurum":{"son_durum_aciklama":"TESLİM EDİLDİ","teslim_tarihi":"20260727",...}}]`
    Kayıtsız/elektronik-tebligat durumunda (barkodun PTT'de karşılığı yoksa):
    `[{"errorMessage":"KAYIT YOK","errorState":false,"hareketDongu":null,
       "kabul":null,"sondurum":null}]` — yani `errorState` GERÇEK/kayıt VAR mı
    anlamına geliyor (isim yanıltıcı, ama iki canlı örnekle doğrulandı).
  * "Elektronik Tebligat" (UETS/KEP ile teslim, `aciklama` alanında bu ibare
    geçer) evrakının PDF'inde de aynı barkod deseni var, ama bu bir UYAP
    belge takip numarasıdır — fiziki postaya hiç çıkmadığından PTT'de HİÇBİR
    ZAMAN kaydı olmaz (canlı doğrulandı: 2026/94270 örneği "KAYIT YOK"
    döndürdü). Bu yüzden bu modül elektronik tebligatlar için PTT'ye hiç
    sorgu ATMAZ, doğrudan "Elektronik Tebligat (PTT'de takip yok)" yazar —
    hem gereksiz dış istek hem de yanıltıcı "KAYIT YOK" satırından kaçınmak
    için.

KULLANICI BULGUSU (2026-08-03, üçüncü tur) — üç kural eklendi:
  1. [ARTIK GEÇERSİZ — bkz. altıncı tur notu aşağıda: bu kural CANLI olarak
     YANLIŞ çıktığı için TAMAMEN KALDIRILDI, tarihçe için burada bırakıldı.]
     "Kapalı Tebligat" türündeki evrak SADECE ödeme emri DEĞİL, borçlunun
     bankasına/işverenine giden haciz ihbarnamesi gibi ÜÇÜNCÜ ŞAHISLARA
     giden tebligatları da kapsıyor — bunları ayrım yapmadan işlemek
     veritabanına "kirli veri" yazıyordu. Artık PDF'in gövde metninde o
     dosyanın borçlusunun adı GEÇMİYORSA sonuç DB'ye YAZILMAZ (bkz.
     `_borclu_adi_govdede_var_mi`). UYARI: bu ayrım kullanıcının önerdiği bir
     SEZGİDİR — CANLI bir haciz ihbarnamesi örneğiyle doğrulanmadı (böyle bir
     ihbarnamenin gövdesinde de borçlunun adı geçiyor olabilir, bu durumda
     ayrım hiçbir şeyi elemez). PDF gövde metni çıkarımı da yalnız barkod
     fontu (Code-39, ASCII rakam) için CANLI doğrulanan yöntemin genel metne
     UYARLANMASIDIR — asıl gövde metni özel/altküme bir yazı tipi
     kodlamasıyla gömülüyse okunaksız çıkabilir; böyle durumda İHTİYATLI
     TARAF seçilir (ad bulunamadı sayılır, DB'ye YAZILMAZ).
  2. Bir evrağın barkodu daha önce DB'ye yazılmışsa (barkod bir kez atanır,
     değişmez) bir sonraki turda PDF TEKRAR İNDİRİLMEZ/barkod TEKRAR
     ÇIKARILMAZ — yalnız PTT durumu tazelenir (bkz. `_bilinen_barkod_kaydini_al`).
  3. Mazbata (Tebliğ Mazbatası / Kapalı Tebliğ Mazbatası) VARLIĞI zaten her
     turda `evrak_listesi_getir` üzerinden kontrol ediliyordu (bkz.
     `tm_evrak`/`ktm_var`) — bu artık barkod ister önbellekten gelsin ister
     yeni sorgulansın HER turda DB'ye yazılıp güncellenir.

KULLANICI BULGUSU (2026-08-03, beşinci tur) — tebligat türü sınıflandırması:
  "Kapalı Tebligat" evrakının PDF'indeki tebligat zarf pusulasında "...
  ihtiva eder" biçiminde bir cümle bulunuyor (kullanıcı bildirdi — CANLI bir
  örnekle DOĞRULANMADI); ayrıca elektronik tebligatlarda evrak açıklaması
  "Elektronik tebligat [tebligat yapılacak üçüncü taraf: <ad>]" biçiminde
  geliyor (bu ikinci biçim CANLI doğrulandı — bankaya giden bir örnekte
  görüldü). Bu iki kaynaktan `_tebligat_turunu_belirle` sabit kategorilere
  (Ödeme Emri / Banka (Haciz İhbarnamesi) / Maaş (Haciz İhbarnamesi) / 103
  Davetiyesi / Kıymet Takdiri Raporu / Bilirkişi Raporu) sınıflandırma
  dener; hiçbiri eşleşmezse "ihtiva eder" sonrası yakalanan HAM metni döner
  (kullanıcı testte eksik/yanlış anahtar kelimeyi görüp bildirebilsin diye —
  sessizce boş bırakmak yerine). UYARI: anahtar kelimeler kullanıcının
  verdiği TEK TEK örneklerden çıkarıldı, toplu canlı veriyle doğrulanmadı.

KULLANICI BULGUSU (2026-08-03, altıncı tur — CANLI doğrulandı, gerçek dosya
2026/84410) — üçüncü turdaki "borçlu adı gövdede yoksa DB'ye YAZILMAZ" filtresi
(`_borclu_adi_govdede_var_mi`) TAMAMEN KALDIRILDI:
  * O filtrenin dayandığı varsayım YANLIŞ çıktı: "Kapalı Tebligat" evrakı için
    indirilen PDF, ihbarnamenin/mektubun KENDİSİ DEĞİL — genel bir "Tebliğ
    Mazbatası" zarf/kapak şablonudur ("BU ZARFTA İcra Daireleri Haciz Tebliğ
    Yazısı VARDIR" biçiminde boş alanlı bir form). Bankaya/işverene giden bir
    haciz ihbarnamesinde bu şablonun "Adı ve Soyadı" alanı kurumun kendisine
    ait olduğundan (ya da elektronik/KEP teslimde hiç doldurulmadığından)
    borçlunun adı YAPISAL OLARAK bu metinde hiç geçmiyor — canlı örnekte
    (dosya 2026/84410, DENİZBANK'a giden Kapalı Tebligat, birimEvrakNo
    5761471) doğrulandı: gövde metninde "Adı ve Soyadı:" alanı boş, borçlu adı
    (KURTULUŞ YARDIMCI) hiç geçmiyor. Yani filtre "kirli veri"yi değil,
    3. şahsa giden HER haciz ihbarnamesini (banka + muhtemelen işveren/maaş)
    %100 oranda eliyordu — kullanıcı bunu "tebligat listesine hiç yok" olarak
    fark etti.
  * Ayrıca `_UCUNCU_TARAF_RE`'nin aradığı "üçüncü taraf:" (iki nokta üst üste
    ile) biçimi de YANLIŞ çıktı — gerçek UYAP metni "3. Taraf" (rakamla, İKİ
    NOKTA ÜST ÜSTE YOK) diyor: `"Elektronik Tebligat[Tebligat Yapılacak 3.
    Taraf DENİZBANK ANONİM ŞİRKETİ]"` (bu, beşinci turdaki "CANLI doğrulandı"
    notuna rağmen yanlıştı — bu sefer gerçekten canlı indirilip kontrol
    edildi). Düzeltilen desen artık bunu doğru yakalıyor.
  * Karar (kullanıcı, 2026-08-03): borçlu adı kontrolü olmadan HER "Kapalı
    Tebligat" evrakı (borçlunun kendisine giden VE bankaya/işverene giden
    dahil) DB'ye yazılır; kimin tebligatı olduğu ayrımı artık dışlama için
    değil, yalnız "Tebligat Türü" sütununu (Ödeme Emri / Banka (Haciz
    İhbarnamesi) / Maaş (Haciz İhbarnamesi) / vs.) doldurmak için kullanılır.

Doğrudan çalıştır:  python barkod_sorgu.py ["Dosya Sorgulama.xlsx"]
Panel içinden (Üretilen Modüller): EXCEL_GIRDI sayesinde "Dosya Seç…" düğmesi
çıkar; Çalıştır'a basınca excel_isle(excel_yolu, log_fn=...) çağrılır.
"""

import base64
import io
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zlib

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

import icra_core  # noqa: E402  (birim_id_bul, birim_listesi_getir, SorguMotoru, BIRIM_TURU2/3)
import dosya_core  # noqa: E402  (evrak_listesi_getir, öncelik geçidi, DOC_VIEWER_HAZIRLIK_ENDPOINT)

SorguMotoru = icra_core.SorguMotoru
BIRIM_TURU2 = icra_core.BIRIM_TURU2
BIRIM_TURU3 = icra_core.BIRIM_TURU3


_YEREL_JETON_YOLU = os.path.join(
    os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "UyapIcra", "gw_local_token")


def _yerel_jetonu_oku():
    """Jetonu HER ÇAĞRIDA dosyadan taze okur — ÖNBELLEKLEMEZ. Kullanıcı bulgusu
    (2026-08-14): ofis (uyap_app.py) her açılışta YENİ rastgele bir jeton
    üretip dosyayı ÜZERİNE yazıyor (bkz. uyap_proxy.py configure()); eskiden
    bu fonksiyon jetonu yalnız import anında BİR KEZ okuyup bir closure'a
    kilitliyordu — ofis Panel açıkken yeniden başlatılırsa (ör. çöküp
    kendiliğinden ya da elle yeniden başlatılırsa) Panel'in elindeki jeton
    ANINDA geçersiz kalıyor, "ofis yetki hatası (HTTP 403)" ile sonuçlanan
    tarama hatalarına yol açıyordu — Panel'i yeniden başlatmadan düzelmiyordu."""
    try:
        with open(_YEREL_JETON_YOLU, "r", encoding="utf-8") as f:
            return f.read().strip()
    except OSError:
        return ""


class _JetonEnjektor(urllib.request.BaseHandler):
    def http_request(self, req):
        try:
            if req.host and req.host.split(":")[0] in ("127.0.0.1", "localhost") \
                    and not req.has_header("X-uyap-local-token"):
                token = _yerel_jetonu_oku()
                if token:
                    req.add_unredirected_header("X-Uyap-Local-Token", token)
        except Exception:
            pass
        return req

    https_request = http_request


def _yerel_yetki_jetonunu_kur():
    """Bu betik Panel'in İÇİNDE (uyap_app ile aynı süreçte) çalışıyorsa ofis
    zaten süreç-global bir urllib opener kurup jetonu otomatik ekliyordur —
    bu durumda burada yapılan da onunla aynı jetonu taşıyacağından ZARARSIZDIR.
    Bağımsız `python ...py` süreci olarak çalıştırılırsa (bu dosyanın normal
    kullanım şekli) ofisin kullanıcıya özel jeton dosyasını okuyup aynı
    başlığı (X-Uyap-Local-Token) ekleyen bir opener kurar (bkz.
    baro_pulu_makbuzu_indiren.py — aynı desen). Jeton HER İSTEKTE taze
    okunur (bkz. _yerel_jetonu_oku) — ofis yeniden başlayıp jetonu
    değiştirse bile bu opener otomatik uyum sağlar."""
    urllib.request.install_opener(urllib.request.build_opener(_JetonEnjektor()))


_yerel_yetki_jetonunu_kur()

# Panel runner'ı (modules/uretilmis_runner.py) bunu görünce "Dosya Seç…"
# düğmesini kurar; Çalıştır'a basınca excel_isle(excel_yolu, log_fn=...) çağrılır.
EXCEL_GIRDI = {
    "etiket": "Dosya Sorgulama Excel'i (A: birim, B: dosya no)",
    "fonksiyon": "excel_isle",
    "uzanti": [("Excel dosyası", "*.xlsx"), ("Tüm dosyalar", "*.*")],
}

# Excel'e ALTERNATİF girdi: dosya seçilmezse (ne masaüstünde ne webde) Panel bu
# alanları formda gösterir ve calistir(girdi_sozlugu, log_fn) çağrılır — bkz.
# UYAP'ın "Dosya Sorgulama" ekranındaki tarih aralığı filtresiyle AYNI
# search_phrase_detayli.ajx sorgusu (icra_core.build_payload/ENDPOINT).
PARAMETRELER = [
    ("birim_adi", "Birim (İcra Dairesi) — tam ad", ""),
    ("baslangic", "Açılış Başlangıç Tarihi (GG.AA.YYYY)", ""),
    ("bitis", "Açılış Bitiş Tarihi (GG.AA.YYYY)", ""),
]

HEDEF_TUR = "Kapalı Tebligat"
PTT_ENDPOINT = "https://api.ptt.gov.tr/api/ShipmentTracking"
ELEKTRONIK_ANAHTAR = "Elektronik Tebligat"

# Avukat portalından atılan HER talep (21/2 yeniden tebliğ dahil) dosyanın evrak
# listesine bu türde düşüyor — dosya açılışındaki otomatik "Kapalı Tebligat"/
# "Ödeme İcra Emri" bu türde DEĞİL (kullanıcı isteği, 2026-08-13: canlı
# doğrulandı — dosya 2026/98353'te tam olarak bu türde 1 evrak vardı ve
# kullanıcı orada gerçekten yeniden tebliğ talebi göndermiş olduğunu
# doğruladı). Yalnız VARLIĞI kontrol edilir, PDF içeriği okunmaz.
TALEP_EVRAK_TUR = "Avukat Portal Tebligat Talebi"

# "Kapalı E-Tebliğ Mazbatası" — kullanıcı bulgusu (2026-08-03, yedinci tur —
# CANLI doğrulandı, dosya 2026/84410): elektronik tebligat gönderildiğinde
# UYAP, "Kapalı Tebligat" evrakından AYRI bir evrak türü daha üretiyor; bu
# mazbatanın PDF'i GERÇEK UETS teslim geçmişini (Tarih/Açıklama tablosu) ve
# barkodu ("Taahhütlü No") taşıyor — bu barkod, karşılık gelen "Kapalı
# Tebligat" evrakının barkoduyla AYNI, yani ikisi eşleştirilebiliyor (bkz.
# `_e_teblig_haritasini_olustur`).
KAPALI_E_TEBLIG_MAZBATASI_TUR = "Kapalı E-Tebliğ Mazbatası"

# "Tebliğ Mazbatası" — 2026/94270, 2026/94269, 2026/94274 gibi gerçek dosyalarda
# CANLI DOĞRULANDI (2026-07-28): "tur" alanı birebir bu string, "aciklama"
# alanı zaten "... - Barkod Numarası: ... - Tebliğ Edildi/Edilmedi" biçiminde
# dolu geliyor — ayrıca barkod/PDF okumaya gerek yok, sadece VARLIĞI (ve
# aciklama'daki hazır özeti) kontrol edilir.
TEBLIG_MAZBATASI_TUR = "Tebliğ Mazbatası"

# "Kapalı Tebliğ Mazbatası" — kullanıcı tarafından bildirildi (fiziki tebligatın
# taranmış/fotoğraf mazbatası, içeriği okunamaz — bkz. modül başlığı), ANCAK
# canlı örneklemde (27+ gerçek dosya) HİÇ rastlanmadı; bu yüzden tam string
# eşleşmesi yerine Türkçe karakter/boşluk farkına toleranslı `_tur_anahtar`
# ile eşleştirilir. Yalnızca VARLIĞI işaretlenir, PDF'i İNDİRİLMEZ/OKUNMAZ.
KAPALI_TEBLIG_MAZBATASI_TUR = "Kapalı Tebliğ Mazbatası"

_TR_CEVIRI = str.maketrans({
    "ç": "c", "Ç": "c", "ğ": "g", "Ğ": "g", "ı": "i", "İ": "i",
    "ö": "o", "Ö": "o", "ş": "s", "Ş": "s", "ü": "u", "Ü": "u",
})


def _tur_anahtar(tur):
    """Türkçe karakter/boşluk farkına toleranslı karşılaştırma anahtarı —
    canlı doğrulanamamış evrak türü adları için (bkz. KAPALI_TEBLIG_MAZBATASI_TUR)."""
    return re.sub(r"\s+", " ", (tur or "").translate(_TR_CEVIRI).lower()).strip()


# Dosyalar arası bekleme — UYAP'ı toplu turda yormamak için (bkz.
# baro_pulu_makbuzu_indiren.py'deki aynı desen).
SATIR_ARASI_SN = 2.0


def _dosya_no_ayir(dosya_no):
    """'2026/94270' -> (2026, 94270). Format uymazsa (None, None)."""
    m = re.match(r"^\s*(\d{4})\s*/\s*(\d+)\s*$", str(dosya_no or ""))
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def excel_oku(xlsx_yolu):
    """A: Birim/İcra Dairesi, B: Dosya No sütunlarını okur (başlık satırına göre
    kolon bulunur — 'Birim'/'Dosya No' içeren başlıklar aranır; bulunamazsa ilk
    iki sütun kullanılır). Döner: list[(birim_adi, dosya_no)]."""
    wb = openpyxl.load_workbook(xlsx_yolu, read_only=True, data_only=True)
    ws = wb.active
    satirlar = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    if not satirlar:
        return []
    basliklar = [str(h or "").strip().lower() for h in satirlar[0]]
    try:
        bi = next(i for i, h in enumerate(basliklar) if "birim" in h or "daire" in h)
    except StopIteration:
        bi = 0
    try:
        di = next(i for i, h in enumerate(basliklar) if "dosya no" in h)
    except StopIteration:
        di = 1

    sonuc = []
    for row in satirlar[1:]:
        if bi >= len(row) or di >= len(row):
            continue
        birim, dosya_no = row[bi], row[di]
        if not birim or not dosya_no:
            continue
        sonuc.append((str(birim).strip(), str(dosya_no).strip()))
    return sonuc


def _dosya_id_coz(motor, birim_id, yil, sira, log):
    """birimId+yıl+sıra -> ham UYAP kaydı (search_phrase_detayli.ajx). birim_id
    VERİLEREK aranır — aynı yıl/sıra farklı birimlerde de olabileceğinden
    birimsiz arama YANLIŞ dosyayı döndürebilir (bkz.
    baro_pulu_makbuzu_indiren.py — aynı fonksiyon). TAM kaydı döner (yalnız
    dosyaId değil) — DB'ye kapak künyesi yazarken (bkz. `_dosya_isle`,
    `icra_models.ingest.dosya_kunyesi_kaydet`) gereken diğer alanlar
    (birimAdi/birimTuru*/dosyaDurumKod/... ) da bu kayıtta gelir."""
    payload = {
        "dosyaDurumKod": 0, "pageSize": 500, "pageNumber": 1,
        "dosyaYil": yil, "dosyaSira": sira,
        "birimId": birim_id, "birimTuru2": BIRIM_TURU2, "birimTuru3": BIRIM_TURU3,
    }
    veri = dosya_core._post_eszamanli_korumali(
        dosya_core._arka_plan_istek, motor, icra_core.ENDPOINT, payload, log)
    try:
        kayitlar = veri[0] if isinstance(veri, list) else []
        if kayitlar:
            return kayitlar[0]
    except Exception as e:
        # Önceden sessizce yutuluyordu (bare except: pass) — gerçek bir
        # ayrıştırma hatası da "dosya bulunamadı" gibi görünüyor, kullanıcı
        # sebebi hiç göremiyordu (kullanıcı bulgusu, 2026-08-13). Akış aynı
        # kalır (yine None döner), yalnız sebep artık loglanır.
        log(f"  ⚠️ Dosya kaydı ayrıştırılamadı: {e}")
    return None


def _tarih_araligi_dosyalari_getir(motor, birim_adi, baslangic, bitis, log):
    """PARAMETRELER'den gelen tarih aralığı girdisini search_phrase_detayli.ajx
    ile sorgular — UYAP'ın 'Dosya Sorgulama' ekranındaki tarih aralığı filtresiyle
    AYNI endpoint (bkz. icra_core.py başlığı). icra_core.IcraSorgu.ara() da bunu
    yapıyor ama HER dosya için AYRICA borçlu-detay isteği + Django DB senkronu
    çalıştırıyor — bu modül zaten kendi evrak/PDF/PTT işini yapacağından o ekstra
    ağır adımlar gereksiz; bu yüzden `build_payload`/`parse_records` doğrudan
    kullanılır (Django bağımlılığı da eklenmemiş olur). Döner: ham UYAP kayıt
    listesi — dosyaId zaten yanıtta geldiğinden AYRICA `_dosya_id_coz` ile
    çözülmesine gerek YOK."""
    birim_adi = (birim_adi or "").strip()
    baslangic = (baslangic or "").strip()
    bitis = (bitis or "").strip()
    if not birim_adi:
        raise RuntimeError("Birim adı girilmedi.")
    if not baslangic and not bitis:
        raise RuntimeError("Açılış başlangıç/bitiş tarihinden en az biri girilmeli "
                            "(yoksa birimdeki TÜM dosyalar taranır — çok uzun sürer).")
    if not icra_core.birim_id_bul(birim_adi):
        raise RuntimeError(f"'{birim_adi}' birimi bulunamadı — tam adını yazın.")

    values = {"birimAdi": birim_adi, "dosyaAcilisTarihiStart": baslangic, "dosyaAcilisTarihiEnd": bitis}
    tum, gorulen, sayfa = [], set(), 1
    while True:
        payload = icra_core.build_payload(values, icra_core.DURUM_VARSAYILAN, None)
        payload["pageNumber"] = sayfa
        _status, veri = motor._post(icra_core.ENDPOINT, payload)
        kayitlar = icra_core.parse_records(veri)
        for rec in kayitlar:
            did = rec.get("dosyaId")
            if did and did in gorulen:
                continue
            if did:
                gorulen.add(did)
            tum.append(rec)
        if len(kayitlar) < icra_core.PAGE_SIZE:
            break
        sayfa += 1
        if sayfa > 200:
            log("⚠️ Sayfa sınırına (200) ulaşıldı; bazı dosyalar alınmamış olabilir.")
            break
        log(f"… {len(tum)} dosya alındı, sayfa {sayfa} çekiliyor")
    log(f"📅 Tarih aralığında {len(tum)} dosya bulundu.")
    return tum


def _evrak_pdf_indir(motor, dosya_id, evrak_id, log_fn=None, istek_sarici=None, timeout=90):
    """`dosya_core.evrak_html_indir`'in ham-bayt eşdeğeri — o fonksiyon içeriği
    utf-8 ile decode ettiğinden ("Kapalı Tebligat" GERÇEKTE bir PDF, HTML
    değil — bkz. modül başlığındaki canlı doğrulama notu) burada AYNI istek
    sırası (önce `getDocViewerParameters.ajx`, sonra GET) tekrarlanır ama
    yanıt bayt olarak (decode edilmeden) döner. Döner: (content_type, ham_bayt)."""
    log = log_fn or (lambda *a, **k: None)
    sarici = istek_sarici or dosya_core._bireysel_istek

    dosya_core._post_eszamanli_korumali(sarici, motor, dosya_core.DOC_VIEWER_HAZIRLIK_ENDPOINT, {}, log)

    qs = urllib.parse.urlencode({"evrakId": evrak_id, "dosyaId": dosya_id})
    url = f"{motor.base}/{dosya_core.EVRAK_ICERIK_YOLU}?{qs}"
    req = urllib.request.Request(url, method="GET")
    req.add_header("accept", "*/*")
    req.add_header("referer", f"{motor.base}/dosya-sorgulama")

    with sarici():
        with urllib.request.urlopen(req, timeout=timeout) as r:
            status = r.status
            content_type = r.headers.get("Content-Type", "")
            data = r.read()
    if status >= 400:
        raise RuntimeError(f"Evrak indirilemedi (HTTP {status})")
    return content_type, data


# ── PDF içi barkod çıkarımı (bkz. modül başlığı — CANLI DOĞRULANDI) ─────────
_PDF_STREAM_RE = re.compile(rb"\d+ 0 obj\s*<<(.*?)>>\s*stream\r?\n", re.DOTALL)
_BARKOD_PDF_RE = re.compile(rb"\(\*([0-9A-Za-z]{6,20})\*\)\s*Tj")


def _pdf_metin_coz(pdf_bytes):
    """PDF'teki tüm FlateDecode content stream'lerini çözüp birleştirir
    (font/resource gibi Flate OLMAYAN stream'ler sessizce atlanır)."""
    parcalar = []
    for m in _PDF_STREAM_RE.finditer(pdf_bytes):
        start = m.end()
        end = pdf_bytes.find(b"endstream", start)
        if end == -1:
            continue
        try:
            parcalar.append(zlib.decompress(pdf_bytes[start:end]))
        except zlib.error:
            continue
    return b"\n".join(parcalar)


def barkod_bul(pdf_bytes, log_fn=None):
    """PDF içindeki Code-39 barkod fontuyla (`/F4`, IDAutomationHC39M) çizilen
    metni bulur — Code-39 kuralınca değer `*...*` arasında geçer (bkz. modül
    başlığı). Birden fazla eşleşme varsa (mazbatanın iki nüshası aynı barkodu
    taşıyor) ilkini döner. Bulunamazsa None döner."""
    log = log_fn or (lambda *a, **k: None)
    if not pdf_bytes.startswith(b"%PDF"):
        log(f"    ⚠️ Beklenmeyen içerik türü (PDF değil, {len(pdf_bytes)} bayt) — barkod aranmadı.")
        return None
    metin = _pdf_metin_coz(pdf_bytes)
    m = _BARKOD_PDF_RE.search(metin)
    if not m:
        log("    ⚠️ PDF içinde Code-39 barkod örüntüsü bulunamadı.")
        return None
    return m.group(1).decode("ascii")


# ── PDF içi gövde metni çıkarımı (borçlu adı eşleşmesi için — bkz. modül
# başlığındaki KULLANICI BULGUSU notu, CANLI DOĞRULANMADI) ─────────────────
_PDF_TEXT_STR_RE = re.compile(rb"\((?:[^()\\]|\\.)*\)")
_PDF_ESCAPE_RE = re.compile(rb"\\([0-7]{1,3}|.)", re.DOTALL)


def _pdf_dize_coz(ham_parantezli):
    """PDF metin operatörünün ('(...)Tj'/TJ) dizge operandındaki kaçış
    dizilerini çözer: '\\(' '\\)' '\\\\' ve sekizli kaçışlar (`\\ddd`)."""
    ic = ham_parantezli[1:-1]

    def _cevir(m):
        g = m.group(1)
        if g in (b"(", b")", b"\\"):
            return g
        if re.match(rb"^[0-7]{1,3}$", g):
            return bytes([int(g, 8) & 0xFF])
        return b""  # satır-sonu kaçışı (\\<newline>) gibi görmezden gelinenler
    return _PDF_ESCAPE_RE.sub(_cevir, ic)


def _pdf_govde_metni(pdf_bytes, log_fn=None):
    """PDF içeriğindeki TÜM metin-gösterme operatörlerinin dizge
    operandlarını sırayla çıkarıp birleştirir — tam bir PDF ayrıştırıcı
    DEĞİL, yalnız düz metin arama (borçlu adı eşleşmesi) için yeterli kaba
    bir çıkarım (bkz. modül başlığındaki UYARI — CANLI DOĞRULANMADI)."""
    metin = _pdf_metin_coz(pdf_bytes)
    parcalar = []
    for m in _PDF_TEXT_STR_RE.finditer(metin):
        try:
            parcalar.append(_pdf_dize_coz(m.group(0)))
        except Exception:
            continue
    ham = b"".join(parcalar)
    try:
        return ham.decode("cp1254")
    except UnicodeDecodeError:
        return ham.decode("latin-1", "replace")


# ── Kapalı E-Tebliğ Mazbatası çözümü (kullanıcı bulgusu, 2026-08-03, yedinci
# tur — CANLI doğrulandı, dosya 2026/84410, örnek dosya
# "2026_84410_Kapalı_E-Tebliğ_Mazbatası_07_07_2026.pdf") ─────────────────────
# Gövde metni (tr_lower ile) şu düz metne çözülüyor (satır sonları YOK, tablo
# hücreleri art arda eklenmiş): "...Örnek No: 25TarihAçıklama30/06/2026
# 15:36Muhatap hesabına teslim edilmek üzere UETStarafından teslim
# alındı.30/06/2026 15:36Tebligat alıcı için ayrılmış tebligat alanına
# (hesabına)başarılı bir şekilde konuldu.05/07/2026Tebligat, alıcının
# hesabına iletilmesine müteakip mevzuat gereği belirlenen süre sonunda
# otomatik olarak okundu sayıldı.E-TEBLİĞ MAZBATASIKarar No:*5002864776768*
# Yukarıda Belirtilen Tebliğ Bilgileri...". "TarihAçıklama" başlığı ile
# "E-TEBLİĞ MAZBATASI" alt başlığı ARASINDAKİ bölüm, art arda
# "GG/AA/YYYY[ SS:DD]<açıklama>" satırlarından oluşan bir tabloya karşılık
# geliyor; SON satır UETS'in vardığı NİHAİ sonucu (gerçekten okundu ya da
# süre sonunda otomatik okundu sayıldı) taşıyor — bu, e-tebligatın GERÇEK
# tebliğ tarihi/durumudur (kullanıcı: "borçluya e-tebligat gönderildiyse
# kapalı e-tebliğ mazbatasında okunmuş sayıldığı tarih yer alır, tebliğ
# tarihi o tarih olur").
_E_TEBLIG_TABLO_RE = re.compile(r"tarihaciklama(.*?)(?:e-teblig mazbatasi|$)", re.DOTALL)
_E_TEBLIG_SATIR_RE = re.compile(r"(\d{2}/\d{2}/\d{4})(?:\s+(\d{2}:\d{2}))?(.*?)(?=\d{2}/\d{2}/\d{4}|\Z)", re.DOTALL)


def _e_teblig_mazbatasi_coz(govde_metni, log_fn=None):
    """Kapalı E-Tebliğ Mazbatası PDF gövdesindeki Tarih/Açıklama tablosunun
    SON satırını (tarih, açıklama) ikilisi olarak döner — bkz. yukarıdaki
    canlı doğrulanmış örnek. Tablo bulunamazsa ("", "") döner."""
    log = log_fn or (lambda *a, **k: None)
    govde_l = dosya_core.tr_lower(govde_metni or "")
    m = _E_TEBLIG_TABLO_RE.search(govde_l)
    if not m:
        log("    ⚠️ E-Tebliğ Mazbatası tablosu (Tarih/Açıklama) bulunamadı.")
        return "", ""
    # tr_lower karakter-karşılığı 1:1 (uzunluk değişmez — bkz. icra_core.tr_lower),
    # bu yüzden govde_l üzerindeki konumlar orijinal (küçültülmemiş) govde_metni
    # üzerinde de aynı aralığa karşılık gelir — orijinal metin ekranda daha
    # okunaklı olduğundan (Türkçe karakterler korunur) o dilimlenir.
    tablo = govde_metni[m.start(1):m.end(1)]
    satirlar = list(_E_TEBLIG_SATIR_RE.finditer(tablo))
    if not satirlar:
        return "", ""
    son = satirlar[-1]
    tarih = son.group(1)
    if son.group(2):
        tarih = f"{tarih} {son.group(2)}"
    aciklama = re.sub(r"\s+", " ", son.group(3)).strip()
    return tarih, aciklama


def _e_teblig_haritasini_olustur(motor, dosya_id, evraklar, log):
    """Dosyadaki TÜM 'Kapalı E-Tebliğ Mazbatası' evraklarını indirir, her
    birinin barkodunu (aynı Code-39 font yöntemiyle, bkz. `barkod_bul` —
    bu mazbatalarda da 'Taahhütlü No' alanı aynı fontla basılıyor) ve
    Tarih/Açıklama tablosunun son satırını çözüp barkod -> (tarih, açıklama)
    sözlüğüne ekler. Bu barkod, karşılık gelen 'Kapalı Tebligat' evrakının
    barkoduyla AYNI olduğundan (canlı doğrulandı) elektronik tebligatın
    GERÇEK sonucunu bulmak için kullanılır (bkz. `_dosya_isle`). Yalnız
    GERÇEKTEN bir elektronik evrakla karşılaşılınca ÇAĞRILIR — dosyada
    elektronik tebligat yoksa hiç mazbata indirilmez."""
    harita = {}
    mazbatalar = [e for e in evraklar if e.get("tur") == KAPALI_E_TEBLIG_MAZBATASI_TUR]
    for mz in mazbatalar:
        try:
            _, mz_pdf = _evrak_pdf_indir(
                motor, dosya_id, mz.get("evrakId"), log_fn=log, istek_sarici=dosya_core._arka_plan_istek)
            mz_barkod = barkod_bul(mz_pdf, log)
            if not mz_barkod:
                continue
            mz_govde = _pdf_govde_metni(mz_pdf, log)
            tarih, aciklama = _e_teblig_mazbatasi_coz(mz_govde, log)
            if tarih:
                harita[mz_barkod] = (tarih, aciklama)
        except Exception as e:
            log(f"    ⚠️ E-Tebliğ Mazbatası okunamadı (evrak {mz.get('birimEvrakNo')}): {e}")
    if mazbatalar:
        log(f"  📬 {len(harita)}/{len(mazbatalar)} E-Tebliğ Mazbatası çözüldü.")
    return harita


# Kullanıcı bulgusu (2026-08-03, altıncı tur): burada eskiden bir
# `_borclu_adi_govdede_var_mi` fonksiyonu vardı ("Kapalı Tebligat" PDF'inin
# gövde metninde borçlunun adı geçmiyorsa 3. şahıs tebligatı sayılıp DB'ye
# YAZILMAZDI). KALDIRILDI: indirilen PDF ihbarnamenin kendisi değil genel bir
# "Tebliğ Mazbatası" zarf şablonu olduğundan, bankaya/işverene giden bir haciz
# ihbarnamesinde borçlunun adı bu metinde YAPISAL OLARAK hiç geçmiyor (canlı
# doğrulandı, dosya 2026/84410) — filtre "kirli veri"yi değil 3. şahsa giden
# HER haciz ihbarnamesini eliyordu. Artık ayrım yapılmadan HER 'Kapalı
# Tebligat' evrakı DB'ye yazılır; kime gittiği bilgisi (borçlu/banka/işveren)
# yalnız `_tebligat_turunu_belirle`nin ürettiği "Tebligat Türü" sütununda
# gösterilir (bkz. modül başlığı).


# ── Tebligat türü sınıflandırması (bkz. modül başlığındaki KULLANICI
# BULGUSU, 2026-08-03/beşinci tur — CANLI toplu veriyle DOĞRULANMADI) ──────
_IHTIVA_EDER_RE = re.compile(r"(.{0,140}?)\s*ihtiva\s+ed(?:er|ilmi|iyor)", re.IGNORECASE)
# Kullanıcı bulgusu (2026-08-03, altıncı tur — CANLI doğrulandı, dosya
# 2026/84410): gerçek UYAP metni "üçüncü taraf:" (iki nokta üst üste ile)
# DEĞİL, "3. Taraf" (rakamla, iki nokta üst üste YOK) diyor — ör.
# "Elektronik Tebligat[Tebligat Yapılacak 3. Taraf DENİZBANK ANONİM
# ŞİRKETİ]". Her iki biçim de (eski varsayım + gerçek biçim) kabul edilir.
_UCUNCU_TARAF_RE = re.compile(r"(?:ucuncu taraf|3\.\s*taraf)\s*:?\s*([^\]]*)\]")

# Kullanıcı bulgusu (2026-08-03, yedinci tur — CANLI doğrulandı, dosya
# 2026/84410'un "Kapalı E-Tebliğ Mazbatası" PDF'i): zarfın içeriği "BU
# ZARFTA <belge adı> VARDIR" cümlesiyle beyan ediliyor — ör. "BU ZARFTA
# Ödeme İcra Emri VARDIR", "BU ZARFTA İcra Daireleri Haciz Tebliğ Yazısı
# VARDIR". Beşinci turda varsayılan "... ihtiva eder" biçimi hâlâ olası bir
# şablon farkı için `_IHTIVA_EDER_RE` ile ayrıca denenir (o CANLI
# doğrulanmadı).
_ZARF_ICERIGI_RE = re.compile(r"bu zarfta\s+(.*?)\s+vardir")


def _tebligat_turunu_belirle(aciklama, pdf_govde_metni=None):
    """Evrak açıklaması + (varsa) PDF gövde metninden tebligat türünü
    sınıflandırır. Döner: 'Ödeme Emri' | '103 Davetiyesi' |
    'Kıymet Takdiri Raporu' | 'Bilirkişi Raporu' | 'Banka (Haciz
    İhbarnamesi)' | 'Maaş (Haciz İhbarnamesi)' | 'Haciz İhbarnamesi' (taraf
    belirsiz) | 'ihtiva eder' sonrası yakalanan ham metin (bilinen bir
    kategoriyle eşleşmediyse — kullanıcı testte anahtar kelime eksikse
    görüp bildirebilsin diye SESSİZCE boş bırakılmaz) | "" (hiçbir ipucu
    bulunamadıysa). Karşılaştırma `tr_lower` (Türkçe karakterleri ASCII'ye
    çevirir) ile yapılır, bu yüzden anahtar kelimeler ASCII yazılmıştır."""
    aciklama_l = dosya_core.tr_lower(aciklama or "")
    govde_l = dosya_core.tr_lower(pdf_govde_metni or "")
    birlesik = f"{aciklama_l} {govde_l}"

    if "odeme emri" in birlesik:
        return "Ödeme Emri"
    if "kiymet takdir" in birlesik:
        return "Kıymet Takdiri Raporu"
    if "bilirkisi" in birlesik:
        return "Bilirkişi Raporu"
    if "103" in birlesik and ("davet" in birlesik or "mal beyan" in birlesik or "madde" in birlesik):
        return "103 Davetiyesi"

    ucuncu_taraf_m = _UCUNCU_TARAF_RE.search(aciklama_l)
    ucuncu_taraf = ucuncu_taraf_m.group(1) if ucuncu_taraf_m else ""

    # Kullanıcı bulgusu (2026-08-03, yedinci tur — CANLI doğrulandı, dosya
    # 2026/84410'un "Kapalı E-Tebliğ Mazbatası" PDF'i): zarfın NE İÇERDİĞİ
    # ("Ödeme İcra Emri", "İcra Daireleri Haciz Tebliğ Yazısı" vb.) "BU
    # ZARFTA ... VARDIR" cümlesinde geçiyor — "3. Taraf" parantezi ise SADECE
    # KİME gönderildiğini (kurum/kişi adı) söylüyor. "Maaş"/"ücret"/"işveren"
    # anahtar kelimesi bu yüzden ZARF İÇERİĞİNDE aranmalı, alıcı adında DEĞİL
    # (aksi halde işverenin şirket adı hiçbir zaman bu kelimeleri
    # içermeyeceğinden maaş haczi hiç yakalanamaz).
    zarf_icerigi_m = _ZARF_ICERIGI_RE.search(govde_l)
    zarf_icerigi = zarf_icerigi_m.group(1) if zarf_icerigi_m else ""

    if "haciz ihbarname" in birlesik or "haciz ihbarname" in zarf_icerigi:
        # Kullanıcı bulgusu (2026-08-03, altıncı tur — CANLI doğrulandı):
        # "banka" YERİNE "bank" aranır — DENİZBANK/AKBANK/ŞEKERBANK/ING
        # BANK/ENPARA BANK gibi çoğu banka adı "...banka" değil "...bank"
        # (sonunda 'a' YOK) ile bitiyor, "banka" substring'i bunları kaçırıyordu.
        if "bank" in (ucuncu_taraf or birlesik):
            return "Banka (Haciz İhbarnamesi)"
        if any(k in zarf_icerigi for k in ("maas", "ucret", "isveren", "is yeri", "hizmet")):
            return "Maaş (Haciz İhbarnamesi)"
        return "Haciz İhbarnamesi"

    # "haciz ihbarname" ibaresi geçmese bile elektronik tebligatın üçüncü
    # taraf parantezi (bkz. modül başlığı, CANLI doğrulandı) banka ayrımı
    # için tek başına yeterli olabilir (kurum ADI zaten "bank" içeriyor).
    if ucuncu_taraf and "bank" in ucuncu_taraf:
        return "Banka"
    if any(k in zarf_icerigi for k in ("maas", "ucret", "isveren", "is yeri", "hizmet")):
        return "Maaş"

    if pdf_govde_metni:
        m2 = _IHTIVA_EDER_RE.search(pdf_govde_metni)
        if m2:
            aday = m2.group(1).strip(" .,:;\"'()[]")
            if aday:
                return aday[:60]
    return ""


def ptt_sorgula(barkod, log_fn=None, timeout=20):
    """PTT gönderi takip servisini sorgular (bkz. modül başlığı — CANLI
    DOĞRULANDI 2026-07-28). Gövde ham JSON STRING olmalı (nesne DEĞİL —
    `{"barkod": ...}` göndermek 400 döner)."""
    log = log_fn or (lambda *a, **k: None)
    govde = json.dumps(barkod, ensure_ascii=False).encode("utf-8")
    istek = urllib.request.Request(PTT_ENDPOINT, data=govde, method="POST")
    istek.add_header("Content-Type", "application/json")
    istek.add_header("Accept", "application/json")
    try:
        with urllib.request.urlopen(istek, timeout=timeout) as r:
            metin = r.read().decode("utf-8", "replace")
    except urllib.error.HTTPError as e:
        return {"_hata": f"PTT HTTP {e.code}"}
    except urllib.error.URLError as e:
        return {"_hata": f"PTT'ye ulaşılamadı ({getattr(e, 'reason', e)})"}
    try:
        veri = json.loads(metin)
    except Exception:
        return {"_hata": f"PTT yanıtı JSON değil: {metin[:200]!r}"}
    if isinstance(veri, list):
        return veri[0] if veri else {"_hata": "PTT boş liste döndü"}
    return veri


def _ptt_tarih_bicimle(ham):
    """PTT'nin `sondurum` alanındaki tarihler 'YYYYMMDD' (ör. 20260727, int ya
    da string) biçiminde geliyor — ekranda 'GG/AA/YYYY' olarak gösterilir.
    `hareketDongu` içindeki tarihler zaten 'GG/AA/YYYY' geldiğinden bu
    fonksiyondan GEÇİRİLMEZ (bkz. `_ptt_durum_ve_tarih`)."""
    s = str(ham or "").strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[6:8]}/{s[4:6]}/{s[0:4]}"
    return s


def _ptt_durum_ve_tarih(veri):
    """PTT yanıtını (durum_metni, tarih_metni) çiftine çevirir (alan adları
    CANLI DOĞRULANDI — bkz. modül başlığı). `errorState` GERÇEK/kayıt var mı
    demek (isim yanıltıcı ama iki canlı örnekle doğrulandı: teslim edilmiş
    gerçek bir gönderide True, kayıtsızda False).

    `sondurum` — CANLI DOĞRULANDI (2026-07-28, 28 dosyalık toplu koşu) — yalnız
    TESLİM EDİLDİ'de değil, "DAĞITIMDA"/"İADE" gibi HENÜZ sonuçlanmamış
    durumlarda da dolu geliyor (`son_durum_aciklama` her zaman güncel durumu
    yansıtıyor). Ancak `teslim_tarihi` alanı yalnız GERÇEKTEN teslim
    edildiğinde gerçek bir tarih taşıyor — teslim edilmemişken PTT bu alana
    yer-tutucu olarak LİTERAL "0" STRING'i koyuyor (canlı örnek: DAĞITIMDA
    durumunda `teslim_tarihi: "0"`, gerçek güncel tarih ise
    `son_islem_tarihi`'nde). "0" niteliksiz bir değer olduğundan (boş değil,
    `or` ile atlanmaz) AYRICA kontrol edilip `son_islem_tarihi`'ne düşülür.
    Bu aracın asıl kullanım amacı çoğunlukla HENÜZ sonuçlanmamış
    (yolda/dağıtımda) tebligatları görmek olduğundan bu düzeltme önemli."""
    if not isinstance(veri, dict):
        return str(veri), ""
    if veri.get("_hata"):
        return veri["_hata"], ""
    if not veri.get("errorState"):
        return veri.get("errorMessage") or "Kayıt yok", ""
    son = veri.get("sondurum") or {}
    if son.get("son_durum_aciklama"):
        teslim_tarihi = son.get("teslim_tarihi")
        if not teslim_tarihi or str(teslim_tarihi).strip() == "0":
            teslim_tarihi = son.get("son_islem_tarihi")
        tarih = _ptt_tarih_bicimle(teslim_tarihi)
        return son["son_durum_aciklama"], tarih
    hareket = veri.get("hareketDongu")
    if isinstance(hareket, list) and hareket:
        son_hareket = hareket[-1]
        tarih = son_hareket.get("tarih", "")
        if son_hareket.get("saat"):
            tarih = f"{tarih} {son_hareket['saat']}".strip()
        return son_hareket.get("aciklama") or veri.get("errorMessage") or "Bilinmeyen durum", tarih
    return veri.get("errorMessage") or "Bilinmeyen durum", ""


# Ekrandaki/excel çıktısındaki tablo hep bu sırayla, hep bu kolonlarla
# gösterilir (satırlar arasında kolon farkı olursa Panel tablosu satırları
# hizasız gösterir — bu yüzden her satır bu şablondan üretilir).
_KOLONLAR = ["Birim", "Dosya No", "Evrak Açıklaması", "Tebligat Türü", "Evrak Tarihi", "Barkod",
             "PTT Durumu", "Son İşlem Tarihi", "Tebliğ Mazbatası",
             "Kapalı Tebliğ Mazbatası", "Yeniden Tebliğ Talebi", "Durum", "Ham Yanıt"]


def _satir_olustur(temel, **degerler):
    satir = {k: "" for k in _KOLONLAR}
    satir.update(temel)
    satir.update(degerler)
    return satir


class _SonucListesi(list):
    """Panel'e dönen satır listesiyle AYNI zamanda üretilen Excel çıktısının
    baytlarını da taşıyan liste alt sınıfı. `_uretilmis_sonuc` (web) ve
    `uretilmis_runner` (masaüstü) sıradan bir `list` bekler ve tablo olarak
    çizer — `dosya_b64`/`dosya_ad` öznitelikleri onlar için görünmez/zararsız,
    ama web tarafı bunları görüp bir 'İndir' düğmesi ekleyebilir (bkz.
    server.py `uretilmis_run`/`_uretilmis_sonuc`)."""
    dosya_b64 = None
    dosya_ad = None


def _excel_bayt_uret(sonuclar):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barkod Sorgu Sonuc"
    ws.append(_KOLONLAR)
    for satir in sonuclar:
        ws.append([satir.get(k, "") for k in _KOLONLAR])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sonucu_excele_yaz(sonuclar, taban_ad, kaynak_klasoru, log):
    """İşlem bitince sonuç tablosunu bir .xlsx'e yazar. Web'de yüklenen dosya
    geçici (ve iş bitince silinen) bir klasörde yaşadığından "girdiyle aynı
    klasöre yaz" YETERSİZ — bu yüzden baytlar HER ZAMAN üretilip döndürülür
    (Panel web bunu base64 olarak yanıta ekleyip tarayıcıda indirme düğmesi
    gösterir); `kaynak_klasoru` GERÇEKTEN yazılabilir bir yerse (masaüstü/CLI
    kullanımı) ayrıca oraya da best-effort bir kopya bırakılır — tarih aralığı
    modunda (girdi bir Excel dosyası DEĞİL) `kaynak_klasoru` None geçilir ve
    bu adım atlanır."""
    veri = _excel_bayt_uret(sonuclar)
    damga = time.strftime("%Y%m%d_%H%M%S")
    dosya_ad = f"{taban_ad}_Barkod_Sonuc_{damga}.xlsx"

    if kaynak_klasoru:
        try:
            cikti_yolu = os.path.join(kaynak_klasoru, dosya_ad)
            with open(cikti_yolu, "wb") as f:
                f.write(veri)
            log(f"📊 Excel çıktısı diske de yazıldı: {cikti_yolu}")
        except OSError:
            pass  # web'de girdi geçici klasördeyse beklenen durum — b64 zaten yeterli

    return dosya_ad, veri


# ── Veritabanı entegrasyonu ("Dosyalarım (Tümü)" ile AYNI kalıcı PostgreSQL) ─
# icra_core.py'nin IcraSorgu.ara() içinde kullandığı desenin AYNISI (django
# setup + gevşek hata yutma) — o sorgu Birim/Dosya kapak künyesini zaten bu
# veritabanına yazıyor; burada AYNI Dosya satırına (doğal anahtarla upsert
# edilerek) barkod sonucu bağlanır.
_DB_MODULLERI = None  # None=henüz denenmedi, False=kullanılamıyor, dict=hazır


def _db_baglantisini_kur(log):
    """icra_models Django uygulamasına bağlanır. PostgreSQL kapalı/erişilemezse
    sessizce None döner — barkod sorgusu yine de canlı UYAP+PTT üzerinden
    çalışmaya devam eder, yalnız DB kaydı atlanır."""
    global _DB_MODULLERI
    if _DB_MODULLERI is not None:
        return _DB_MODULLERI or None
    _DB_MODULLERI = False
    try:
        # Bağlantı bootstrap'i artık dosya_core._django_hazirla()'dan gelir —
        # aynı sys.path/DJANGO_SETTINGS_MODULE/django.setup() bloğu 5 ayrı
        # dosyada tekrarlanıyordu (kullanıcı isteği, 2026-08-13: bağlantı
        # kodu sadeleştirme). icra_core.py KASTEN dokunulmadan bırakıldı
        # (bkz. dosya_core.py başlığı) — yalnız burası, dosya_core'un zaten
        # modül seviyesinde import edilmiş olan hazır fonksiyonunu kullanır.
        dosya_core._django_hazirla()
        from icra_models.models import Dosya, DosyaTaraf, TebligatBarkod
        from icra_models.ingest import dosya_kunyesi_kaydet
        Dosya.objects.filter(pk=1).exists()  # bağlantıyı doğrula
        _DB_MODULLERI = {
            "TebligatBarkod": TebligatBarkod,
            "DosyaTaraf": DosyaTaraf,
            "dosya_kunyesi_kaydet": dosya_kunyesi_kaydet,
        }
    except Exception as e:
        log(f"⚠️ Yerel veritabanı aktif değil (barkod sonuçları DB'ye yazılmayacak): {e}")
    return _DB_MODULLERI or None


def _borclu_coz_tekil(dosya_obj):
    """Dosyada rol=borclu olan DosyaTaraf sayısı TEK ise onu döner, aksi
    halde (çok borçlulu ya da hiç borçlusuz dosya) None döner. Çok borçlulu
    dosyada bir tebligat evrakının HANGİ borçluya gittiğini ayırt eden CANLI
    doğrulanmış bir UYAP alanı/desen henüz yok (kullanıcı sorusu, 2026-08-04)
    — bu yüzden orada UYDURULMAZ, TebligatBarkod.borclu boş kalır."""
    if not dosya_obj:
        return None
    try:
        borclular = list(dosya_obj.taraf_baglari.filter(rol="borclu"))
    except Exception:
        return None
    return borclular[0] if len(borclular) == 1 else None


def _barkodu_veritabanina_kaydet(db, dosya_obj, evrak, log, **degerler):
    """Tek bir evrak/barkod sonucunu TebligatBarkod'a upsert eder — KALICI
    anahtar (dosya, birim_evrak_no) [uq_dosya_tebligat_barkod]. `dosya_obj`
    ÖNCEDEN çözülmüş bir Django `Dosya` kaydıdır (bkz. `_dosya_isle` — dosya
    başına BİR KEZ `dosya_kunyesi_kaydet` ile çözülür, evrak döngüsünde
    tekrar tekrar çözülmez). `db=None`/`dosya_obj=None` (bağlantı ya da künye
    yok) veya `birimEvrakNo` eksikse sessizce atlanır."""
    if not db or not dosya_obj:
        return
    birim_evrak_no = evrak.get("birimEvrakNo")
    if not birim_evrak_no:
        return
    try:
        db["TebligatBarkod"].objects.update_or_create(
            dosya=dosya_obj, birim_evrak_no=int(birim_evrak_no),
            defaults=degerler,
        )
    except Exception as e:
        log(f"  ⚠️ DB'ye kaydedilemedi: {e}")


def _bilinen_barkod_kaydini_al(db, dosya_obj, birim_evrak_no):
    """Bu evrak için DAHA ÖNCE kaydedilmiş bir TebligatBarkod satırı varsa
    döner (kullanıcı bulgusu, 2026-08-03: barkod bir kez atanır, değişmez —
    bkz. modül başlığı) — yoksa ya da DB/dosya_obj/birim_evrak_no eksikse
    None döner (çağıran bu durumda PDF'i normal şekilde indirip işler)."""
    if not db or not dosya_obj or not birim_evrak_no:
        return None
    try:
        return db["TebligatBarkod"].objects.filter(
            dosya=dosya_obj, birim_evrak_no=int(birim_evrak_no)).first()
    except Exception:
        return None


def _dosya_isle(motor, birim_adi, dosya_no, log, _ekle, db, rec_bilinen=None):
    """Tek bir (birim, dosya no) için 'Kapalı Tebligat' evrak(lar)ının barkodunu
    bulup PTT'de sorgular; sonuç satır(lar)ını `_ekle` ile biriktirir (ve
    DB bağlantısı varsa `db` üzerinden TebligatBarkod'a da yazar).
    `rec_bilinen` verilmişse (tarih aralığı modu — arama sonucunda zaten
    geliyor, dosyaId DAHİL) search_phrase_detayli.ajx ile YENİDEN çözülmez;
    Excel modunda (yalnız birim+dosya no varsa) None geçilip aşağıda çözülür."""
    temel = {"Birim": birim_adi, "Dosya No": dosya_no}
    try:
        rec = rec_bilinen
        dosya_id = rec.get("dosyaId") if rec else None
        if not dosya_id:
            yil, sira = _dosya_no_ayir(dosya_no)
            if yil is None:
                log(f"  ⚠️ Dosya no formatı tanınmadı: {dosya_no!r}")
                _ekle(_satir_olustur(temel, Durum="⚠️ Atlandı (dosya no formatı)"), "atlanan")
                return

            birim_id = icra_core.birim_id_bul(birim_adi)
            if not birim_id:
                log(f"  ⚠️ Birim bulunamadı: {birim_adi!r}")
                _ekle(_satir_olustur(temel, Durum="⚠️ Atlandı (birim bulunamadı)"), "atlanan")
                return

            rec = _dosya_id_coz(motor, birim_id, yil, sira, log)
            dosya_id = rec.get("dosyaId") if rec else None
            if not dosya_id:
                log("  ⚠️ Dosya bulunamadı.")
                _ekle(_satir_olustur(temel, Durum="⚠️ Atlandı (dosya bulunamadı)"), "atlanan")
                return

        evraklar = dosya_core.evrak_listesi_getir(
            dosya_id, log_fn=log, istek_sarici=dosya_core._arka_plan_istek)

        tm_evrak = next((e for e in evraklar if e.get("tur") == TEBLIG_MAZBATASI_TUR), None)
        tm_aciklama = (tm_evrak.get("aciklama") or "Var") if tm_evrak else ""
        temel["Tebliğ Mazbatası"] = tm_aciklama if tm_evrak else "Yok"
        log(f"  📎 Tebliğ Mazbatası: {temel['Tebliğ Mazbatası']}")

        ktm_anahtar = _tur_anahtar(KAPALI_TEBLIG_MAZBATASI_TUR)
        ktm_var = any(_tur_anahtar(e.get("tur")) == ktm_anahtar for e in evraklar)
        temel["Kapalı Tebliğ Mazbatası"] = "Var" if ktm_var else "Yok"

        # Yeniden tebliğ (21/2 vb.) talebi zaten gönderilmiş mi — evrak listesi
        # zaten burada elimizde, ekstra istek GEREKMEZ (kullanıcı isteği,
        # 2026-08-13).
        yeniden_teblig_talebi_var = any(e.get("tur") == TALEP_EVRAK_TUR for e in evraklar)
        temel["Yeniden Tebliğ Talebi"] = "Var" if yeniden_teblig_talebi_var else "Yok"

        eslesen = [e for e in evraklar if e.get("tur") == HEDEF_TUR]
        if not eslesen:
            log(f"  ℹ️ '{HEDEF_TUR}' evrakı yok ({len(evraklar)} evrak tarandı).")
            _ekle(_satir_olustur(temel, Durum=f"ℹ️ '{HEDEF_TUR}' yok"), "atlanan")
            return

        # Dosya başına BİR KEZ çözülür (evrak döngüsünde tekrar tekrar
        # çözülmez — bkz. `_barkodu_veritabanina_kaydet` docstring'i).
        dosya_obj = None
        if db:
            try:
                dosya_obj, _ = db["dosya_kunyesi_kaydet"](rec)
            except Exception as e:
                log(f"  ⚠️ Dosya künyesi kaydedilemedi (DB kaydı bu dosya için atlanacak): {e}")
        # Yalnız dosyada TEK borçlu varsa otomatik atanır (bkz.
        # `_borclu_coz_tekil` docstring'i) — çok borçlulu dosyada boş kalır.
        borclu_obj = _borclu_coz_tekil(dosya_obj)

        # ASLA GERİYE DÜŞMEZ (kullanıcı bulgusu, 2026-08-17: "gönderdim ama
        # Yeniden Tebliğ Talebi durumu Var'a dönmedi" — az önce GERÇEKTEN
        # gönderilmiş (bkz. teblig_21_2_core.dosyalari_gonderildi_isaretle,
        # finalize()'ın KESİN başarı yanıtına dayanır) bir dosyada, BU tarama
        # yeniden_teblig_talebi_var'ı BURADA (satır ~944) UYAP'ın evrak
        # listesinden yeniden hesaplarken, o evrak UYAP'ın listesine henüz
        # YANSIMAMIŞ olabilir — "şu an evrak listesinde YOK" (False) "hiç
        # gönderilmedi" ANLAMINA GELMEZ. Aşağıdaki _barkodu_veritabanina_
        # kaydet çağrıları bu alanı defaults= ile HER taramada YENİDEN YAZAR
        # (update_or_create), yani bir SONRAKİ taramanın yanlış-negatif
        # vermesi, az önceki DOĞRU "gönderildi" kaydını SESSİZCE siliyordu.
        # Burada: veritabanında BU dosya için ZATEN True olan bir kayıt varsa
        # canlı sonuç False çıksa bile geriye düşürülmez.
        if db and dosya_obj and not yeniden_teblig_talebi_var:
            try:
                if db["TebligatBarkod"].objects.filter(
                        dosya=dosya_obj, yeniden_teblig_talep_edildi=True).exists():
                    yeniden_teblig_talebi_var = True
                    temel["Yeniden Tebliğ Talebi"] = "Var"
            except Exception:
                pass

        # barkod -> (tarih, açıklama) — yalnız GERÇEKTEN bir elektronik evrakla
        # karşılaşılınca (aşağıda) doldurulur (bkz. `_e_teblig_haritasini_olustur`);
        # dosyada hiç elektronik tebligat yoksa hiç mazbata indirilmez.
        e_teblig_haritasi = None

        for evrak in eslesen:
            evrak_id = evrak.get("evrakId")
            birim_evrak_no = evrak.get("birimEvrakNo")
            aciklama = evrak.get("aciklama", "") or ""
            temel_evrak = {**temel, "Evrak Açıklaması": aciklama,
                            "Evrak Tarihi": evrak.get("onaylandigiTarih", "")}
            try:
                bilinen = _bilinen_barkod_kaydini_al(db, dosya_obj, birim_evrak_no)
                # Kullanıcı bulgusu (2026-08-04, YEDİNCİ tur — ÖNEMLİ bug):
                # "barkod SONUCU elinde varsa tekrar sorgulama" kullanıcının
                # kendi isteğiydi, ama önceki kod bunu YALNIZ `bilinen.barkod`
                # dolu mu diye kontrol ederek uyguluyordu — `tebligat_turu`
                # BOŞ kalmış olsa BİLE "sonuç elimde" sayılıp PDF bir daha HİÇ
                # indirilmiyordu. aciklama-only sınıflandırma (`pdf_govde_metni`
                # olmadan çağrılan `_tebligat_turunu_belirle`) çoğu evrakta
                # (Maaş/Banka'nın zarf-içeriği tespiti, "ihtiva eder" yakalama
                # dahil TÜM gövde-metni-gerektiren kurallar) HİÇBİR ZAMAN
                # sonuç üretemez — bu yüzden ilk turda türü boş kalan bir evrak
                # kaç kez "güncelleme" çalıştırılırsa çalıştırılsın SONSUZA DEK
                # boş kalıyordu (dosya "atlanıyormuş" gibi görünüyordu, aslında
                # her turda işleniyordu ama körlemesine — PDF'e hiç bakmadan).
                # Şart artık `bilinen.tebligat_turu` DOLU olmasını da istiyor:
                # tür hâlâ boşsa PDF'i (aşağıdaki "yavaş yol") yeniden indirip
                # gövde metniyle YENİDEN sınıflandırmayı dener; barkod'un
                # kendisi değişmez kuralına aykırı değildir (aynı PDF'ten aynı
                # barkod yeniden çıkar, yalnız fazladan bir istek maliyeti var).
                if bilinen and bilinen.barkod and bilinen.tebligat_turu:
                    barkod = bilinen.barkod
                    log(f"  ♻️ Barkod+tür zaten biliniyor ({barkod}) — PDF tekrar indirilmedi, PTT durumu tazeleniyor.")
                    tebligat_turu = bilinen.tebligat_turu
                    if bilinen.elektronik_tebligat:
                        if e_teblig_haritasi is None:
                            e_teblig_haritasi = _e_teblig_haritasini_olustur(motor, dosya_id, evraklar, log)
                        e_teblig = e_teblig_haritasi.get(barkod)
                        if e_teblig:
                            tarih, mz_aciklama = e_teblig
                            ozet, ptt_veri = f"E-Tebligat: {mz_aciklama}", None
                            log(f"  ✅ E-Tebliğ Mazbatası: {ozet} ({tarih})")
                        else:
                            ozet, tarih, ptt_veri = "Elektronik Tebligat (PTT'de takip yok)", "", None
                    else:
                        ptt_veri = ptt_sorgula(barkod, log)
                        ozet, tarih = _ptt_durum_ve_tarih(ptt_veri)
                        log(f"  ✅ PTT: {ozet}" + (f" ({tarih})" if tarih else ""))
                    _ekle(_satir_olustur(
                        temel_evrak, Barkod=barkod,
                        **{"PTT Durumu": ozet, "Son İşlem Tarihi": tarih, "Tebligat Türü": tebligat_turu},
                        Durum="✅ Sorgulandı (barkod önbellekten)",
                        **{"Ham Yanıt": json.dumps(ptt_veri, ensure_ascii=False) if isinstance(ptt_veri, dict) else ""}),
                        "basarili")
                    _barkodu_veritabanina_kaydet(
                        db, dosya_obj, evrak, log,
                        evrak_aciklama=aciklama, evrak_tarihi=str(evrak.get("onaylandigiTarih", "") or ""),
                        barkod=barkod, elektronik_tebligat=bilinen.elektronik_tebligat,
                        ptt_durumu=ozet, son_islem_tarihi=tarih, tebligat_turu=tebligat_turu,
                        ham_yanit=ptt_veri if isinstance(ptt_veri, dict) else None,
                        tebligat_mazbatasi_var=tm_evrak is not None,
                        tebligat_mazbatasi_aciklama=tm_aciklama,
                        kapali_tebligat_mazbatasi_var=ktm_var, borclu=borclu_obj,
                        yeniden_teblig_talep_edildi=yeniden_teblig_talebi_var)
                    continue

                _, pdf_bytes = _evrak_pdf_indir(
                    motor, dosya_id, evrak_id, log_fn=log, istek_sarici=dosya_core._arka_plan_istek)
                if bilinen and bilinen.barkod:
                    # Barkod zaten biliniyor (yalnız tür eksikti, bkz. yukarı)
                    # — "barkod bir kez atanır, değişmez" kuralı gereği YENİDEN
                    # ÇIKARILMAZ, PDF yalnız gövde metni için indirildi.
                    barkod = bilinen.barkod
                    log(f"  ♻️ Barkod zaten biliniyor ({barkod}) — tür için PDF yeniden indirildi.")
                else:
                    barkod = barkod_bul(pdf_bytes, log)
                    if not barkod:
                        _ekle(_satir_olustur(temel_evrak, Durum="⚠️ Barkod bulunamadı"), "atlanan")
                        continue

                # Gövde metni BİR KEZ çıkarılır — tebligat türü sınıflandırması
                # bunu kullanır (bkz. `_tebligat_turunu_belirle`). Kullanıcı
                # bulgusu (2026-08-03, altıncı tur): burada eskiden bir
                # "belge gövdesinde borçlu adı yoksa DB'ye YAZILMAZ" filtresi
                # vardı — KALDIRILDI (bkz. modül başlığı): indirilen PDF genel
                # bir zarf/mazbata şablonu olduğundan bankaya/işverene giden
                # haciz ihbarnamelerinde borçlu adı bu metinde YAPISAL OLARAK
                # hiç geçmiyor, filtre bunları %100 dışlıyordu.
                govde_metni = _pdf_govde_metni(pdf_bytes, log)
                tebligat_turu = _tebligat_turunu_belirle(aciklama, govde_metni)

                log(f"  📮 Barkod: {barkod}" + (f" — Tür: {tebligat_turu}" if tebligat_turu else ""))
                if ELEKTRONIK_ANAHTAR in aciklama:
                    # bkz. modül başlığı — e-Tebligat fiziki postaya hiç
                    # çıkmadığından PTT'de asla kaydı olmaz; ama GERÇEK tebliğ
                    # tarihi/durumu ayrı bir "Kapalı E-Tebliğ Mazbatası"
                    # evrakında var (kullanıcı bulgusu, yedinci tur — CANLI
                    # doğrulandı) — bkz. `_e_teblig_haritasini_olustur`.
                    if e_teblig_haritasi is None:
                        e_teblig_haritasi = _e_teblig_haritasini_olustur(motor, dosya_id, evraklar, log)
                    e_teblig = e_teblig_haritasi.get(barkod)
                    if e_teblig:
                        tarih, mz_aciklama = e_teblig
                        ozet = f"E-Tebligat: {mz_aciklama}"
                        log(f"  ✅ E-Tebliğ Mazbatası: {ozet} ({tarih})")
                    else:
                        ozet, tarih = "Elektronik Tebligat (PTT'de takip yok)", ""
                    _ekle(_satir_olustur(
                        temel_evrak, Barkod=barkod,
                        **{"PTT Durumu": ozet, "Son İşlem Tarihi": tarih, "Tebligat Türü": tebligat_turu},
                        Durum="ℹ️ Elektronik"), "basarili")
                    _barkodu_veritabanina_kaydet(
                        db, dosya_obj, evrak, log,
                        evrak_aciklama=aciklama, evrak_tarihi=str(evrak.get("onaylandigiTarih", "") or ""),
                        barkod=barkod, elektronik_tebligat=True,
                        ptt_durumu=ozet, son_islem_tarihi=tarih,
                        tebligat_turu=tebligat_turu, ham_yanit=None,
                        tebligat_mazbatasi_var=tm_evrak is not None,
                        tebligat_mazbatasi_aciklama=tm_aciklama,
                        kapali_tebligat_mazbatasi_var=ktm_var, borclu=borclu_obj,
                        yeniden_teblig_talep_edildi=yeniden_teblig_talebi_var)
                    continue

                ptt_veri = ptt_sorgula(barkod, log)
                ozet, tarih = _ptt_durum_ve_tarih(ptt_veri)
                log(f"  ✅ PTT: {ozet}" + (f" ({tarih})" if tarih else ""))
                _ekle(_satir_olustur(
                    temel_evrak, Barkod=barkod,
                    **{"PTT Durumu": ozet, "Son İşlem Tarihi": tarih, "Tebligat Türü": tebligat_turu},
                    Durum="✅ Sorgulandı",
                    **{"Ham Yanıt": json.dumps(ptt_veri, ensure_ascii=False)}), "basarili")
                _barkodu_veritabanina_kaydet(
                    db, dosya_obj, evrak, log,
                    evrak_aciklama=aciklama, evrak_tarihi=str(evrak.get("onaylandigiTarih", "") or ""),
                    barkod=barkod, elektronik_tebligat=False,
                    ptt_durumu=ozet, son_islem_tarihi=tarih, tebligat_turu=tebligat_turu,
                    ham_yanit=ptt_veri if isinstance(ptt_veri, dict) else None,
                    tebligat_mazbatasi_var=tm_evrak is not None,
                    tebligat_mazbatasi_aciklama=tm_aciklama,
                    kapali_tebligat_mazbatasi_var=ktm_var, borclu=borclu_obj,
                    yeniden_teblig_talep_edildi=yeniden_teblig_talebi_var)

            except Exception as e:
                log(f"  ❌ Hata: {e}")
                _ekle(_satir_olustur(temel_evrak, Durum=f"❌ Hata: {e}"), "hatali")

        # Bu dosyanın Ödeme Emri tebligatı yeni işlendiğine göre kesinleşme
        # durumu da tazelenir (bkz. dosya_core.kesinlesme_durumlarini_guncelle
        # docstring'i) — kullanıcı bulgusu, 2026-08-04: kesinleşme sütunu
        # hiçbir akışta hesaplanmıyordu.
        if db and dosya_obj:
            try:
                dosya_core.kesinlesme_durumlarini_guncelle(dosya_obj=dosya_obj, log_fn=log)
            except Exception as e:
                log(f"  ⚠️ Kesinleşme durumu güncellenemedi: {e}")

    except Exception as e:
        log(f"  ❌ Hata: {e}")
        _ekle(_satir_olustur(temel, Durum=f"❌ Hata: {e}"), "hatali")


def calistir(girdi, log_fn=None, kontrol=None):
    """Ana giriş noktası — üç modu destekler:
      * Excel modu: `girdi` bir dosya yolu (str) — Excel'deki (Birim, Dosya No)
        satırları okunur (bkz. `excel_oku`).
      * Tarih aralığı modu: `girdi` bir sözlük ({"birim_adi","baslangic","bitis"}
        — bkz. PARAMETRELER) — UYAP'ın "Dosya Sorgulama" ekranındaki tarih
        aralığı filtresiyle AYNI şekilde o aralıkta açılmış dosyalar listelenir
        (bkz. `_tarih_araligi_dosyalari_getir`); dosyaId zaten yanıtta geldiğinden
        AYRICA search_phrase_detayli.ajx ile çözülmez.
      * Seçili dosyalar modu: `girdi` bir liste — Panel'in "Dosyalarım (Tümü)"
        tarzı yerel DB listesinden (bkz. dosya_core.dosyalarim_db_listele)
        kullanıcının seçtiği satırlar, her biri en az {"birimAdi","dosyaNo"}
        içeren bir sözlük. BİLEREK `rec_bilinen` OLUŞTURULMAZ: bu sözlükler
        UYAP'ın ham search_phrase_detayli.ajx kaydıyla AYNI BİÇİMDE DEĞİL
        (ör. birimTuru1/2/3 yok, tarih alanı başka ad taşıyor) — `_dosya_isle`'ye
        böyle geçirilirse `dosya_kunyesi_kaydet` var olan doğru
        Birim.turu1/2/3 ve Dosya.acilis_tarihi alanlarını SESSİZCE boşaltırdı;
        bu yüzden her satır Excel modundaki gibi `_dosya_id_coz` ile YENİDEN
        çözülür (aşağıda `onceden_cozulmus` bu modda hiç doldurulmaz).
    Her modda da her dosya için 'Kapalı Tebligat' evrak(lar)ının barkodu
    bulunup PTT'de sorgulanır. Döner: satır başına sözlük listesi (Panel'de
    tablo olarak gösterilir); işlem bitince bir .xlsx çıktısı da üretilir
    (bkz. `_sonucu_excele_yaz`)."""
    log = log_fn or print
    icra_core.birim_listesi_getir(log)  # birim adı -> id önbelleğini doldur
    motor = SorguMotoru(log)
    db = _db_baglantisini_kur(log)  # None ise DB kaydı sessizce atlanır

    onceden_cozulmus = {}  # (birim, dosya_no) -> ham UYAP kaydı — yalnız tarih aralığı modunda dolu
    if isinstance(girdi, list):
        satirlar = []
        for rec in girdi:
            b = str(rec.get("birimAdi") or rec.get("Birim") or "").strip()
            no = str(rec.get("dosyaNo") or rec.get("Dosya No") or "").strip()
            if b and no:
                satirlar.append((b, no))
        log(f"🗂️ {len(satirlar)} seçili dosya işlenecek")
        taban_ad = "Barkod_Sorgu_Secili_Dosyalar"
        kaynak_klasoru = None
    elif isinstance(girdi, dict):
        birim_adi_g = (girdi.get("birim_adi") or "").strip()
        baslangic = (girdi.get("baslangic") or "").strip()
        bitis = (girdi.get("bitis") or "").strip()
        kayitlar = _tarih_araligi_dosyalari_getir(motor, birim_adi_g, baslangic, bitis, log)
        satirlar = []
        for rec in kayitlar:
            dosya_id = rec.get("dosyaId")
            dosya_no = rec.get("dosyaNo")
            if not dosya_id or not dosya_no:
                continue
            b = rec.get("birimAdi") or rec.get("birim_adi") or birim_adi_g
            anahtar = (b, str(dosya_no))
            satirlar.append(anahtar)
            onceden_cozulmus[anahtar] = rec
        taban_ad = f"Barkod_Sorgu_{birim_adi_g or 'TarihAraligi'}".replace(" ", "_")
        kaynak_klasoru = None
    else:
        satirlar = excel_oku(girdi)
        log(f"📄 {len(satirlar)} satır okundu")
        taban_ad = os.path.splitext(os.path.basename(girdi))[0]
        kaynak_klasoru = os.path.dirname(os.path.abspath(girdi))

    sonuclar = []
    sayac = {"basarili": 0, "atlanan": 0, "hatali": 0}

    def _ekle(satir, anahtar):
        sayac[anahtar] += 1
        sonuclar.append(satir)

    for i, (birim_adi, dosya_no) in enumerate(satirlar, 1):
        if kontrol:
            kontrol.tur_bitti()
            if not kontrol.nokta():
                log("⏹ Durduruldu.")
                break
        log(f"\n▶ [{i}/{len(satirlar)}] {birim_adi} {dosya_no}")
        _dosya_isle(motor, birim_adi, dosya_no, log, _ekle, db,
                    onceden_cozulmus.get((birim_adi, dosya_no)))
        time.sleep(SATIR_ARASI_SN)

    log(f"\n{'=' * 60}\n"
        f"✅ İşlenen: {sayac['basarili']}   ⚠️ Atlanan: {sayac['atlanan']}   ❌ Hatalı: {sayac['hatali']}")

    cikti = _SonucListesi(sonuclar)
    if sonuclar:
        try:
            dosya_ad, veri = _sonucu_excele_yaz(sonuclar, taban_ad, kaynak_klasoru, log)
            cikti.dosya_ad = dosya_ad
            cikti.dosya_b64 = base64.b64encode(veri).decode("ascii")
        except Exception as e:
            log(f"  ⚠️ Excel çıktısı üretilemedi: {e}")

    return cikti


def excel_isle(excel_yolu, log_fn=print, kontrol=None):
    """Panel runner'ının (EXCEL_GIRDI) çağırdığı giriş noktası — calistir()
    ile birebir aynı, yalnızca imza runner'ın çağrı biçimiyle uyumlu."""
    return calistir(excel_yolu, log_fn, kontrol=kontrol)


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    _varsayilan_xlsx = os.path.normpath(os.path.join(_HERE, "..", "..", "Dosya Sorgulama.xlsx"))
    _xlsx = sys.argv[1] if len(sys.argv) > 1 else _varsayilan_xlsx
    for _satir in calistir(_xlsx, log_fn=print):
        print(_satir)
