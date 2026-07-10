# -*- coding: utf-8 -*-
"""
SGK Sorgu — paylaşılan headless motor (GUI + web ortak)
=======================================================
ÇALIŞMA MANTIĞI orijinal "Sorgu/SGK Sorgu/sgk_sorgu_gui.py" dosyasından AYNEN
import edilip kullanılır; UYAP ağ/ayrıştırma/özetleme mantığı YENİDEN YAZILMAZ:
  • SorguMotoru (127.0.0.1:8800 ofis proxy'sine .ajx istekleri)
  • hucre_metni / ozetle / degerlendir_ssk_calisani / tam_metin (özetleyiciler)
  • SORGULAR, SONUC_BASLANGIC_KOL, HATA_ON, SIRKET_NOT, sabitler

Bu sınıf yalnızca tkinter GUI sınıfının içindeki "worker + Excel + kaydetme"
orkestrasyonunu — algoritması BİREBİR korunarak — arayüzden bağımsız (headless)
hale getirir. Böylece aynı motor hem masaüstü GUI'de hem web sunucusunda çalışır.
"""

import os
import sys
import time
import threading

import openpyxl

# Orijinal SGK modülünün bulunduğu klasörü import yoluna ekle
_HERE = os.path.dirname(os.path.abspath(__file__))
_SGK_DIR = os.path.normpath(os.path.join(_HERE, "..", "..", "Sorgu", "SGK Sorgu"))
if _SGK_DIR not in sys.path:
    sys.path.insert(0, _SGK_DIR)

import sgk_sorgu_gui as _sgk  # noqa: E402  (orijinal — değiştirilmez)

# Orijinalden AYNEN alınan motor + sabitler (yeniden yazılmaz)
SorguMotoru = _sgk.SorguMotoru
OturumHatasi = _sgk.OturumHatasi
SORGULAR = _sgk.SORGULAR
SONUC_BASLANGIC_KOL = _sgk.SONUC_BASLANGIC_KOL
HATA_ON = _sgk.HATA_ON
SIRKET_NOT = _sgk.SIRKET_NOT
SIRKET_ANAHTARLARI = _sgk.SIRKET_ANAHTARLARI
tr_normalize = _sgk.tr_normalize
SATIR_ARASI_BEKLEME = _sgk.SATIR_ARASI_BEKLEME
MOLA_HATA_ESIGI = _sgk.MOLA_HATA_ESIGI
MOLA_SURESI = _sgk.MOLA_SURESI
OTURUM_KONTROL_ESIGI = _sgk.OTURUM_KONTROL_ESIGI
YAPILANLAR_EKI = _sgk.YAPILANLAR_EKI
OZET_EKI = _sgk.OZET_EKI

KOLONLAR = ["No", "Ad Soyad", "Birim", "Dosya No"] + [s[0] for s in SORGULAR]
SORGU_ANAHTARLARI = [s[2] for s in SORGULAR]          # ["kamuCalisani", ...]
SORGU_BASLIKLARI = [(s[2], s[0]) for s in SORGULAR]    # [(anahtar, başlık), ...]


def _kisa(metin):
    """Tabloda göstermek için tek satıra indir (orijinal _kisa ile aynı)."""
    if metin is None:
        return ""
    t = str(metin).split("\n", 1)[0]
    return (t[:90] + "…") if len(t) > 90 else t


class SgkBatch:
    """Headless SGK toplu sorgu motoru. Orijinal SgkSorguGUI worker'ının
    arayüzden arındırılmış, algoritması birebir korunmuş eşi."""

    def __init__(self, log_fn=None):
        self._ext_log = log_fn
        self.kolonlar = list(KOLONLAR)

        self.kaynak_yolu = None
        self.excel_yolu = None
        self.ozet_yolu = None
        self.wb = None
        self.ws = None

        self.satirlar = []           # veri satırlarının excel satır numaraları (sıralı)
        self.row_state = {}          # r -> {"base":[..4], "results":[..7], "tag":""}
        self.row_rev = {}            # r -> rev (artımlı güncelleme için)
        self.rev = 0
        self.log_lines = []
        self.status_text = ""

        self.running = False
        self._paused = False
        self._stop = False
        self._lock = threading.Lock()

    # ─────────────────────────── günlük / durum ───────────────────────────
    def log(self, mesaj):
        with self._lock:
            self.log_lines.append(str(mesaj))
        if self._ext_log:
            try:
                self._ext_log(str(mesaj))
            except Exception:
                pass

    def _durum(self, text):
        self.status_text = text

    def _touch(self, r):
        self.rev += 1
        self.row_rev[r] = self.rev

    # ─────────────────────────── Excel yükleme (orijinal excel_yukle mantığı) ───────────────────────────
    def load(self, secilen_yol):
        """Excel'i yükler; varsa '_yapilanlar' dosyasından DEVAM eder. Seçilen
        kök dosya hiç değiştirilmez. Döner: yüklenen satır sayısı."""
        kok, uzanti = os.path.splitext(secilen_yol)
        for ek in (OZET_EKI, YAPILANLAR_EKI):
            if kok.endswith(ek):
                kok = kok[: -len(ek)]
                break
        cikti_yolu = kok + YAPILANLAR_EKI + uzanti
        ozet_yolu = kok + OZET_EKI + uzanti

        if os.path.exists(cikti_yolu):
            self.wb = openpyxl.load_workbook(cikti_yolu)
            self.log(f"↩️ Yapılanlar dosyasından devam ediliyor: {os.path.basename(cikti_yolu)}")
        else:
            self.wb = openpyxl.load_workbook(secilen_yol)
        self.ws = self.wb.active

        # Başlık satırı yoksa ekle (orijinal davranış)
        if not self._baslik_satiri_mi(self.ws):
            self.ws.insert_rows(1)
            for c, ad in enumerate(self.kolonlar, start=1):
                self.ws.cell(1, c, value=ad)

        self.kaynak_yolu = secilen_yol
        self.excel_yolu = cikti_yolu
        self.ozet_yolu = ozet_yolu
        self._tabloyu_doldur()
        self.save()  # başlıkları/biçimi hemen yapılanlar dosyasına yaz
        self.log(f"📂 Yüklendi: {os.path.basename(secilen_yol)}  ({len(self.satirlar)} satır)")
        self.log(f"💾 Tam veri → {os.path.basename(cikti_yolu)}")
        self.log(f"💾 Özet → {os.path.basename(ozet_yolu)}")
        return len(self.satirlar)

    def _baslik_satiri_mi(self, ws):
        try:
            deg = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, 5)]
            return deg == [b.lower() for b in self.kolonlar[:4]]
        except Exception:
            return False

    def _tabloyu_doldur(self):
        self.satirlar = []
        self.row_state = {}
        self.row_rev = {}
        for r in range(2, self.ws.max_row + 1):
            temel = [self.ws.cell(r, c).value for c in range(1, 5)]
            if all(v is None for v in temel):
                continue
            results = [_kisa(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                       for i in range(len(SORGULAR))]
            self.satirlar.append(r)
            self.row_state[r] = {
                "base": ["" if v is None else str(v) for v in temel],
                "results": results,
                "tag": "",
            }
            self._touch(r)

    # ─────────────────────────── satır durumu yardımcıları (orijinal) ───────────────────────────
    @staticmethod
    def _bos(v):
        return v is None or str(v).strip() == ""

    def _hata_cell(self, v):
        return (not self._bos(v)) and str(v).lstrip().startswith(HATA_ON)

    def gereken_anahtarlar(self, r, mode, secili):
        """Bu satırda hangi sorguların çalışması gerektiğini belirler (orijinal mantık)."""
        gerekli = set()
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if anahtar not in secili:
                continue
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if mode == "retry":
                if self._hata_cell(v):
                    gerekli.add(anahtar)
            else:
                if self._bos(v) or self._hata_cell(v):
                    gerekli.add(anahtar)
        return gerekli

    @staticmethod
    def _sirket_mi(isim):
        if not isim:
            return False
        kelimeler = set(tr_normalize(isim).split())
        return bool(kelimeler & SIRKET_ANAHTARLARI)

    def _satir_tamamlandi_mi(self, r, secili):
        secili_var = False
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if anahtar not in secili:
                continue
            secili_var = True
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value
            if self._bos(v) or self._hata_cell(v):
                return False
        return secili_var

    # ─────────────────────────── kontroller ───────────────────────────
    def start(self, mode="normal", secili=None):
        if self.running or not self.ws:
            return False
        secili = set(secili) if secili else set(SORGU_ANAHTARLARI)
        if mode == "retry":
            toplam = sum(len(self.gereken_anahtarlar(r, "retry", secili)) for r in self.satirlar)
            if toplam == 0:
                self.log("Tekrar denenecek hatalı hücre yok.")
                return False
        self.running = True
        self._stop = False
        self._paused = False
        threading.Thread(target=self._worker, args=(mode, secili), daemon=True).start()
        return True

    def pause(self):
        if self.running:
            self._paused = True

    def resume(self):
        self._paused = False

    def toggle_pause(self):
        if self.running:
            self._paused = not self._paused
        return self._paused

    def stop(self):
        self._stop = True
        self._paused = False
        self._durum("Durduruluyor...")

    @property
    def paused(self):
        return self._paused

    def _bekle_devam(self):
        """Duraklatma boyunca bekler; durdurulduysa False döner (orijinal)."""
        while self._paused and not self._stop:
            self._durum("⏸ Duraklatıldı")
            time.sleep(0.3)
        return not self._stop

    # ─────────────────────────── worker (orijinal _sorgu_worker birebir) ───────────────────────────
    def _worker(self, mode, secili):
        motor = SorguMotoru(self.log)
        try:
            motor.baslat()
        except Exception as e:
            self.log(f"❌ Başlatma hatası: {e}")
            motor.kapat()
            self.running = False
            return

        satirlar = list(self.satirlar)
        toplam = len(satirlar)
        offset = {anahtar: i for i, (_, _, anahtar) in enumerate(SORGULAR)}
        hata_sayaci = 0
        pespese = 0
        oturum_yenilendi = False

        def oturum_yenile():
            nonlocal oturum_yenilendi, pespese
            oturum_yenilendi = True
            pespese = 0
            try:
                motor.yeniden_baglan()
                return True
            except Exception as e:
                self.log(f"❌ Oturum yenilenemedi: {e}")
                return False

        def mola_kontrol():
            nonlocal hata_sayaci
            if hata_sayaci >= MOLA_HATA_ESIGI:
                hata_sayaci = 0
                self.log(f"⏳ {MOLA_HATA_ESIGI} hata oluştu, sunucu yükünü azaltmak için "
                         f"{MOLA_SURESI} sn mola veriliyor...")
                bekleyen = MOLA_SURESI
                while bekleyen > 0 and not self._stop:
                    self._durum(f"⏳ Otomatik mola: {bekleyen} sn")
                    time.sleep(1)
                    bekleyen -= 1
                if not self._stop:
                    self.log("▶ Molaya devam ediliyor.")
            return not self._stop

        for idx, r in enumerate(satirlar, 1):
            if not self._bekle_devam():
                break
            if not mola_kontrol():
                break
            self._durum(f"İşleniyor {idx}/{toplam}  (satır {r})")

            gereken = self.gereken_anahtarlar(r, mode, secili)
            if not gereken:
                continue

            d = self.ws.cell(r, 4).value
            isim = self.ws.cell(r, 2).value
            if self._sirket_mi(isim):
                self.log(f"satır {r}: '{isim}' şirket (tüzel kişi) — sorgu yapılmadı, taşınıyor.")
                self._sirketi_isaretle(r, secili)
                continue
            if not d or "/" not in str(d):
                self._yaz_hepsi(r, gereken, offset, secili, f"{HATA_ON} Geçersiz dosya no: {d}")
                hata_sayaci += 1
                continue

            yil, _, sira = str(d).partition("/")
            try:
                dosya_id, _ = motor.dosya_id_bul(yil.strip(), sira.strip())
            except OturumHatasi as e:
                self.log(f"⛔ Oturum hatası: {e}")
                if not oturum_yenilendi and oturum_yenile():
                    continue
                self._oturum_dur()
                self._stop = True
                break
            except Exception as e:
                self._yaz_hepsi(r, gereken, offset, secili, f"{HATA_ON} arama: {e}")
                hata_sayaci += 1
                pespese += 1
                if pespese >= OTURUM_KONTROL_ESIGI and not oturum_yenilendi:
                    self.log(f"⚠️ {OTURUM_KONTROL_ESIGI} ardışık hata — oturum kontrol ediliyor...")
                    oturum_yenile()
                continue
            if not dosya_id:
                self._yaz_hepsi(r, gereken, offset, secili, f"{HATA_ON} Dosya bulunamadı")
                self.log(f"satır {r}: {d} -> dosya bulunamadı")
                hata_sayaci += 1
                pespese += 1
                if pespese >= OTURUM_KONTROL_ESIGI and not oturum_yenilendi:
                    self.log(f"⚠️ {OTURUM_KONTROL_ESIGI} ardışık 'dosya bulunamadı' — "
                             "oturum kopmuş olabilir, yenileniyor...")
                    oturum_yenile()
                continue

            try:
                borclular = motor.borclular(dosya_id)
            except OturumHatasi as e:
                self.log(f"⛔ Oturum hatası: {e}")
                if not oturum_yenilendi and oturum_yenile():
                    continue
                self._oturum_dur()
                self._stop = True
                break
            except Exception as e:
                self._yaz_hepsi(r, gereken, offset, secili, f"{HATA_ON} borçlu: {e}")
                hata_sayaci += 1
                continue
            borclu, eslesti = motor.borclu_sec(borclular, isim)
            if not borclu:
                self._yaz_hepsi(r, gereken, offset, secili, f"{HATA_ON} Borçlu bulunamadı")
                hata_sayaci += 1
                continue

            pespese = 0
            oturum_yenilendi = False
            self.log(f"satır {r}: {d} -> {borclu['adi']} {borclu['soyadi']}"
                     + ("" if eslesti else "  (⚠ isim eşleşmedi, ilk borçlu)"))

            ozetler = motor.sgk_sorgula(dosya_id, borclu["kisiKurumId"], gereken, self._bekle_devam)

            sparse = {}
            for anahtar, (ok, metin) in ozetler.items():
                sparse[offset[anahtar]] = metin
                if not ok:
                    hata_sayaci += 1
            if sparse:
                self._apply_sonuc(r, sparse, secili)

            if self._stop:
                break
            if idx % 10 == 0:
                self.save()
            time.sleep(SATIR_ARASI_BEKLEME)

        self.save()
        if self._stop:
            self._durum("Durduruldu")
            self.log("⏹ Durduruldu.")
        else:
            self._durum("Tamamlandı")
            self.log("🎉 İşlem tamamlandı.")
        motor.kapat()
        self.running = False

    # ─────────────────────────── sonuç yazımı (orijinal _kuyruk_isle eşi) ───────────────────────────
    def _apply_sonuc(self, r, sparse, secili):
        st = self.row_state.get(r)
        for off, metin in sparse.items():
            self.ws.cell(r, SONUC_BASLANGIC_KOL + off, value=metin)
            if st is not None:
                st["results"][off] = _kisa(metin)
        if st is not None:
            st["tag"] = "tamam" if self._satir_tamamlandi_mi(r, secili) else ""
            self._touch(r)

    def _yaz_hepsi(self, r, gereken, offset, secili, metin):
        sparse = {offset[a]: metin for a in gereken}
        self._apply_sonuc(r, sparse, secili)

    def _sirketi_isaretle(self, r, secili):
        st = self.row_state.get(r)
        for i, (_, _, anahtar) in enumerate(SORGULAR):
            if anahtar not in secili:
                continue
            self.ws.cell(r, SONUC_BASLANGIC_KOL + i, value=SIRKET_NOT)
            if st is not None:
                st["results"][i] = _kisa(SIRKET_NOT)
        if st is not None:
            st["tag"] = "sirket"
            self._touch(r)

    def _oturum_dur(self):
        self._durum("⛔ Ofis bağlantısı yok")
        self.log("⛔ Yerel ofis bağlantısına (127.0.0.1:8800) ulaşılamadı ya da yetki reddedildi. "
                 "'UYAP Bağlantısı' modülünden Paylaş/Al ile bağlantıyı başlatıp tekrar deneyin.")

    # ─────────────────────────── kaydetme (orijinal _kaydet / _ozet_yaz) ───────────────────────────
    def save(self):
        if not self.wb or not self.excel_yolu:
            return
        try:
            self.wb.save(self.excel_yolu)
        except PermissionError:
            self.log("⚠️ Yapılanlar dosyası Excel'de açık olabilir, kapatın.")
            return
        except Exception as e:
            self.log(f"⚠️ Kaydetme hatası: {e}")
            return
        self._ozet_yaz()

    @staticmethod
    def _ilk_satir(deger):
        if deger is None:
            return None
        return str(deger).split("\n", 1)[0]

    def _ozet_yaz(self):
        if not self.ozet_yolu or not self.ws:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(list(self.kolonlar))
            for r in self.satirlar:
                satir = [self.ws.cell(r, c).value for c in range(1, 5)]
                satir += [self._ilk_satir(self.ws.cell(r, SONUC_BASLANGIC_KOL + i).value)
                          for i in range(len(SORGULAR))]
                ws.append(satir)
            wb.save(self.ozet_yolu)
        except PermissionError:
            self.log("⚠️ Özet dosyası Excel'de açık olabilir, kapatın.")
        except Exception as e:
            self.log(f"⚠️ Özet yazma hatası: {e}")

    # ─────────────────────────── anlık görüntü (GUI/web polling) ───────────────────────────
    def hucre_tam_deger(self, r, kol_index):
        """Detay için tam (kısaltılmamış) değer."""
        if self.ws is None or kol_index < 0:
            return ""
        if kol_index >= 4:
            v = self.ws.cell(r, SONUC_BASLANGIC_KOL + (kol_index - 4)).value
            return "" if v is None else str(v)
        v = self.ws.cell(r, kol_index + 1).value
        return "" if v is None else str(v)

    def snapshot(self, since_log=0, since_rev=0):
        """Artımlı görüntü: yeni günlük satırları + son güncellemeden bu yana değişen satırlar."""
        with self._lock:
            logs = self.log_lines[since_log:] if since_log < len(self.log_lines) else []
            log_n = len(self.log_lines)
        rows = []
        for r in self.satirlar:
            if self.row_rev.get(r, 0) > since_rev:
                st = self.row_state[r]
                rows.append({"r": r, "values": st["base"] + st["results"], "tag": st["tag"]})
        return {
            "running": self.running,
            "paused": self._paused,
            "status": self.status_text,
            "rev": self.rev,
            "log_n": log_n,
            "logs": logs,
            "rows": rows,
        }
